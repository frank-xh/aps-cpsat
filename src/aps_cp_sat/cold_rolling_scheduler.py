from __future__ import annotations

"""
Compatibility facade for the legacy scheduler entry points.

New production code must use `aps_cp_sat.cold_rolling_pipeline.ColdRollingPipeline`.
This module remains only for backward-compatible entry points and legacy adapter
behavior.
"""

import math
from dataclasses import dataclass
from pathlib import Path
from time import perf_counter
from typing import Dict, List, Set, Tuple

import pandas as pd
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model
from aps_cp_sat.constraints_config import DEFAULT_CONSTRAINT_CONFIG
from aps_cp_sat.config.rule_config import RuleConfig
from aps_cp_sat.io.result_writer import export_schedule_results as _export_schedule_results_v2
from aps_cp_sat.preprocess.grade_catalog import MergedGradeRuleCatalog as PREPROCESS_GRADE_CATALOG
from aps_cp_sat.preprocess.order_preparation import prepare_orders as PREPROCESS_PREPARE_ORDERS
from aps_cp_sat.transition.bridge_rules import (
    DUE_RANK as SHARED_DUE_RANK,
    build_virtual_spec_views as SHARED_BUILD_VIRTUAL_SPEC_VIEWS,
    _bridge_need as SHARED_BRIDGE_NEED,
    _bridge_pair as SHARED_BRIDGE_PAIR,
    _temp_overlap_len as SHARED_TEMP_OVERLAP_LEN,
    _txt as SHARED_TXT,
)


@dataclass(frozen=True)
class SolveConfig:
    max_orders: int = 720
    rounds: int = 4
    time_limit_seconds: float = 20.0
    
    # Validation/experiment fields
    profile_name: str = "constructive_lns_search"
    main_solver_strategy: str | None = None
    
    # 主模型编排参数：控制外层编排复杂度，默认保持轻量。
    master_profile_count: int = 2
    master_seed_count: int = 2
    master_num_workers: int = 4
    allow_fallback: bool = False
    # 仅在显式生产兼容模式下允许回退到 legacy 路径。
    allow_legacy_fallback: bool = False
    # 验证模式：默认禁用回退，要求结果来自联合主模型路径。
    validation_mode: bool = False
    production_compatibility_mode: bool = False
    # 默认使用分层目标，混合单层目标作为非默认调试模式。
    enable_tiered_objective: bool = True
    min_real_schedule_ratio: float = 0.90
    max_virtual_chain: int = 12
    max_unbridgeable_drop_ratio: float = 0.08
    lot_max_tons: float = 120.0
    sparse_k_same_group: int = 10
    sparse_k_same_thickness: int = 8
    sparse_k_cross_group: int = 6
    sparse_k_due_tight: int = 6
    template_min_out_degree: int = 1
    template_min_in_degree: int = 1
    template_top_k: int = 40
    # 联合主模型弧稀疏化强度：<=0 表示不做全局剪枝。
    global_prune_max_pairs_per_from: int = 0
    # 单槽位允许的路段数软上限（超过后高罚），用于降低低吨位碎片轧期。
    max_routes_per_slot: int = 6
    # 虚拟宽度策略
    strict_virtual_width_levels: bool = False
    physical_reverse_step_mode: bool = True


@dataclass(frozen=True)
class SteelSpec:
    steel_group: str
    roll_capability: str
    priority: int


@dataclass(frozen=True)
class PenaltyWeights:
    seq_unique: int = DEFAULT_CONSTRAINT_CONFIG.weights.seq_unique
    seq_contiguous: int = DEFAULT_CONSTRAINT_CONFIG.weights.seq_contiguous
    seq_start: int = DEFAULT_CONSTRAINT_CONFIG.weights.seq_start
    roll_line_mismatch: int = DEFAULT_CONSTRAINT_CONFIG.weights.roll_line_mismatch
    ton_over: int = DEFAULT_CONSTRAINT_CONFIG.weights.ton_over
    ton_under: int = DEFAULT_CONSTRAINT_CONFIG.weights.ton_under
    temp_shortage: int = DEFAULT_CONSTRAINT_CONFIG.weights.temp_shortage
    width_violation: int = DEFAULT_CONSTRAINT_CONFIG.weights.width_violation
    thick_violation: int = DEFAULT_CONSTRAINT_CONFIG.weights.thick_violation
    non_pc_switch: int = DEFAULT_CONSTRAINT_CONFIG.weights.non_pc_switch
    virtual_chain_excess: int = DEFAULT_CONSTRAINT_CONFIG.weights.virtual_chain_excess
    unassigned_real: int = DEFAULT_CONSTRAINT_CONFIG.weights.unassigned_real
    virtual_ratio: int = DEFAULT_CONSTRAINT_CONFIG.weights.virtual_ratio
    pure_virtual_campaign: int = DEFAULT_CONSTRAINT_CONFIG.weights.pure_virtual_campaign
    virtual_use: int = DEFAULT_CONSTRAINT_CONFIG.weights.virtual_use
    width_smooth: int = DEFAULT_CONSTRAINT_CONFIG.weights.width_smooth
    thick_smooth: int = DEFAULT_CONSTRAINT_CONFIG.weights.thick_smooth
    temp_margin: int = DEFAULT_CONSTRAINT_CONFIG.weights.temp_margin
    ton_target: int = DEFAULT_CONSTRAINT_CONFIG.weights.ton_target
    ton_target_value: float = DEFAULT_CONSTRAINT_CONFIG.thresholds.campaign_ton_target


class SteelSpecCatalog:
    SPECS: Dict[str, SteelSpec] = {
        "DC01": SteelSpec("PC", "dual", 1),
        "DC01-D": SteelSpec("PC", "dual", 1),
        "SPCC": SteelSpec("PC", "dual", 1),
        "DC04": SteelSpec("IF", "dual", 2),
        "DC05": SteelSpec("IF", "dual", 2),
        "DC06": SteelSpec("IF", "dual", 2),
        "T1500HS": SteelSpec("HOTFORM", "small_only", 3),
        "T280VK": SteelSpec("DP", "small_only", 3),
        "HC180Y": SteelSpec("BH", "small_only", 3),
        "HC340/590DP": SteelSpec("DP", "small_only", 3),
    }

    @classmethod
    def get(cls, grade: str) -> SteelSpec | None:
        return cls.SPECS.get(_txt(grade).upper())


class MergedGradeRuleCatalog:
    """
    统一钢种规则目录（已固化快照）：
    本字典由“先读文件 + 合并硬编码参数”生成后写入代码，后续直接查字典。
    """

    RULES: Dict[str, dict] = {
        '180Y': {'steel_group': '加磷高强、烘烤硬化钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        '210P1': {'steel_group': '加磷高强、烘烤硬化钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        '280VK': {'steel_group': '双相钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'CR300LA': {'steel_group': '低合金高强', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'CR330Y590T-DP': {'steel_group': '双相钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'CR340/590DP': {'steel_group': '双相钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'CR420LA': {'steel_group': '低合金高强', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'DC01': {'steel_group': '普碳', 'roll_capability': 'dual', 'priority': 1, 'source': 'data+catalog'},
        'DC01-D': {'steel_group': '普碳', 'roll_capability': 'dual', 'priority': 1, 'source': 'data+catalog'},
        'DC03': {'steel_group': 'IF钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'DC03EK': {'steel_group': 'IF钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'DC04': {'steel_group': 'IF钢', 'roll_capability': 'dual', 'priority': 2, 'source': 'data+catalog'},
        'DC05': {'steel_group': 'IF钢', 'roll_capability': 'dual', 'priority': 2, 'source': 'data+catalog'},
        'DC06': {'steel_group': 'IF钢', 'roll_capability': 'dual', 'priority': 2, 'source': 'data+catalog'},
        'HC180B': {'steel_group': '加磷高强、烘烤硬化钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC180Y': {'steel_group': '加磷高强、烘烤硬化钢', 'roll_capability': 'small_only', 'priority': 3, 'source': 'data+catalog'},
        'HC220B': {'steel_group': '加磷高强、烘烤硬化钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC220Y': {'steel_group': '加磷高强、烘烤硬化钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC260/450DP': {'steel_group': '双相钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC260LA': {'steel_group': '低合金高强', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC260Y': {'steel_group': '加磷高强、烘烤硬化钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC300/500DP': {'steel_group': '双相钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC300LA': {'steel_group': '低合金高强', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC340/590DP': {'steel_group': '双相钢', 'roll_capability': 'small_only', 'priority': 3, 'source': 'data+catalog'},
        'HC340LA': {'steel_group': '低合金高强', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC380LA': {'steel_group': '低合金高强', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC420/780DP': {'steel_group': '顶级牌号780以上', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC420LA': {'steel_group': '低合金高强', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC550/980DP': {'steel_group': '顶级牌号780以上', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC600/980QP': {'steel_group': '顶级牌号780以上', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC700/980DP': {'steel_group': '顶级牌号780以上', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'HC820/1180DP': {'steel_group': '顶级牌号780以上', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'Q235': {'steel_group': '低合金1', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'Q235A': {'steel_group': '低合金高强', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'Q235B': {'steel_group': '低合金1', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'SPA-C': {'steel_group': '结构钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'SPCC': {'steel_group': '普碳', 'roll_capability': 'dual', 'priority': 1, 'source': 'data+catalog'},
        'SPCC-SD': {'steel_group': '普碳', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'ST13': {'steel_group': 'IF钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'ST14': {'steel_group': 'IF钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'ST16': {'steel_group': 'IF钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'ST37-2G': {'steel_group': '低合金高强', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'T1500HS': {'steel_group': '热冲压钢', 'roll_capability': 'small_only', 'priority': 3, 'source': 'data+catalog'},
        'T170P1': {'steel_group': '加磷高强、烘烤硬化钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'T210P1': {'steel_group': '加磷高强、烘烤硬化钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'T250P1': {'steel_group': '加磷高强、烘烤硬化钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'T280VK': {'steel_group': '双相钢', 'roll_capability': 'small_only', 'priority': 3, 'source': 'data+catalog'},
        'THS2-TY': {'steel_group': '普碳', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'THS3-TY': {'steel_group': '普碳', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'TLA': {'steel_group': '普碳', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'TTC1': {'steel_group': 'IF钢', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'TYH': {'steel_group': '普碳', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'TZT1': {'steel_group': '普碳', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
        'TZT2': {'steel_group': '普碳', 'roll_capability': 'dual', 'priority': 1, 'source': 'data'},
    }

    def __init__(self, rules: Dict[str, dict] | None = None):
        self.rules = rules or dict(self.RULES)

    @classmethod
    def build(cls, steel_info_path: Path) -> "MergedGradeRuleCatalog":
        # 已固化为代码内独立字典；参数仅保留接口兼容
        _ = steel_info_path
        return cls(dict(cls.RULES))

    def get(self, grade: str) -> dict:
        g = _txt(grade).upper()
        return self.rules.get(g, {"steel_group": "UNKNOWN", "roll_capability": "dual", "priority": 1, "source": "default"})


V_WIDTHS = [1000.0, 1250.0, 1500.0]
V_THICKS = [round(0.5 + 0.1 * i, 1) for i in range(16)]
# 虚拟板坯温度规格：600~950，按 20℃ 切片
V_TEMP_BANDS = [(float(s), float(s + 20)) for s in range(600, 951, 20) if s + 20 <= 950]
DUE_RANK = {"overdue": 0, "urgent": 1, "normal": 2, "slack": 3}
TH = DEFAULT_CONSTRAINT_CONFIG.thresholds


def _txt(v, default: str = "") -> str:
    if v is None:
        return default
    if isinstance(v, float) and pd.isna(v):
        return default
    return str(v).strip()


def _col_by_name_or_index(df: pd.DataFrame, names: List[str], idx: int) -> str:
    cols = [str(c) for c in df.columns]
    for n in names:
        for c in cols:
            if n.lower() in c.lower():
                return c
    if idx < len(cols):
        return cols[idx]
    raise ValueError(f"Cannot resolve column: {names} / {idx}")


def _col_exact_or_fallback(df: pd.DataFrame, exact_names: List[str], fuzzy_names: List[str], idx: int) -> str:
    """优先按精确列名匹配，找不到再走模糊匹配/索引兜底。"""
    cols = [str(c) for c in df.columns]
    for n in exact_names:
        if n in cols:
            return n
    return _col_by_name_or_index(df, fuzzy_names, idx)


def _normalize_order_id(v) -> str:
    """规范订单号文本，避免出现 60.0 这类展示。"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if float(v).is_integer():
            return str(int(v))
        return str(v)
    s = str(v).strip()
    # 仅对纯数字小数格式做去尾 .0；像 0020054109-000010 这类合同号保持原样
    if "." in s:
        left, right = s.split(".", 1)
        if left.isdigit() and right.isdigit() and set(right) <= {"0"}:
            return left
    return s


def _roll_capability(raw: str) -> str:
    t = _txt(raw).lower()
    has_big = ("大" in t) or ("big" in t)
    has_small = ("小" in t) or ("small" in t)
    if has_big and has_small:
        return "dual"
    if has_big:
        return "big_only"
    if has_small:
        return "small_only"
    return "dual"


def _merge_roll_capability(source_cap: str, spec_cap: str) -> str:
    if source_cap in {"big_only", "small_only", "dual"}:
        return source_cap
    return spec_cap if spec_cap in {"big_only", "small_only", "dual"} else "dual"


def _is_pc(group: str) -> bool:
    return _txt(group).upper() in {"PC", "普碳", "VIRTUAL_PC"}


def _thick_ok(prev_t: float, cur_t: float) -> bool:
    diff = abs(cur_t - prev_t)
    if prev_t > 0.8:
        return (diff / prev_t) <= 0.30
    if prev_t >= 0.6:
        return diff <= 0.2
    return diff <= 0.1


def _temp_overlap_len(a_min: float, a_max: float, b_min: float, b_max: float) -> float:
    return max(0.0, min(float(a_max), float(b_max)) - max(float(a_min), float(b_min)))


def _required_temp_overlap(a: dict, b: dict) -> float:
    # 实物-实物：必须重叠 >=10℃
    # 任何一侧是虚拟板坯：允许按切片边界衔接（>=0）
    if bool(a.get("is_virtual", False)) or bool(b.get("is_virtual", False)):
        return 0.0
    return float(TH.min_temp_overlap_real_real)


def _temp_transition_ok(a: dict, b: dict) -> bool:
    overlap = _temp_overlap_len(a.get("temp_min", 0.0), a.get("temp_max", 0.0), b.get("temp_min", 0.0), b.get("temp_max", 0.0))
    return overlap >= _required_temp_overlap(a, b)


def _temp_center(a: dict) -> float:
    lo = float(a.get("temp_min", 0.0))
    hi = float(a.get("temp_max", 0.0))
    return (lo + hi) / 2.0


def _nearest_temp_band(center: float) -> Tuple[float, float]:
    return min(V_TEMP_BANDS, key=lambda x: abs(((x[0] + x[1]) / 2.0) - center))


def _hard_direct_step_ok(a: dict, b: dict) -> bool:
    """检查两卷是否可直接相邻（硬约束）。"""
    if b["width"] > a["width"]:
        return False
    if (a["width"] - b["width"]) > float(TH.max_width_drop):
        return False
    if not _thick_ok(a["thickness"], b["thickness"]):
        return False
    if not _temp_transition_ok(a, b):
        return False
    ga = _txt(a.get("steel_group", "")).upper()
    gb = _txt(b.get("steel_group", "")).upper()
    if ga != gb and (not _is_pc(ga)) and (not _is_pc(gb)):
        return False
    return True


def _width_reverse_virtual_need(
    a_width: float,
    b_width: float,
    strict_virtual_width_levels: bool = False,
) -> int:
    """
    逻辑逆宽（b>a）所需最少“虚拟块数”：
    物理逆宽每次最多 step，步次数=ceil(delta/step)，
    中间插入块数=步次数-1。
    """
    delta = float(b_width) - float(a_width)
    if delta <= 0.0:
        return 0
    step = max(1.0, float(TH.max_width_rise_physical_step))
    logical_max = float(getattr(TH, "virtual_reverse_attach_max_mm", 50.0))
    if delta > logical_max:
        return 10**9
    if strict_virtual_width_levels:
        # 严格离散宽度模式：逆宽只能使用 V_WIDTHS，步次数按离散宽度跳变估算。
        w0 = _nearest(V_WIDTHS, float(a_width))
        w1 = _nearest(V_WIDTHS, float(b_width))
        if w1 <= w0:
            return 0
        physical_steps = int(math.ceil((w1 - w0) / step))
    else:
        physical_steps = int(math.ceil(delta / step))
    # 逆宽必须通过“普材/虚拟”过渡，至少 1 块过渡板坯。
    return max(1, physical_steps - 1)


def _bridge_need(
    a: dict,
    b: dict,
    max_virtual_chain: int,
    strict_virtual_width_levels: bool = False,
) -> int:
    """
    估算从 a 到 b 至少需要多少个虚拟卷。
    返回值 > max_virtual_chain 代表该相邻在当前配置下不可行。
    """
    max_virtual_chain = int(max_virtual_chain)
    if _hard_direct_step_ok(a, b):
        return 0
    # 逆宽规则：允许逻辑逆宽，但必须通过桥接实现，且每次物理逆宽<=20mm，逻辑逆宽总量<=200mm。
    if float(b["width"]) > float(a["width"]):
        need_reverse = _width_reverse_virtual_need(
            float(a["width"]),
            float(b["width"]),
            strict_virtual_width_levels=bool(strict_virtual_width_levels),
        )
        if need_reverse >= 10**9:
            return max_virtual_chain + 1
    else:
        need_reverse = 0
    # 前置离散规格可行性过滤：
    # 若需要虚拟桥接，但目标宽度高于虚拟最大宽度，则无法通过“虚拟->实物”收敛；
    # 若起点宽度超过虚拟最大宽度太多（>max_width_drop），则无法通过“实物->虚拟”起桥。
    v_w_max = float(max(V_WIDTHS)) if V_WIDTHS else 0.0
    if float(b["width"]) > v_w_max:
        return max_virtual_chain + 1
    if float(a["width"]) - v_w_max > float(TH.max_width_drop):
        return max_virtual_chain + 1
    need_w = max(
        0,
        math.ceil(
            max(0.0, a["width"] - b["width"] - float(TH.max_width_drop))
            / max(1.0, float(TH.max_width_drop))
        ),
    )
    need_group = 1 if (_txt(a["steel_group"]).upper() != _txt(b["steel_group"]).upper() and (not _is_pc(a["steel_group"])) and (not _is_pc(b["steel_group"]))) else 0
    if _thick_ok(a["thickness"], b["thickness"]):
        need_t = 0
    else:
        need_t = max_virtual_chain + 1
        for n in range(1, max_virtual_chain + 1):
            prev = a["thickness"]
            ok = True
            for k in range(1, n + 1):
                t = a["thickness"] + (b["thickness"] - a["thickness"]) * k / (n + 1)
                t = min(2.0, max(0.5, round(t, 1)))
                if not _thick_ok(prev, t):
                    ok = False
                    break
                prev = t
            if ok and _thick_ok(prev, b["thickness"]):
                need_t = n
                break
    # 温度硬约束（区间重叠>=10）桥接块估算
    need_temp = 0
    if not _temp_transition_ok(a, b):
        need_temp = max_virtual_chain + 1
        ca = _temp_center(a)
        cb = _temp_center(b)
        for n in range(1, max_virtual_chain + 1):
            prev_min = float(a.get("temp_min", 0.0))
            prev_max = float(a.get("temp_max", 0.0))
            ok = True
            for k in range(1, n + 1):
                c = ca + (cb - ca) * k / (n + 1)
                tmin, tmax = _nearest_temp_band(c)
                # ?????????????????????????
                if _temp_overlap_len(prev_min, prev_max, tmin, tmax) < 0.0:
                    ok = False
                    break
                prev_min, prev_max = tmin, tmax
            # ???????? -> ????????????????
            if ok and _temp_overlap_len(prev_min, prev_max, float(b.get("temp_min", 0.0)), float(b.get("temp_max", 0.0))) >= 0.0:
                need_temp = n
                break
    need = max(need_w, need_group, need_t, need_temp, int(need_reverse))
    return need if need <= max_virtual_chain else max_virtual_chain + 1


def _nearest(vals: List[float], x: float) -> float:
    return min(vals, key=lambda v: abs(v - x))


def _nearest_n(vals: List[float], x: float, n: int) -> List[float]:
    return sorted(vals, key=lambda v: abs(v - x))[: max(1, int(n))]


def _build_grade_lookup(steel_info_path: Path) -> Dict[str, str]:
    if not steel_info_path.exists():
        return {}
    grade_to_group: Dict[str, str] = {}
    xls = pd.ExcelFile(steel_info_path)
    for sheet in xls.sheet_names:
        df = pd.read_excel(steel_info_path, sheet_name=sheet)
        if df.empty:
            continue
        cols = [str(c) for c in df.columns]
        grade_col = None
        group_col = None
        for c in cols:
            lc = c.lower()
            if grade_col is None and ("牌号" in c or "grade" in lc):
                grade_col = c
            if group_col is None and (("钢" in c and "种" in c) or "group" in lc):
                group_col = c
        if grade_col is None or group_col is None:
            continue
        for _, row in df.iterrows():
            g = _txt(row.get(grade_col, "")).upper()
            sg = _txt(row.get(group_col, "")).upper()
            if g and sg and sg not in {"NAN", "(空白)"}:
                grade_to_group.setdefault(g, sg)
    return grade_to_group


def _due_bucket(v) -> str:
    d = pd.to_datetime(v, errors="coerce")
    if pd.isna(d):
        return "slack"
    delta = (d.normalize() - pd.Timestamp.now().normalize()).days
    if delta < 0:
        return "overdue"
    if delta <= 3:
        return "urgent"
    if delta <= 14:
        return "normal"
    return "slack"


def _proc_hours(tons: float, thickness: float, line: str) -> float:
    base_tph = 95.0 if line == "small_roll" else 120.0
    if thickness < 0.7:
        base_tph *= 0.82
    elif thickness < 1.0:
        base_tph *= 0.90
    return float(tons / max(30.0, base_tph))


def _prepare_orders(
    orders_path: Path,
    steel_info_path: Path,
    cfg: SolveConfig,
    grade_catalog: MergedGradeRuleCatalog | None = None,
) -> pd.DataFrame:
    """
    读取原始订单并标准化为排程输入：
    1) 字段映射（钢种组/宽厚/吨位/辊型）
    2) planning lot 拆分（大单拆批）
    3) 生成 due bucket 与线体能力字段
    """
    df = pd.read_excel(orders_path, sheet_name=0).copy()
    grade_catalog = grade_catalog or MergedGradeRuleCatalog.build(steel_info_path)

    col_id = _col_exact_or_fallback(df, ["合同完全号", "物料号", "订单号"], ["合同", "物料", "订单号"], 1)
    col_due = _col_by_name_or_index(df, ["交货", "due"], 4)
    col_grade = _col_by_name_or_index(df, ["牌号", "grade"], 5)
    col_backlog = _col_by_name_or_index(df, ["欠交", "backlog"], 9)
    col_group = _col_by_name_or_index(df, ["Steel Grade Class", "钢种", "group"], 11)
    # 宽厚必须取成品字段，避免误用“热轧宽度/热轧厚度”。
    col_width = _col_exact_or_fallback(df, ["宽度"], ["宽度", "width"], 15)
    col_thick = _col_exact_or_fallback(df, ["厚度"], ["厚度", "thickness"], 16)
    col_roll = _col_by_name_or_index(df, ["辊", "roll"], 30)
    col_temp_max = _col_exact_or_fallback(df, ["均热段温度最大值"], ["温度最大", "temp max"], 28)
    col_temp_min = _col_exact_or_fallback(df, ["均热段温度最小值"], ["温度最小", "temp min"], 29)
    col_tons = df.columns[33] if len(df.columns) > 33 else _col_by_name_or_index(df, ["吨", "重量", "weight"], 33)

    base = pd.DataFrame(
        {
            "source_order_id": df[col_id].map(_normalize_order_id),
            "grade": df[col_grade].astype(str),
            "steel_group_raw": df[col_group].astype(str),
            "width": pd.to_numeric(df[col_width], errors="coerce"),
            "thickness": pd.to_numeric(df[col_thick], errors="coerce"),
            "temp_min": pd.to_numeric(df[col_temp_min], errors="coerce"),
            "temp_max": pd.to_numeric(df[col_temp_max], errors="coerce"),
            "tons": pd.to_numeric(df[col_tons], errors="coerce"),
            "backlog": pd.to_numeric(df[col_backlog], errors="coerce"),
            "due_date": pd.to_datetime(df[col_due], errors="coerce"),
            "roll_type": df[col_roll].astype(str),
        }
    )
    base = base.dropna(subset=["width", "thickness", "temp_min", "temp_max", "tons", "backlog"]).copy()
    base["temp_min"], base["temp_max"] = (
        base[["temp_min", "temp_max"]].min(axis=1),
        base[["temp_min", "temp_max"]].max(axis=1),
    )
    base = base[base["backlog"] > 0].copy()

    rows: List[dict] = []
    for ridx, row in base.iterrows():
        src_id = _normalize_order_id(row["source_order_id"])
        grade = _txt(row["grade"]).upper()
        source_cap = _roll_capability(_txt(row["roll_type"]))
        rule = grade_catalog.get(grade)
        # 后续规则统一从独立字典读取
        steel_group = _txt(rule.get("steel_group", "UNKNOWN")).upper()
        line_capability = _merge_roll_capability(source_cap, _txt(rule.get("roll_capability", "dual")))
        priority = int(rule.get("priority", 1))

        # 当前业务口径：一个订单视作一个物料，不做拆 lot。
        rows.append(
            {
                "order_id": src_id,  # 物料号
                "source_order_id": src_id,  # 来源订单号
                "parent_order_id": src_id,  # 父订单号
                "lot_id": src_id,  # 规划批号（与订单一致）
                "grade": grade,
                "steel_group": steel_group,
                "steel_group_raw": _txt(row.get("steel_group_raw", "")),
                "width": float(row["width"]),
                "thickness": float(row["thickness"]),
                "temp_min": float(row["temp_min"]),
                "temp_max": float(row["temp_max"]),
                "temp_mean": (float(row["temp_min"]) + float(row["temp_max"])) / 2.0,
                "tons": float(row["tons"]),
                "backlog": float(row["backlog"]),
                "due_date": row["due_date"],
                "roll_type": _txt(row["roll_type"]),
                "line_capability": line_capability,
                "priority": priority,
            }
        )

    out = pd.DataFrame(rows)
    out["due_bucket"] = out["due_date"].map(_due_bucket)
    out["due_rank"] = out["due_bucket"].map(DUE_RANK).fillna(3).astype(int)
    out["line"] = out["line_capability"].map({"big_only": "big_roll", "small_only": "small_roll", "dual": ""}).fillna("")
    out["line_source"] = out["line"].map({"big_roll": "big_only", "small_roll": "small_only", "": "dual"}).fillna("dual")
    out = out.sort_values(["priority", "due_rank", "backlog", "tons"], ascending=[False, True, False, False]).head(cfg.max_orders).reset_index(drop=True)
    out["proc_hours_big"] = out.apply(lambda r: _proc_hours(float(r["tons"]), float(r["thickness"]), "big_roll"), axis=1)
    out["proc_hours_small"] = out.apply(lambda r: _proc_hours(float(r["tons"]), float(r["thickness"]), "small_roll"), axis=1)
    out["proc_hours"] = 0.0
    return out

def _assign_dual_master(data: pd.DataFrame, time_limit: float = 8.0) -> pd.DataFrame:
    """
    dual 订单分配主模型（小规模 CP-SAT）：
    以工时平衡为主，同时兼顾交期桶、宽度中心和钢组混杂度。
    """
    out = data.copy()
    fixed_big = out[out["line"] == "big_roll"].copy()
    fixed_small = out[out["line"] == "small_roll"].copy()
    dual = out[out["line"] == ""].copy()
    if dual.empty:
        out["proc_hours"] = out.apply(lambda r: r["proc_hours_big"] if r["line"] == "big_roll" else r["proc_hours_small"], axis=1)
        return out

    model = cp_model.CpModel()
    dual_idx = dual.index.tolist()
    x: Dict[Tuple[int, str], cp_model.IntVar] = {}
    for i in dual_idx:
        x[(i, "big_roll")] = model.NewBoolVar(f"x_{i}_big")
        x[(i, "small_roll")] = model.NewBoolVar(f"x_{i}_small")
        model.Add(x[(i, "big_roll")] + x[(i, "small_roll")] == 1)

    h_big_fixed = float(fixed_big["proc_hours_big"].sum())
    h_small_fixed = float(fixed_small["proc_hours_small"].sum())
    big_terms = [int(round(float(out.at[i, "proc_hours_big"]) * 1000)) * x[(i, "big_roll")] for i in dual_idx]
    small_terms = [int(round(float(out.at[i, "proc_hours_small"]) * 1000)) * x[(i, "small_roll")] for i in dual_idx]
    big_total = model.NewIntVar(0, 10**9, "big_total")
    small_total = model.NewIntVar(0, 10**9, "small_total")
    model.Add(big_total == int(round(h_big_fixed * 1000)) + sum(big_terms))
    model.Add(small_total == int(round(h_small_fixed * 1000)) + sum(small_terms))
    hours_diff = model.NewIntVar(0, 10**9, "hours_diff")
    model.AddAbsEquality(hours_diff, big_total - small_total)

    due_diff_terms = []
    for b in ["overdue", "urgent", "normal"]:
        idxs = [i for i in dual_idx if _txt(out.at[i, "due_bucket"]) == b]
        if not idxs:
            continue
        vb = model.NewIntVar(0, len(idxs), f"due_{b}_big")
        vs = model.NewIntVar(0, len(idxs), f"due_{b}_small")
        model.Add(vb == sum(x[(i, "big_roll")] for i in idxs))
        model.Add(vs == sum(x[(i, "small_roll")] for i in idxs))
        d = model.NewIntVar(0, len(idxs), f"due_diff_{b}")
        model.AddAbsEquality(d, vb - vs)
        due_diff_terms.append(d)

    all_width_mean = float(out["width"].mean())
    big_anchor = float(fixed_big["width"].mean()) if not fixed_big.empty else all_width_mean
    small_anchor = float(fixed_small["width"].mean()) if not fixed_small.empty else all_width_mean
    width_dev_terms = []
    for i in dual_idx:
        wb = int(round(abs(float(out.at[i, "width"]) - big_anchor)))
        ws = int(round(abs(float(out.at[i, "width"]) - small_anchor)))
        width_dev_terms.append(wb * x[(i, "big_roll")])
        width_dev_terms.append(ws * x[(i, "small_roll")])

    groups = sorted(set(_txt(out.at[i, "steel_group"]) for i in dual_idx))
    overlap_terms = []
    for g in groups:
        idxs = [i for i in dual_idx if _txt(out.at[i, "steel_group"]) == g]
        if not idxs:
            continue
        yb = model.NewBoolVar(f"g_{g}_big")
        ys = model.NewBoolVar(f"g_{g}_small")
        for i in idxs:
            model.Add(x[(i, "big_roll")] <= yb)
            model.Add(x[(i, "small_roll")] <= ys)
        model.Add(sum(x[(i, "big_roll")] for i in idxs) >= yb)
        model.Add(sum(x[(i, "small_roll")] for i in idxs) >= ys)
        ov = model.NewBoolVar(f"ov_{g}")
        model.Add(ov <= yb)
        model.Add(ov <= ys)
        model.Add(ov >= yb + ys - 1)
        overlap_terms.append(ov)

    model.Minimize(1500 * hours_diff + 800 * sum(due_diff_terms) + 6 * sum(width_dev_terms) + 1000 * sum(overlap_terms))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = max(2.0, min(20.0, float(time_limit)))
    solver.parameters.random_seed = 42
    solver.parameters.num_search_workers = 1
    status = solver.Solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        big_h = h_big_fixed
        small_h = h_small_fixed
        for i in dual.sort_values(["due_rank", "priority", "tons"], ascending=[True, False, False]).index:
            hb = big_h + float(out.at[i, "proc_hours_big"])
            hs = small_h + float(out.at[i, "proc_hours_small"])
            if hb <= hs:
                out.at[i, "line"] = "big_roll"
                big_h = hb
            else:
                out.at[i, "line"] = "small_roll"
                small_h = hs
    else:
        for i in dual_idx:
            out.at[i, "line"] = "big_roll" if solver.Value(x[(i, "big_roll")]) else "small_roll"

    out["proc_hours"] = out.apply(lambda r: r["proc_hours_big"] if r["line"] == "big_roll" else r["proc_hours_small"], axis=1)
    return out


def _heuristic_order(df: pd.DataFrame) -> List[int]:
    h = (
        df.assign(_idx=range(1, len(df) + 1))
        .sort_values(["width", "thickness", "due_rank", "priority", "tons"], ascending=[False, False, True, False, False])
    )
    return h["_idx"].tolist()


def _component_fallback_order(df: pd.DataFrame, cfg: SolveConfig) -> List[int]:
    """
    当 CP-SAT 不可行时的回退排序：
    先按“可桥接连通块”分组，再在组内按宽厚优先排序，尽量减少硬冲突。
    """
    n = len(df)
    if n <= 1:
        return [1] if n == 1 else []
    recs = df.to_dict("records")
    neigh: Dict[int, set] = {i: set() for i in range(n)}
    for i in range(n):
        ai = recs[i]
        for j in range(i + 1, n):
            bj = recs[j]
            ij = _bridge_need(ai, bj, cfg.max_virtual_chain, strict_virtual_width_levels=cfg.strict_virtual_width_levels) <= cfg.max_virtual_chain
            ji = _bridge_need(bj, ai, cfg.max_virtual_chain, strict_virtual_width_levels=cfg.strict_virtual_width_levels) <= cfg.max_virtual_chain
            if ij or ji:
                neigh[i].add(j)
                neigh[j].add(i)

    comps: List[List[int]] = []
    seen = set()
    for i in range(n):
        if i in seen:
            continue
        stack = [i]
        seen.add(i)
        comp = []
        while stack:
            u = stack.pop()
            comp.append(u)
            for v in neigh[u]:
                if v not in seen:
                    seen.add(v)
                    stack.append(v)
        comps.append(comp)

    comps = sorted(
        comps,
        key=lambda c: (
            -max(float(recs[i]["width"]) for i in c),
            min(int(recs[i]["due_rank"]) for i in c),
            -len(c),
        ),
    )
    seq: List[int] = []
    for comp in comps:
        comp_sorted = sorted(
            comp,
            key=lambda i: (
                -float(recs[i]["width"]),
                -float(recs[i]["thickness"]),
                int(recs[i]["due_rank"]),
                -int(recs[i]["priority"]),
            ),
        )
        seq.extend([i + 1 for i in comp_sorted])
    return seq


def _build_sparse_arcs(df: pd.DataFrame, cfg: SolveConfig) -> Tuple[Dict[Tuple[int, int], int], Dict[Tuple[int, int], int]]:
    """
    构建稀疏候选弧：
    - 不可能相邻的边直接过滤
    - 可桥接边保留并附加惩罚
    - 通过多策略选边保证每个点有基本连通性
    """
    recs = df.to_dict("records")
    n = len(recs)
    arc_cost: Dict[Tuple[int, int], int] = {}
    arc_bridge: Dict[Tuple[int, int], int] = {}

    def _edge_cost(a: dict, b: dict, need: int) -> int:
        wdiff = abs(float(a["width"]) - float(b["width"]))
        tdiff = abs(float(a["thickness"]) - float(b["thickness"]))
        tempdiff = abs(float(a.get("temp_mean", 0.0)) - float(b.get("temp_mean", 0.0)))
        group_switch = 1 if _txt(a["steel_group"]).upper() != _txt(b["steel_group"]).upper() else 0
        due_inv = max(0, int(a["due_rank"]) - int(b["due_rank"]))
        return int(2 * wdiff + 180 * tdiff + 30 * tempdiff + 70 * group_switch + 300 * due_inv + 1200 * need)

    for i in range(n):
        a = recs[i]
        candidates: List[Tuple[int, int, int]] = []
        for j in range(n):
            if i == j:
                continue
            b = recs[j]
            need = _bridge_need(a, b, cfg.max_virtual_chain, strict_virtual_width_levels=cfg.strict_virtual_width_levels)
            if need > cfg.max_virtual_chain:
                continue
            candidates.append((j, need, _edge_cost(a, b, need)))

        if not candidates:
            continue

        same_group = [c for c in candidates if _txt(recs[c[0]]["steel_group"]).upper() == _txt(a["steel_group"]).upper()]
        same_group = sorted(same_group, key=lambda x: (abs(float(recs[x[0]]["width"]) - float(a["width"])), x[2]))[: cfg.sparse_k_same_group]
        same_thick = sorted(candidates, key=lambda x: (abs(float(recs[x[0]]["thickness"]) - float(a["thickness"])), x[2]))[: cfg.sparse_k_same_thickness]
        cross_bridge = [c for c in candidates if c[1] > 0]
        cross_bridge = sorted(cross_bridge, key=lambda x: (x[1], x[2]))[: cfg.sparse_k_cross_group]
        due_tight = sorted(candidates, key=lambda x: (int(recs[x[0]]["due_rank"]), x[2]))[: cfg.sparse_k_due_tight]

        kept = {(j, need, score) for (j, need, score) in (same_group + same_thick + cross_bridge + due_tight)}
        if len(kept) < 3:
            for c in sorted(candidates, key=lambda x: x[2])[:3]:
                kept.add(c)

        for j, need, score in kept:
            arc_cost[(i + 1, j + 1)] = score
            arc_bridge[(i + 1, j + 1)] = need

    for i in range(1, n + 1):
        rec = recs[i - 1]
        start_cost = int(500 * int(rec["due_rank"]) + 0.5 * max(0.0, 1800.0 - float(rec["width"])))
        arc_cost[(0, i)] = start_cost
        arc_bridge[(0, i)] = 0
        arc_cost[(i, 0)] = 0
        arc_bridge[(i, 0)] = 0

    return arc_cost, arc_bridge


def _solve_line_once_cp_sat(df: pd.DataFrame, cfg: SolveConfig, seed: int) -> Tuple[List[int], Dict[Tuple[int, int], int], Dict[str, float]]:
    """
    单线一次求解：
    1) 稀疏图 CP-SAT + hint
    2) 若不可行，回退到更密集弧图再次求解
    """
    def _solve_with_arcs(arc_cost_map: Dict[Tuple[int, int], int], hint_seq: List[int], time_limit: float) -> Tuple[int, cp_model.CpSolver, Dict[Tuple[int, int], cp_model.IntVar]]:
        # 通用子过程：给定弧集合后建立 circuit 并求解。
        model = cp_model.CpModel()
        vars_map: Dict[Tuple[int, int], cp_model.IntVar] = {}
        arcs = []
        terms = []
        for (i, j), c in arc_cost_map.items():
            v = model.NewBoolVar(f"a_{i}_{j}")
            vars_map[(i, j)] = v
            arcs.append((i, j, v))
            terms.append(int(c) * v)
        model.AddCircuit(arcs)
        model.Minimize(sum(terms))
        hint_edges = []
        if hint_seq:
            hint_edges.append((0, hint_seq[0]))
            hint_edges.extend((hint_seq[k], hint_seq[k + 1]) for k in range(len(hint_seq) - 1))
            hint_edges.append((hint_seq[-1], 0))
        hint_set = set(hint_edges)
        for k, var in vars_map.items():
            model.AddHint(var, 1 if k in hint_set else 0)
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = max(1.0, float(time_limit))
        solver.parameters.random_seed = seed
        solver.parameters.randomize_search = True
        solver.parameters.num_search_workers = 1
        solver.parameters.use_optimization_hints = True
        status = solver.Solve(model)
        return status, solver, vars_map

    t_model = perf_counter()
    arc_cost, arc_bridge = _build_sparse_arcs(df, cfg)
    hint_seq = _heuristic_order(df)
    model_seconds = perf_counter() - t_model

    t_solve = perf_counter()
    status, solver, vars_map = _solve_with_arcs(arc_cost, hint_seq, cfg.time_limit_seconds)
    solve_seconds = perf_counter() - t_solve

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        # Fallback to dense feasible arcs while keeping hard-edge filtering.
        recs = df.to_dict("records")
        n = len(recs)
        dense_cost: Dict[Tuple[int, int], int] = {}
        dense_bridge: Dict[Tuple[int, int], int] = {}
        for i in range(n):
            a = recs[i]
            for j in range(n):
                if i == j:
                    continue
                b = recs[j]
                need = _bridge_need(a, b, cfg.max_virtual_chain, strict_virtual_width_levels=cfg.strict_virtual_width_levels)
                if need > cfg.max_virtual_chain:
                    continue
                c = int(
                    2 * abs(float(a["width"]) - float(b["width"]))
                    + 180 * abs(float(a["thickness"]) - float(b["thickness"]))
                    + 30 * abs(float(a.get("temp_mean", 0.0)) - float(b.get("temp_mean", 0.0)))
                    + 70 * (_txt(a["steel_group"]).upper() != _txt(b["steel_group"]).upper())
                    + 300 * max(0, int(a["due_rank"]) - int(b["due_rank"]))
                    + 1200 * need
                )
                dense_cost[(i + 1, j + 1)] = c
                dense_bridge[(i + 1, j + 1)] = need
        for i in range(1, n + 1):
            dense_cost[(0, i)] = 0
            dense_bridge[(0, i)] = 0
            dense_cost[(i, 0)] = 0
            dense_bridge[(i, 0)] = 0

        t_solve2 = perf_counter()
        status, solver, vars_map = _solve_with_arcs(dense_cost, hint_seq, cfg.time_limit_seconds * 0.6)
        solve_seconds += perf_counter() - t_solve2
        if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            raise RuntimeError("CP-SAT 未找到可行解")
        arc_bridge = dense_bridge

    succ: Dict[int, int] = {}
    for (i, j), var in vars_map.items():
        if solver.Value(var) == 1:
            succ[i] = j

    seq = []
    cur = succ.get(0, 0)
    seen = set()
    while cur != 0 and cur not in seen:
        seq.append(cur)
        seen.add(cur)
        cur = succ.get(cur, 0)
    if len(seq) != len(df):
        seq.extend(sorted(set(range(1, len(df) + 1)) - set(seq)))
    return seq, arc_bridge, {"model_seconds": model_seconds, "solve_seconds": solve_seconds}


def _make_virtual(
    idx: int,
    line: str,
    width: float,
    thickness: float,
    temp_min: float,
    temp_max: float,
    campaign_id: int,
) -> dict:
    return {
        "order_id": f"VIRTUAL-{idx:05d}",
        "source_order_id": "",
        "parent_order_id": "",
        "lot_id": "",
        "grade": "VIRTUAL_PC",
        "steel_group": "PC",
        "width": float(width),
        "thickness": float(thickness),
        "temp_min": float(temp_min),
        "temp_max": float(temp_max),
        "temp_mean": (float(temp_min) + float(temp_max)) / 2.0,
        "tons": 20.0,
        "backlog": 0.0,
        "line": line,
        "line_source": "virtual",
        "due_date": pd.NaT,
        "due_bucket": "slack",
        "due_rank": DUE_RANK["slack"],
        "proc_hours": 0.0,
        "priority": 0,
        "campaign_id": int(campaign_id),
        "is_virtual": True,
    }


def _bridge_pair(
    a: dict,
    b: dict,
    max_virtual_chain: int,
    vidx_start: int,
    campaign_id: int,
    strict_virtual_width_levels: bool = False,
    physical_reverse_step_mode: bool = True,
) -> Tuple[List[dict], int]:
    max_virtual_chain = int(max_virtual_chain)
    need = _bridge_need(
        a,
        b,
        max_virtual_chain,
        strict_virtual_width_levels=bool(strict_virtual_width_levels),
    )
    if need == 0:
        return [], vidx_start
    if need > max_virtual_chain:
        raise RuntimeError("bridge unavailable")

    # 逆宽专用：按“每次物理逆宽<=20mm”构造阶梯桥接。
    if bool(physical_reverse_step_mode) and float(b["width"]) > float(a["width"]):
        delta = float(b["width"]) - float(a["width"])
        if delta > float(getattr(TH, "virtual_reverse_attach_max_mm", 50.0)):
            raise RuntimeError("reverse-width logical delta exceeds hard limit")
        physical_step = max(1.0, float(TH.max_width_rise_physical_step))
        physical_steps = int(math.ceil(delta / physical_step))
        n = max(0, physical_steps - 1)
        if n > max_virtual_chain:
            raise RuntimeError("reverse-width bridge exceeds max virtual chain")
        chain: List[dict] = []
        prev = a
        for k in range(n):
            frac = (k + 1) / (n + 1)
            if bool(strict_virtual_width_levels):
                raw_w = float(a["width"]) + delta * frac
                w = _nearest(V_WIDTHS, raw_w)
            else:
                w = float(a["width"]) + delta * frac
                w = round(w / physical_step) * physical_step
                w = min(float(b["width"]), max(float(a["width"]), w))
            raw_t = float(a["thickness"] + (b["thickness"] - a["thickness"]) * frac)
            t = _nearest(V_THICKS, min(2.0, max(0.5, raw_t)))
            raw_temp = float(_temp_center(a) + (_temp_center(b) - _temp_center(a)) * frac)
            tb_min, tb_max = _nearest_temp_band(raw_temp)
            v = _make_virtual(vidx_start + k, _txt(a["line"]), float(w), float(t), float(tb_min), float(tb_max), campaign_id)
            if not _hard_direct_step_ok(prev, v):
                raise RuntimeError("reverse-width virtual step invalid")
            prev = v
            chain.append(v)
        if not _hard_direct_step_ok(prev, b):
            raise RuntimeError("reverse-width tail step invalid")
        return chain, vidx_start + n

    def _try_with_n(n: int, start_vidx: int) -> Tuple[List[dict], int] | None:
        chain: List[dict] = []
        prev = a
        for k in range(n):
            frac = (k + 1) / (n + 1)
            raw_w = float(a["width"] + (b["width"] - a["width"]) * frac)
            raw_t = float(a["thickness"] + (b["thickness"] - a["thickness"]) * frac)
            raw_temp = float(_temp_center(a) + (_temp_center(b) - _temp_center(a)) * frac)
            w_cands = _nearest_n(V_WIDTHS, raw_w, 3)
            t_cands = _nearest_n(V_THICKS, min(2.0, max(0.5, raw_t)), 5)
            tb_cands = sorted(V_TEMP_BANDS, key=lambda x: abs(((x[0] + x[1]) / 2.0) - raw_temp))[:3]
            rem = n - k - 1
            best_v = None
            best_score = None
            # 先局部候选，若无解再扩展到全离散规格，减少“可桥接却断期”的情况。
            pools = [
                (w_cands, t_cands, tb_cands),
                (V_WIDTHS, V_THICKS, V_TEMP_BANDS),
            ]
            for ww, tt, bb in pools:
                for w in ww:
                    for t in tt:
                        for tb_min, tb_max in bb:
                            v = _make_virtual(start_vidx + k, _txt(a["line"]), float(w), float(t), float(tb_min), float(tb_max), campaign_id)
                            if not _hard_direct_step_ok(prev, v):
                                continue
                            if rem == 0:
                                if not _hard_direct_step_ok(v, b):
                                    continue
                            else:
                                if _bridge_need(v, b, rem, strict_virtual_width_levels=bool(strict_virtual_width_levels)) > rem:
                                    continue
                            score = (
                                2.0 * abs(float(w) - raw_w)
                                + 300.0 * abs(float(t) - raw_t)
                                + abs(((tb_min + tb_max) / 2.0) - raw_temp)
                            )
                            if (best_score is None) or (score < best_score):
                                best_score = score
                                best_v = v
                if best_v is not None:
                    break
            if best_v is None:
                return None
            chain.append(best_v)
            prev = best_v
        if not _hard_direct_step_ok(prev, b):
            return None
        return chain, start_vidx + n

    # 先尝试最少桥接块，失败再逐步增加到上限，尽量避免强制断期。
    for n in range(need, max_virtual_chain + 1):
        res = _try_with_n(n, vidx_start)
        if res is not None:
            return res
    raise RuntimeError("virtual step invalid")


# LEGACY ONLY / NOT FOR PRODUCTION PATH:
# 兼容层仍保留一份显式 legacy 规则快照，仅用于旧入口复用共享桥接逻辑。
LEGACY_RULE = RuleConfig()
_legacy_spec = SHARED_BUILD_VIRTUAL_SPEC_VIEWS(LEGACY_RULE)
V_WIDTHS = _legacy_spec["widths"]
V_THICKS = _legacy_spec["thicks"]
V_TEMP_RANGE = _legacy_spec["temp_range"]
DUE_RANK = SHARED_DUE_RANK
_txt = SHARED_TXT
_temp_overlap_len = SHARED_TEMP_OVERLAP_LEN


def _bridge_need(*args, **kwargs):
    return SHARED_BRIDGE_NEED(*args, rule=LEGACY_RULE, **kwargs)


def _bridge_pair(*args, **kwargs):
    return SHARED_BRIDGE_PAIR(*args, rule=LEGACY_RULE, **kwargs)


def _attach_checks(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy().reset_index(drop=True)
    if out.empty:
        return out
    group_cols = ["line", "campaign_id"]
    out["seq"] = out.index + 1
    out["prev_width"] = out.groupby(group_cols)["width"].shift(1)
    out["prev_thickness"] = out.groupby(group_cols)["thickness"].shift(1)
    out["prev_group"] = out.groupby(group_cols)["steel_group"].shift(1)
    out["prev_is_virtual"] = out.groupby(group_cols)["is_virtual"].shift(1).fillna(False).astype(bool)
    out["prev_temp_min"] = out.groupby(group_cols)["temp_min"].shift(1)
    out["prev_temp_max"] = out.groupby(group_cols)["temp_max"].shift(1)
    out["width_drop"] = out["prev_width"] - out["width"]
    out["narrow_to_wide"] = (out["width"] > out["prev_width"]).fillna(False)
    out["width_rise"] = (out["width"] - out["prev_width"]).clip(lower=0.0).fillna(0.0)
    out["physical_reverse_violation"] = (
        out["narrow_to_wide"] & (out["width_rise"] > float(TH.max_width_rise_physical_step))
    ).fillna(False)
    out["logical_reverse_flag"] = (
        out["narrow_to_wide"]
        & (~out["prev_is_virtual"])
        & (~out["is_virtual"])
    ).fillna(False)
    out["width_jump_violation"] = (
        (out["width_drop"] > float(TH.max_width_drop))
        | out["physical_reverse_violation"]
    ).fillna(False)
    out["width_over_mm"] = (
        ((out["width_drop"] - float(TH.max_width_drop)).clip(lower=0.0))
        + ((out["width_rise"] - float(TH.max_width_rise_physical_step)).clip(lower=0.0))
    ).fillna(0.0)
    out["thickness_violation"] = (
        ((out["prev_thickness"] > 0.8) & (((out["thickness"] - out["prev_thickness"]).abs() / out["prev_thickness"]) > 0.30))
        | ((out["prev_thickness"] >= 0.6) & (out["prev_thickness"] <= 0.8) & ((out["thickness"] - out["prev_thickness"]).abs() > 0.2))
        | ((out["prev_thickness"] < 0.6) & ((out["thickness"] - out["prev_thickness"]).abs() > 0.1))
    ).fillna(False)
    out["thick_over_mm"] = 0.0
    mask_hi = out["prev_thickness"] > 0.8
    out.loc[mask_hi, "thick_over_mm"] = (
        (out.loc[mask_hi, "thickness"] - out.loc[mask_hi, "prev_thickness"]).abs()
        - out.loc[mask_hi, "prev_thickness"].abs() * 0.30
    ).clip(lower=0.0)
    mask_mid = (out["prev_thickness"] >= 0.6) & (out["prev_thickness"] <= 0.8)
    out.loc[mask_mid, "thick_over_mm"] = (
        (out.loc[mask_mid, "thickness"] - out.loc[mask_mid, "prev_thickness"]).abs() - 0.2
    ).clip(lower=0.0)
    mask_lo = out["prev_thickness"] < 0.6
    out.loc[mask_lo, "thick_over_mm"] = (
        (out.loc[mask_lo, "thickness"] - out.loc[mask_lo, "prev_thickness"]).abs() - 0.1
    ).clip(lower=0.0)
    out["group_switch"] = (out["steel_group"] != out["prev_group"]).fillna(False)
    out["non_pc_direct_switch"] = (
        out["group_switch"]
        & ~out["steel_group"].astype(str).str.upper().isin(["PC", "VIRTUAL_PC", "普碳"])
        & ~out["prev_group"].astype(str).str.upper().isin(["PC", "VIRTUAL_PC", "普碳"])
    ).fillna(False)
    out["temp_overlap"] = (
        pd.concat([out["prev_temp_max"], out["temp_max"]], axis=1).min(axis=1)
        - pd.concat([out["prev_temp_min"], out["temp_min"]], axis=1).max(axis=1)
    ).clip(lower=0.0)
    out["temp_required_overlap"] = ((~out["is_virtual"]) & (~out["prev_is_virtual"])).map(
        lambda x: float(TH.min_temp_overlap_real_real) if bool(x) else 0.0
    )
    out["temp_conflict"] = out["temp_overlap"] < out["temp_required_overlap"]
    out["temp_shortage"] = (out["temp_required_overlap"] - out["temp_overlap"]).clip(lower=0.0)
    out["width_smooth_mm"] = out["width_drop"].clip(lower=0.0).fillna(0.0)
    out["thick_smooth_mm"] = (out["thickness"] - out["prev_thickness"]).abs().fillna(0.0)
    out["temp_margin_pen"] = (
        (float(TH.temp_margin_target) - out["temp_overlap"]).clip(lower=0.0)
        * (
            (out["temp_overlap"] >= float(TH.min_temp_overlap_real_real))
            & (out["temp_overlap"] < float(TH.temp_margin_target))
        ).astype(int)
    )
    # 同一轧期逻辑逆宽次数硬约束：最多1次（仅统计实物->实物的逆宽起点）。
    rev_cnt = out.groupby(group_cols)["logical_reverse_flag"].transform("sum").fillna(0).astype(int)
    out["logical_reverse_cnt_campaign"] = rev_cnt
    out["logical_reverse_count_violation"] = (
        rev_cnt > int(TH.max_logical_reverse_per_campaign)
    ).fillna(False)
    out["width_jump_violation"] = (
        out["width_jump_violation"] | out["logical_reverse_count_violation"]
    ).fillna(False)
    out["width_over_mm"] = out["width_over_mm"] + (
        (rev_cnt - int(TH.max_logical_reverse_per_campaign)).clip(lower=0).astype(float)
        * float(getattr(TH, "real_reverse_step_max_mm", 50.0))
    )
    first_rows = out.groupby(group_cols).head(1).index
    out.loc[first_rows, ["narrow_to_wide", "width_jump_violation", "thickness_violation", "group_switch", "non_pc_direct_switch", "temp_conflict"]] = False
    out.loc[first_rows, ["width_over_mm", "thick_over_mm", "temp_shortage", "width_smooth_mm", "thick_smooth_mm", "temp_margin_pen"]] = 0.0
    out.loc[first_rows, ["physical_reverse_violation", "logical_reverse_flag", "logical_reverse_count_violation"]] = False
    return out


def _assign_campaigns_and_filter(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    轧期分配与吨位约束处理：
    - 先按顺序切轧期（支持强制断期）
    - campaign_ton_max 上限为硬约束（超上限剔除）
    - campaign_ton_min 下限作为高罚分目标（不在此处剔除）
    """
    if df.empty:
        return df.copy(), pd.DataFrame()
    out = df.copy().reset_index(drop=True)
    if "master_slot" in out.columns and out["master_slot"].notna().any():
        # 第三/四阶段：辊期按主模型槽位落地。
        # 同时尊重槽位内的强制断开点，避免把不可连续段硬拼到同一辊期。
        out["master_slot"] = pd.to_numeric(out["master_slot"], errors="coerce").fillna(0).astype(int)
        out = out[out["master_slot"] > 0].copy().reset_index(drop=True)
        cids = []
        cur_slot = None
        seg = 1
        for _, row in out.iterrows():
            s = int(row["master_slot"])
            fb = bool(row.get("force_break_before", False))
            if (cur_slot is None) or (s != cur_slot):
                cur_slot = s
                seg = 1
            elif fb:
                seg += 1
            cids.append(int(s * 1000 + seg))
        out["campaign_id"] = cids
    else:
        out["campaign_id"] = 0
        cid = 1
        cur_tons = 0.0
        for i, row in out.iterrows():
            if bool(row.get("force_break_before", False)) and cur_tons > 0:
                cid += 1
                cur_tons = 0.0
            tons = float(row["tons"])
            if cur_tons >= float(TH.campaign_ton_min) and (cur_tons + tons > float(TH.campaign_ton_max)):
                cid += 1
                cur_tons = 0.0
            out.at[i, "campaign_id"] = cid
            cur_tons += tons

    # 轧期吨位按“真实订单”统计，不计虚拟板坯。
    real = out[~out["is_virtual"]].copy()
    tons_by = real.groupby(["line", "campaign_id"])["tons"].sum().reset_index(name="campaign_tons")
    bad = tons_by[tons_by["campaign_tons"] > float(TH.campaign_ton_max)][["line", "campaign_id"]]
    if bad.empty:
        return out, pd.DataFrame()

    key = out[["line", "campaign_id"]].astype(str).agg("||".join, axis=1)
    bad_set = set(bad.astype(str).agg("||".join, axis=1).tolist())
    mask_bad = key.isin(bad_set)
    dropped = out[mask_bad & (~out["is_virtual"])].copy()
    dropped["drop_reason"] = "CAMPAIGN_TON_OVER_MAX"
    good = out[~mask_bad].copy().reset_index(drop=True)
    return good, dropped


def _campaign_ton_penalty(df: pd.DataFrame) -> Tuple[int, float]:
    """统计低吨位轧期惩罚（<campaign_ton_min）。返回(低吨位轧期数, 总缺口吨位)。"""
    if df.empty:
        return 0, 0.0
    real = df[~df["is_virtual"]].copy()
    if real.empty:
        return 0, 0.0
    tons_by = real.groupby(["line", "campaign_id"])["tons"].sum().reset_index(name="campaign_tons")
    low = tons_by[tons_by["campaign_tons"] < float(TH.campaign_ton_min)].copy()
    if low.empty:
        return 0, 0.0
    low["gap_tons"] = float(TH.campaign_ton_min) - low["campaign_tons"]
    return int(len(low)), float(low["gap_tons"].sum())


def _campaign_ton_excess(df: pd.DataFrame) -> float:
    if df.empty:
        return 0.0
    real = df[~df["is_virtual"]].copy()
    if real.empty:
        return 0.0
    csum = real.groupby(["line", "campaign_id"])["tons"].sum().reset_index(name="campaign_tons")
    return float((csum["campaign_tons"] - float(TH.campaign_ton_max)).clip(lower=0.0).sum())


def _campaign_ton_target_deviation(df: pd.DataFrame, target: float = float(TH.campaign_ton_target)) -> float:
    if df.empty:
        return 0.0
    real = df[~df["is_virtual"]].copy()
    if real.empty:
        return 0.0
    csum = real.groupby(["line", "campaign_id"])["tons"].sum().reset_index(name="campaign_tons")
    return float((csum["campaign_tons"] * 10.0 - target * 10.0).abs().sum())


def _max_virtual_chain_len(df: pd.DataFrame) -> int:
    if df.empty:
        return 0
    m = 0
    for _, g in df.groupby(["line", "campaign_id"], dropna=False):
        cur = 0
        for iv in g["is_virtual"].tolist():
            if bool(iv):
                cur += 1
                if cur > m:
                    m = cur
            else:
                cur = 0
    return int(m)


def _pure_virtual_campaign_count(df: pd.DataFrame) -> int:
    if df.empty:
        return 0
    cnt = 0
    for _, g in df.groupby(["line", "campaign_id"], dropna=False):
        if bool(g["is_virtual"].all()):
            cnt += 1
    return int(cnt)


def _weighted_penalties(df: pd.DataFrame, dropped_cnt: int, w: PenaltyWeights) -> Dict[str, float]:
    out: Dict[str, float] = {}
    if df.empty:
        out["total_penalty"] = float(w.unassigned_real * dropped_cnt)
        return out
    real = (~df["is_virtual"])
    # 结构序号在当前实现中按构造天然成立，保留为显式项便于后续扩展。
    seq_unique_v = 0
    seq_contiguous_v = 0
    seq_start_v = 0
    roll_line_mismatch_v = int(
        (
            ((df["line_capability"] == "big_only") & (df["line"] != "big_roll"))
            | ((df["line_capability"] == "small_only") & (df["line"] != "small_roll"))
        ).sum()
    ) if "line_capability" in df.columns else 0
    ton_over = _campaign_ton_excess(df)
    low_cnt, ton_under_gap = _campaign_ton_penalty(df)
    temp_short = float(df.loc[real, "temp_shortage"].sum()) if "temp_shortage" in df.columns else 0.0
    width_over = float(df.loc[real, "width_over_mm"].sum()) if "width_over_mm" in df.columns else 0.0
    thick_over = float(df.loc[real, "thick_over_mm"].sum()) if "thick_over_mm" in df.columns else 0.0
    non_pc = int(df.loc[real, "non_pc_direct_switch"].sum()) if "non_pc_direct_switch" in df.columns else 0
    vchain_excess = max(0, _max_virtual_chain_len(df) - int(TH.virtual_chain_penalty_threshold))
    unassigned = int(dropped_cnt)
    virtual_cnt = int(df["is_virtual"].sum())
    virtual_ton10 = int(round(float(df.loc[df["is_virtual"], "tons"].sum()) * 10.0))
    total_ton10 = int(round(float(df["tons"].sum()) * 10.0))
    # 中约束：虚拟吨位占比超限（默认20%）计罚
    virtual_ratio_pen_units = max(
        0,
        virtual_ton10 * int(TH.virtual_ton_ratio_den) - total_ton10 * int(TH.virtual_ton_ratio_num),
    )
    width_smooth = float(df.loc[real, "width_smooth_mm"].sum()) if "width_smooth_mm" in df.columns else 0.0
    thick_smooth = float(df.loc[real, "thick_smooth_mm"].sum()) if "thick_smooth_mm" in df.columns else 0.0
    temp_margin = float(df.loc[real, "temp_margin_pen"].sum()) if "temp_margin_pen" in df.columns else 0.0
    ton_target_dev = _campaign_ton_target_deviation(df, target=w.ton_target_value)
    pure_virtual_campaign_cnt = _pure_virtual_campaign_count(df)

    out["seq_unique_pen"] = float(w.seq_unique * seq_unique_v)
    out["seq_contiguous_pen"] = float(w.seq_contiguous * seq_contiguous_v)
    out["seq_start_pen"] = float(w.seq_start * abs(seq_start_v))
    out["roll_line_mismatch_pen"] = float(w.roll_line_mismatch * roll_line_mismatch_v)
    out["ton_over_pen"] = float(w.ton_over * ton_over)
    out["ton_under_pen"] = float(w.ton_under * ton_under_gap)
    out["temp_short_pen"] = float(w.temp_shortage * temp_short)
    out["width_over_pen"] = float(w.width_violation * width_over)
    out["thick_over_pen"] = float(w.thick_violation * thick_over)
    out["non_pc_pen"] = float(w.non_pc_switch * non_pc)
    out["vchain_excess_pen"] = float(w.virtual_chain_excess * vchain_excess)
    out["unassigned_pen"] = float(w.unassigned_real * unassigned)
    out["virtual_ratio_pen"] = float(w.virtual_ratio * virtual_ratio_pen_units)
    out["pure_virtual_campaign_pen"] = float(w.pure_virtual_campaign * pure_virtual_campaign_cnt)
    out["virtual_use_pen"] = float(w.virtual_use * virtual_cnt)
    out["width_smooth_pen"] = float(w.width_smooth * width_smooth)
    out["thick_smooth_pen"] = float(w.thick_smooth * thick_smooth)
    out["temp_margin_pen"] = float(w.temp_margin * temp_margin)
    out["ton_target_pen"] = float(w.ton_target * ton_target_dev)
    out["low_ton_campaign_cnt"] = float(low_cnt)
    out["low_ton_gap_tons"] = float(ton_under_gap)
    out["pure_virtual_campaign_cnt"] = float(pure_virtual_campaign_cnt)
    out["total_penalty"] = float(
        out["seq_unique_pen"] + out["seq_contiguous_pen"] + out["seq_start_pen"] + out["roll_line_mismatch_pen"]
        + out["ton_over_pen"] + out["ton_under_pen"] + out["temp_short_pen"] + out["width_over_pen"] + out["thick_over_pen"]
        + out["non_pc_pen"] + out["vchain_excess_pen"] + out["unassigned_pen"] + out["virtual_ratio_pen"] + out["pure_virtual_campaign_pen"] + out["virtual_use_pen"]
        + out["width_smooth_pen"] + out["thick_smooth_pen"] + out["temp_margin_pen"] + out["ton_target_pen"]
    )
    return out


def _line_schedule(data_line: pd.DataFrame, cfg: SolveConfig) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    单产线多轮求解入口：
    每轮执行「求序 -> 插入虚拟卷 -> 分轧期 -> 硬约束清洗」，并记录耗时与指标。
    """
    logs = []
    best_df = pd.DataFrame()
    best_dropped = pd.DataFrame()
    best_score: Tuple[float, int, int] | None = None
    weights = PenaltyWeights()
    use_master_slot = ("master_slot" in data_line.columns) and bool(data_line["master_slot"].notna().any())

    line_name = _txt(data_line.iloc[0]["line"]) if not data_line.empty else "unknown"
    print(f"[APS] 产线={line_name} 开始排程, 订单数={len(data_line)}, 轮次={cfg.rounds}, 求解时限={cfg.time_limit_seconds}s")
    if use_master_slot:
        slots = int(pd.to_numeric(data_line["master_slot"], errors="coerce").dropna().nunique())
        print(f"[APS] 产线={line_name} 启用主模型槽位落地, 槽位数={slots}")
    for r in range(1, cfg.rounds + 1):
        t0 = perf_counter()
        print(f"[APS] 产线={line_name} 第{r}/{cfg.rounds}轮 开始求解...")
        model_s = 0.0
        solve_s = 0.0
        dropped_rows: List[dict] = []
        if use_master_slot:
            # 第四阶段（中间版）：主模型确定槽位后，槽位内再用 CP-SAT 路由求序。
            # 这样避免“槽位内纯启发式”导致过多断期。
            with_gid = data_line.assign(_gid=range(1, len(data_line) + 1))
            if "master_seq" in with_gid.columns and with_gid["master_seq"].notna().any():
                seq = (
                    with_gid.sort_values(["master_slot", "master_seq"], kind="mergesort")["_gid"].astype(int).tolist()
                )
                edge_bridge = {}
                for k in range(len(seq) - 1):
                    a = with_gid.iloc[seq[k] - 1].to_dict()
                    b = with_gid.iloc[seq[k + 1] - 1].to_dict()
                    if int(a.get("master_slot", -1)) != int(b.get("master_slot", -1)):
                        edge_bridge[(seq[k], seq[k + 1])] = 0
                    else:
                        edge_bridge[(seq[k], seq[k + 1])] = _bridge_need(a, b, cfg.max_virtual_chain, strict_virtual_width_levels=cfg.strict_virtual_width_levels)
            else:
                seq = []
                edge_bridge = {}
                model_acc = 0.0
                solve_acc = 0.0
                slot_vals = sorted(pd.to_numeric(with_gid["master_slot"], errors="coerce").dropna().astype(int).unique().tolist())
                for sv in slot_vals:
                    part = with_gid[pd.to_numeric(with_gid["master_slot"], errors="coerce").fillna(-1).astype(int) == sv].copy().reset_index(drop=True)
                    if part.empty:
                        continue
                    try:
                        loc_seq, loc_bridge, timing = _solve_line_once_cp_sat(part.drop(columns=["_gid"]), cfg, seed=3000 + 37 * r + sv)
                        model_acc += float(timing.get("model_seconds", 0.0))
                        solve_acc += float(timing.get("solve_seconds", 0.0))
                    except Exception:
                        loc_seq = _component_fallback_order(part.drop(columns=["_gid"]), cfg)
                        loc_bridge = {}
                        for kk in range(len(loc_seq) - 1):
                            a = part.iloc[loc_seq[kk] - 1].to_dict()
                            b = part.iloc[loc_seq[kk + 1] - 1].to_dict()
                            loc_bridge[(loc_seq[kk], loc_seq[kk + 1])] = _bridge_need(a, b, cfg.max_virtual_chain, strict_virtual_width_levels=cfg.strict_virtual_width_levels)
                    gid_seq = [int(part.iloc[i - 1]["_gid"]) for i in loc_seq]
                    seq.extend(gid_seq)
                    for (i1, i2), need in loc_bridge.items():
                        g1 = int(part.iloc[i1 - 1]["_gid"])
                        g2 = int(part.iloc[i2 - 1]["_gid"])
                        edge_bridge[(g1, g2)] = int(need)
                model_s = model_acc
                solve_s = solve_acc
        else:
            try:
                seq, edge_bridge, timing = _solve_line_once_cp_sat(data_line, cfg, seed=1000 + 37 * r)
                model_s = timing["model_seconds"]
                solve_s = timing["solve_seconds"]
            except Exception as e:
                print(f"[APS] 产线={line_name} 第{r}轮 CP-SAT回退到启发式，原因: {e}")
                seq = _component_fallback_order(data_line, cfg)
                edge_bridge = {}
                for k in range(len(seq) - 1):
                    a = data_line.iloc[seq[k] - 1].to_dict()
                    b = data_line.iloc[seq[k + 1] - 1].to_dict()
                    edge_bridge[(seq[k], seq[k + 1])] = _bridge_need(a, b, cfg.max_virtual_chain, strict_virtual_width_levels=cfg.strict_virtual_width_levels)

        t_post = perf_counter()
        ordered = data_line.iloc[[i - 1 for i in seq]].copy().reset_index(drop=True)
        ordered["is_virtual"] = False

        # 解后只做“落地化”：根据相邻边需要插入虚拟卷或断期，不做复杂补救优化。
        rows: List[dict] = []
        vidx = 1
        recs = ordered.to_dict("records")
        force_break_before = set()
        real_cnt = int((~ordered["is_virtual"]).sum()) if not ordered.empty else 0
        drop_budget = max(0, int(math.floor(real_cnt * float(cfg.max_unbridgeable_drop_ratio))))
        dropped_unbridgeable = 0
        dropped_idx: Set[int] = set()

        def _can_drop_order(rec: dict) -> bool:
            if bool(rec.get("is_virtual", False)):
                return False
            # 允许剔除少量非紧急实物，优先处理边角/难桥接订单，避免产生大量低吨位断期。
            due_rank = int(rec.get("due_rank", 3))
            if due_rank <= 1:
                return False
            tons = float(rec.get("tons", 0.0))
            # 非紧急订单均可候选；小吨位订单优先剔除。
            if tons <= 30.0:
                return True
            return due_rank >= 2

        i = 0
        while i < len(recs):
            if i in dropped_idx:
                i += 1
                continue
            cur = recs[i]
            row_cur = dict(cur)
            row_cur["force_break_before"] = (i in force_break_before)
            rows.append(row_cur)

            j = i + 1
            advanced = False
            while j < len(recs):
                if j in dropped_idx:
                    j += 1
                    continue
                nxt = recs[j]
                pair = (seq[i], seq[j])
                if use_master_slot and int(cur.get("master_slot", -1)) != int(nxt.get("master_slot", -1)):
                    force_break_before.add(j)
                    i = j
                    advanced = True
                    break
                need = edge_bridge.get(pair, _bridge_need(cur, nxt, cfg.max_virtual_chain, strict_virtual_width_levels=cfg.strict_virtual_width_levels))
                if need <= 0:
                    i = j
                    advanced = True
                    break
                try:
                    vs, vidx = _bridge_pair(
                        cur,
                        nxt,
                        cfg.max_virtual_chain,
                        vidx,
                        campaign_id=1,
                        strict_virtual_width_levels=cfg.strict_virtual_width_levels,
                        physical_reverse_step_mode=cfg.physical_reverse_step_mode,
                    )
                    for v in vs:
                        v["force_break_before"] = False
                        if use_master_slot:
                            v["master_slot"] = int(cur.get("master_slot", -1))
                    rows.extend(vs)
                    i = j
                    advanced = True
                    break
                except RuntimeError:
                    if (dropped_unbridgeable < drop_budget) and _can_drop_order(nxt):
                        dr = dict(nxt)
                        dr["drop_reason"] = "UNBRIDGEABLE_DROPPED"
                        dropped_rows.append(dr)
                        dropped_idx.add(j)
                        dropped_unbridgeable += 1
                        j += 1
                        continue
                    # 仍不可桥接：保留该单并强制断期
                    force_break_before.add(j)
                    i = j
                    advanced = True
                    break
            if not advanced:
                break

        cand = pd.DataFrame(rows) if rows else ordered.copy()
        if "is_virtual" not in cand.columns:
            cand["is_virtual"] = False

        checked = cand.copy()
        while True:
            checked, dropped_campaign = _assign_campaigns_and_filter(checked)
            if not dropped_campaign.empty:
                dropped_rows.extend(dropped_campaign.to_dict("records"))
            if checked.empty:
                break
            checked = _attach_checks(checked)
            hard_mask = (checked["thickness_violation"] | checked["non_pc_direct_switch"] | checked["width_jump_violation"] | checked["temp_conflict"]) & (~checked["is_virtual"])
            if not bool(hard_mask.any()):
                break
            bad = checked[hard_mask].copy()
            bad["drop_reason"] = "HARD_VIOLATION"
            dropped_rows.extend(bad.to_dict("records"))
            checked = checked[~hard_mask].copy().reset_index(drop=True)
        post_s = perf_counter() - t_post
        elapsed = perf_counter() - t0
        metric = {
            "round": r,
            "line": line_name,
            "rows_total": int(len(checked)),
            "virtual_cnt": int(checked["is_virtual"].sum()) if not checked.empty else 0,
            "dropped_cnt": int(len(dropped_rows)),
            "width_jump_violation_cnt": int((checked["width_jump_violation"] & (~checked["is_virtual"])).sum()) if not checked.empty else 0,
            "thickness_violation_cnt": int((checked["thickness_violation"] & (~checked["is_virtual"])).sum()) if not checked.empty else 0,
            "non_pc_direct_switch_cnt": int((checked["non_pc_direct_switch"] & (~checked["is_virtual"])).sum()) if not checked.empty else 0,
            "temp_conflict_cnt": int((checked["temp_conflict"] & (~checked["is_virtual"])).sum()) if not checked.empty else 0,
            "model_seconds": round(model_s, 3),
            "solve_seconds": round(solve_s, 3),
            "postprocess_seconds": round(post_s, 3),
            "elapsed_seconds": round(elapsed, 3),
        }
        pen = _weighted_penalties(checked, int(len(dropped_rows)), weights)
        metric["low_ton_campaign_cnt"] = int(pen.get("low_ton_campaign_cnt", 0))
        metric["low_ton_gap_tons"] = round(float(pen.get("low_ton_gap_tons", 0.0)), 1)
        metric["penalty_total"] = round(float(pen.get("total_penalty", 0.0)), 1)
        metric["penalty_ton_under"] = round(float(pen.get("ton_under_pen", 0.0)), 1)
        metric["penalty_unassigned"] = round(float(pen.get("unassigned_pen", 0.0)), 1)
        metric["penalty_virtual_ratio"] = round(float(pen.get("virtual_ratio_pen", 0.0)), 1)
        metric["penalty_pure_virtual_campaign"] = round(float(pen.get("pure_virtual_campaign_pen", 0.0)), 1)
        metric["pure_virtual_campaign_cnt"] = int(pen.get("pure_virtual_campaign_cnt", 0.0))
        metric["penalty_virtual_use"] = round(float(pen.get("virtual_use_pen", 0.0)), 1)
        logs.append(metric)
        print(
            f"[APS] 产线={line_name} 第{r}轮 完成, 总耗时={elapsed:.1f}s, 建模={model_s:.1f}s, "
            f"求解={solve_s:.1f}s, 解后处理={post_s:.1f}s, rows={metric['rows_total']}, "
            f"virtual={metric['virtual_cnt']}, dropped={metric['dropped_cnt']}, "
            f"不可桥接剔除={dropped_unbridgeable}/{drop_budget}, "
            f"低吨位轧期={metric['low_ton_campaign_cnt']}, 低吨位缺口={metric['low_ton_gap_tons']:.1f}吨, "
            f"总罚分={metric['penalty_total']:.1f}"
        )

        real_cnt = int((~checked["is_virtual"]).sum()) if not checked.empty else 0
        score = (
            -float(metric["penalty_total"]),
            real_cnt,
            -metric["virtual_cnt"],
        )
        if (best_score is None) or (score > best_score):
            best_score = score
            best_df = checked.copy()
            best_df["line_order_local"] = range(1, len(best_df) + 1)
            best_dropped = pd.DataFrame(dropped_rows)

    return best_df, pd.DataFrame(logs), best_dropped


def _finalize_output(final_df: pd.DataFrame) -> pd.DataFrame:
    """汇总两条产线结果，重编轧期号并生成全局/线内/轧期内序号。"""
    if final_df.empty:
        return final_df
    remapped = []
    for _, g in final_df.groupby("line", sort=False):
        if "line_order_local" in g.columns:
            g = g.sort_values("line_order_local", kind="mergesort")
        uniq = list(dict.fromkeys(g["campaign_id"].tolist()))
        mp = {old: i + 1 for i, old in enumerate(uniq)}
        gg = g.copy()
        gg["campaign_id"] = gg["campaign_id"].map(mp)
        remapped.append(gg)
    out = pd.concat(remapped, ignore_index=True)
    line_order = {"big_roll": 0, "small_roll": 1}
    out["_line_order"] = out["line"].map(lambda x: line_order.get(_txt(x), 99))
    # 当 campaign_id 已经在主模型/解码阶段确定后，最终输出必须按轧期内真实顺序落地，
    # 不能再用旧的 line_order_local 打乱同一轧期内部相邻关系。
    sort_cols = ["_line_order", "campaign_id", "seq"]
    out = out.sort_values(sort_cols, kind="mergesort").reset_index(drop=True)
    out = out.drop(columns=["_line_order"])
    out["global_seq"] = out.index + 1
    out["line_seq"] = out.groupby("line").cumcount() + 1
    out["campaign_seq"] = out.groupby(["line", "campaign_id"]).cumcount() + 1
    # 用户要求：虚拟板坯也要有“轧期内真实序号”。
    # 这里统一按轧期内连续序号赋值，确保每一行都有序号。
    out["campaign_real_seq"] = out.groupby(["line", "campaign_id"]).cumcount() + 1
    out["line_name"] = out["line"].map({"big_roll": "大辊线", "small_roll": "小辊线"}).fillna(out["line"])
    return _attach_checks(out)


def _autosize_excel(writer: pd.ExcelWriter) -> None:
    for ws in writer.book.worksheets:
        for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
            max_len = 0
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(idx)].width = min(60, max(10, max_len + 2))


def _drop_bad_campaigns_final_real(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """最终兜底：按真实订单吨位检查辊期硬约束（仅 >campaign_ton_max）。"""
    if df.empty:
        return df.copy(), pd.DataFrame()
    out = df.copy()
    real = out[~out["is_virtual"]].copy()
    if real.empty:
        return out, pd.DataFrame()
    csum = real.groupby(["line", "campaign_id"])["tons"].sum().reset_index(name="campaign_tons")
    bad = csum[csum["campaign_tons"] > float(TH.campaign_ton_max)][["line", "campaign_id"]]
    if bad.empty:
        return out, pd.DataFrame()
    key = out[["line", "campaign_id"]].astype(str).agg("||".join, axis=1)
    bad_key = set(bad.astype(str).agg("||".join, axis=1).tolist())
    mask_bad = key.isin(bad_key)
    dropped = out[mask_bad & (~out["is_virtual"])].copy()
    dropped["drop_reason"] = "FINAL_CAMPAIGN_TON_OVER_MAX"
    good = out[~mask_bad].copy().reset_index(drop=True)
    return good, dropped


def _schedule_all_lines_once(data: pd.DataFrame, cfg: SolveConfig, pass_idx: int) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """一次完整排程（按当前 line 分配）：两条线分别求解并汇总。"""
    all_seq: List[pd.DataFrame] = []
    all_logs: List[pd.DataFrame] = []
    all_dropped: List[pd.DataFrame] = []
    for line_name in ["big_roll", "small_roll"]:
        d = data[data["line"] == line_name].copy().reset_index(drop=True)
        if d.empty:
            print(f"[APS] 迭代{pass_idx}: 产线={line_name} 无订单，跳过")
            continue
        print(f"[APS] 迭代{pass_idx}: 产线={line_name} 待排订单数={len(d)}")
        seq_df, logs_df, dropped_df = _line_schedule(d, cfg)
        if not seq_df.empty:
            all_seq.append(seq_df)
        if not logs_df.empty:
            logs_df = logs_df.copy()
            logs_df["rebalance_pass"] = pass_idx
            all_logs.append(logs_df)
        if not dropped_df.empty:
            dropped_df = dropped_df.copy()
            dropped_df["rebalance_pass"] = pass_idx
            all_dropped.append(dropped_df)

    if not all_seq:
        raise RuntimeError("No schedulable data")

    final_df = _finalize_output(pd.concat(all_seq, ignore_index=True))
    dropped_df = pd.concat(all_dropped, ignore_index=True) if all_dropped else pd.DataFrame()
    while True:
        final_df, bad_final = _drop_bad_campaigns_final_real(final_df)
        if bad_final.empty:
            break
        bad_final = bad_final.copy()
        bad_final["rebalance_pass"] = pass_idx
        dropped_df = pd.concat([dropped_df, bad_final], ignore_index=True) if not dropped_df.empty else bad_final
        if final_df.empty:
            break
        final_df = _finalize_output(final_df)

    logs_df = pd.concat(all_logs, ignore_index=True) if all_logs else pd.DataFrame()
    return final_df, logs_df, dropped_df


def _flip_dropped_dual_orders(data: pd.DataFrame, dropped_df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
    """将本轮未排上的 dual 实物订单切换到另一条线，返回新分配和迁移数量。"""
    if dropped_df.empty:
        return data, 0
    need = dropped_df.copy()
    if "line_source" in need.columns:
        need = need[need["line_source"] == "dual"]
    if "is_virtual" in need.columns:
        need = need[~need["is_virtual"]]
    if need.empty:
        return data, 0

    ids: Set[str] = set()
    for col in ["source_order_id", "parent_order_id", "order_id"]:
        if col in need.columns:
            ids.update(_normalize_order_id(v) for v in need[col].tolist() if _normalize_order_id(v))
            if ids:
                break
    if not ids:
        return data, 0

    out = data.copy()
    moved = 0
    for idx, row in out.iterrows():
        if _txt(row.get("line_source")) != "dual":
            continue
        key = _normalize_order_id(row.get("source_order_id", row.get("order_id", "")))
        if key not in ids:
            continue
        cur = _txt(row.get("line"))
        nxt = "small_roll" if cur == "big_roll" else "big_roll"
        out.at[idx, "line"] = nxt
        out.at[idx, "proc_hours"] = float(row["proc_hours_small"]) if nxt == "small_roll" else float(row["proc_hours_big"])
        moved += 1
    return out, moved


def export_schedule_results(
    final_df: pd.DataFrame,
    rounds_df: pd.DataFrame,
    dropped_df: pd.DataFrame,
    output_path: str,
    input_order_count: int,
    t_start: float | None = None,
    engine_used: str = "unknown",
    equivalence_summary: Dict[str, object] | None = None,
    failure_diagnostics: Dict[str, object] | None = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """兼容层导出接口：统一转发到 io.result_writer。"""
    return _export_schedule_results_v2(
        final_df=final_df,
        rounds_df=rounds_df,
        dropped_df=dropped_df,
        output_path=output_path,
        input_order_count=input_order_count,
        rule=RuleConfig(
            campaign_ton_min=float(TH.campaign_ton_min),
            campaign_ton_max=float(TH.campaign_ton_max),
            campaign_ton_target=float(TH.campaign_ton_target),
            min_temp_overlap_real_real=float(TH.min_temp_overlap_real_real),
            max_width_drop=float(TH.max_width_drop),
            virtual_reverse_attach_max_mm=float(getattr(TH, "virtual_reverse_attach_max_mm", 50.0)),
            max_width_rise_physical_step=float(TH.max_width_rise_physical_step),
            max_logical_reverse_per_campaign=int(TH.max_logical_reverse_per_campaign),
            virtual_chain_penalty_threshold=int(TH.virtual_chain_penalty_threshold),
            temp_margin_target=float(TH.temp_margin_target),
            virtual_ton_ratio_num=int(TH.virtual_ton_ratio_num),
            virtual_ton_ratio_den=int(TH.virtual_ton_ratio_den),
        ),
        t_start=t_start,
        engine_used=engine_used,
        equivalence_summary=equivalence_summary,
        failure_diagnostics=failure_diagnostics,
    )


def _build_schedule_legacy_impl(
    orders_path: str,
    steel_info_path: str,
    output_path: str,
    config: SolveConfig | None = None,
    prepared_orders: pd.DataFrame | None = None,
    prepared_dropped: pd.DataFrame | None = None,
    use_dual_master: bool = True,
    enable_rebalance: bool = True,
    engine_used: str = "legacy_fallback",
    equivalence_summary: Dict[str, object] | None = None,
    failure_diagnostics: Dict[str, object] | None = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    排程总入口：
    数据准备 -> dual 分配 -> 分线排程 -> 结果汇总 -> Excel/CSV 输出。
    """
    t_start = perf_counter()
    cfg = config or SolveConfig()
    print(f"[APS] 开始构建排程: orders={orders_path}, steel_info={steel_info_path}")
    print(
        f"[APS] 参数: max_orders={cfg.max_orders}, rounds={cfg.rounds}, "
        f"time_limit_seconds={cfg.time_limit_seconds}, max_virtual_chain={cfg.max_virtual_chain}"
    )

    if prepared_orders is not None:
        orders = prepared_orders.copy().reset_index(drop=True)
        print(f"[APS] 使用外部预处理订单, 订单总数(规划粒度)={len(orders)}")
    else:
        grade_catalog = PREPROCESS_GRADE_CATALOG.build(Path(steel_info_path))
        orders = PREPROCESS_PREPARE_ORDERS(Path(orders_path), Path(steel_info_path), cfg, grade_catalog=grade_catalog)
        print(f"[APS] 数据准备完成, 订单总数(规划粒度)={len(orders)}")

    if use_dual_master:
        data = _assign_dual_master(orders, time_limit=min(12.0, cfg.time_limit_seconds * 0.15))
    else:
        data = orders.copy()
        if "proc_hours" not in data.columns:
            data["proc_hours"] = data.apply(
                lambda r: r.get("proc_hours_big", 0.0) if _txt(r.get("line", "")) == "big_roll" else r.get("proc_hours_small", 0.0),
                axis=1,
            )

    src_cnt = data["line_source"].value_counts(dropna=False).to_dict()
    print(
        f"[APS] 订单来源统计: big_only={int(src_cnt.get('big_only', 0))}, "
        f"small_only={int(src_cnt.get('small_only', 0))}, dual={int(src_cnt.get('dual', 0))}"
    )
    dual_assigned = data[data["line_source"] == "dual"]
    if not dual_assigned.empty:
        dual_big = dual_assigned[dual_assigned["line"] == "big_roll"]
        dual_small = dual_assigned[dual_assigned["line"] == "small_roll"]
        print(
            f"[APS] dual分配结果: big_roll={len(dual_big)}单/{dual_big['tons'].sum():.1f}吨, "
            f"small_roll={len(dual_small)}单/{dual_small['tons'].sum():.1f}吨"
        )

    # dual 订单“排中再分”：按当前分配排程 -> 把 dual 丢单切到另一线 -> 再排，迭代取最优。
    max_rebalance_passes = 3 if enable_rebalance else 1
    best_final = pd.DataFrame()
    best_rounds = pd.DataFrame()
    best_dropped = pd.DataFrame()
    best_score: Tuple[float, int, int] | None = None
    weights = PenaltyWeights()
    cur_data = data.copy()
    seen_signatures: Set[Tuple[Tuple[str, str], ...]] = set()

    for p in range(1, max_rebalance_passes + 1):
        dual_now = cur_data[cur_data["line_source"] == "dual"]
        if not dual_now.empty:
            dual_big = dual_now[dual_now["line"] == "big_roll"]
            dual_small = dual_now[dual_now["line"] == "small_roll"]
            print(
                f"[APS] 迭代{p}: dual分配 big_roll={len(dual_big)}单/{dual_big['tons'].sum():.1f}吨, "
                f"small_roll={len(dual_small)}单/{dual_small['tons'].sum():.1f}吨"
            )

        final_p, rounds_p, dropped_p = _schedule_all_lines_once(cur_data, cfg, pass_idx=p)
        pen = _weighted_penalties(final_p, int(len(dropped_p)), weights)
        hard_v = int(
            (
                (final_p["width_jump_violation"] | final_p["thickness_violation"] | final_p["temp_conflict"] | final_p["non_pc_direct_switch"])
                & (~final_p["is_virtual"])
            ).sum()
        )
        low_campaign_cnt = int(pen.get("low_ton_campaign_cnt", 0))
        low_gap_tons = float(pen.get("low_ton_gap_tons", 0.0))
        real_cnt = int((~final_p["is_virtual"]).sum())
        virtual_cnt = int(final_p["is_virtual"].sum())
        dropped_cnt = int(len(dropped_p))
        score = (-float(pen.get("total_penalty", 0.0)), real_cnt, -virtual_cnt)
        print(
            f"[APS] 迭代{p}: 结果 real={real_cnt}, virtual={virtual_cnt}, dropped={dropped_cnt}, "
            f"hard={hard_v}, 低吨位轧期={low_campaign_cnt}, 低吨位缺口={low_gap_tons:.1f}吨, "
            f"总罚分={float(pen.get('total_penalty', 0.0)):.1f}, score={score}"
        )
        if (best_score is None) or (score > best_score):
            best_score = score
            best_final = final_p
            best_rounds = rounds_p
            best_dropped = dropped_p

        if not enable_rebalance:
            print(f"[APS] 迭代{p}: 关闭重分配，结束迭代")
            break
        next_data, moved = _flip_dropped_dual_orders(cur_data, dropped_p)
        if moved <= 0:
            print(f"[APS] 迭代{p}: 无可换线 dual 丢单，结束迭代")
            break
        sig_pairs = next_data[next_data["line_source"] == "dual"][["source_order_id", "line"]].copy()
        sig_pairs["source_order_id"] = sig_pairs["source_order_id"].map(_normalize_order_id)
        sig = tuple(sorted((str(a), str(b)) for a, b in sig_pairs.to_records(index=False)))
        if sig in seen_signatures:
            print(f"[APS] 迭代{p}: 分配状态已重复，停止迭代避免循环")
            break
        seen_signatures.add(sig)
        cur_data = next_data
        print(f"[APS] 迭代{p}: dual丢单换线 {moved} 单，进入下一轮")

    if best_final.empty:
        raise RuntimeError("No schedulable data")
    final_df = best_final
    rounds_df = best_rounds
    dropped_df = best_dropped
    if prepared_dropped is not None and not prepared_dropped.empty:
        dropped_df = pd.concat([prepared_dropped.copy(), dropped_df], ignore_index=True) if not dropped_df.empty else prepared_dropped.copy()

    return export_schedule_results(
        final_df=final_df,
        rounds_df=rounds_df,
        dropped_df=dropped_df,
        output_path=output_path,
        input_order_count=int(len(data)),
        t_start=t_start,
        engine_used=engine_used,
        equivalence_summary=equivalence_summary,
        failure_diagnostics=failure_diagnostics,
    )


def _to_planner_config(config: SolveConfig | object | None):
    from aps_cp_sat.config import ModelConfig, PlannerConfig, RuleConfig, ScoreConfig
    from aps_cp_sat.config.parameters import build_profile_config

    if isinstance(config, PlannerConfig):
        return config

    if config is None:
        return PlannerConfig()

    if isinstance(config, SolveConfig):
        # 1. If user explicitly provided a specific profile_name, use it.
        # Fallback to default (which will be mapped to constructive_lns_search by pipeline)
        # only if profile_name is empty or "default".
        profile = str(config.profile_name).strip() if config.profile_name else ""
        if profile and profile != "default":
            # Build full profile config
            cfg = build_profile_config(
                profile,
                validation_mode=bool(config.validation_mode),
                production_compatibility_mode=bool(config.production_compatibility_mode)
            )
            
            # 2. Check main_solver_strategy constraint if explicitly requested
            if config.main_solver_strategy:
                if config.main_solver_strategy != cfg.model.main_solver_strategy:
                    raise ValueError(
                        f"[APS][SolveConfig] Strategy mismatch: "
                        f"Requested profile '{profile}' uses '{cfg.model.main_solver_strategy}', "
                        f"but explicitly asked for main_solver_strategy='{config.main_solver_strategy}'"
                    )
            
            # Apply SolveConfig overrides on top of the built profile
            cfg.model.max_orders = int(config.max_orders)
            cfg.model.time_limit_seconds = float(config.time_limit_seconds)
            return cfg

        # If no explicit profile requested, fall back to old ad-hoc mapping
        model = ModelConfig(
            max_orders=int(config.max_orders),
            rounds=int(config.rounds),
            time_limit_seconds=float(config.time_limit_seconds),
            master_profile_count=int(config.master_profile_count),
            master_seed_count=int(config.master_seed_count),
            master_num_workers=int(config.master_num_workers),
            allow_fallback=bool(config.allow_fallback),
            allow_legacy_fallback=bool(config.allow_legacy_fallback),
            validation_mode=bool(config.validation_mode),
            production_compatibility_mode=bool(config.production_compatibility_mode),
            enable_tiered_objective=bool(config.enable_tiered_objective),
            min_real_schedule_ratio=float(config.min_real_schedule_ratio),
            max_virtual_chain=int(config.max_virtual_chain),
            max_unbridgeable_drop_ratio=float(config.max_unbridgeable_drop_ratio),
            lot_max_tons=float(config.lot_max_tons),
            sparse_k_same_group=int(config.sparse_k_same_group),
            sparse_k_same_thickness=int(config.sparse_k_same_thickness),
            sparse_k_cross_group=int(config.sparse_k_cross_group),
            sparse_k_due_tight=int(config.sparse_k_due_tight),
            template_min_out_degree=int(config.template_min_out_degree),
            template_min_in_degree=int(config.template_min_in_degree),
            template_top_k=int(config.template_top_k),
            global_prune_max_pairs_per_from=int(config.global_prune_max_pairs_per_from),
            max_routes_per_slot=int(config.max_routes_per_slot),
            strict_virtual_width_levels=bool(config.strict_virtual_width_levels),
            physical_reverse_step_mode=bool(config.physical_reverse_step_mode),
        )
        return PlannerConfig(rule=RuleConfig(max_virtual_chain=model.max_virtual_chain), model=model, score=ScoreConfig())

    raise TypeError(f"Unsupported config type for build_schedule: {type(config)!r}")


def build_schedule(
    orders_path: str,
    steel_info_path: str,
    output_path: str,
    config: SolveConfig | None = None,
    prepared_orders: pd.DataFrame | None = None,
    prepared_dropped: pd.DataFrame | None = None,
    use_dual_master: bool = True,
    enable_rebalance: bool = True,
    engine_used: str = "legacy_fallback",
    equivalence_summary: Dict[str, object] | None = None,
    failure_diagnostics: Dict[str, object] | None = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    兼容层公开入口：
    默认走新 pipeline；仅 legacy 专用参数场景才回落到内部旧实现。
    """
    print("[APS][DEPRECATION] cold_rolling_scheduler.build_schedule 为兼容层，默认转发到 ColdRollingPipeline 联合主模型路径。")

    needs_legacy_impl = (
        prepared_orders is not None
        or prepared_dropped is not None
        or (not bool(use_dual_master))
        or (not bool(enable_rebalance))
    )
    if needs_legacy_impl:
        print("[APS][COMPAT] 检测到 legacy 专用参数，转入内部兼容实现。")
        return _build_schedule_legacy_impl(
            orders_path=orders_path,
            steel_info_path=steel_info_path,
            output_path=output_path,
            config=config,
            prepared_orders=prepared_orders,
            prepared_dropped=prepared_dropped,
            use_dual_master=use_dual_master,
            enable_rebalance=enable_rebalance,
            engine_used=engine_used,
            equivalence_summary=equivalence_summary,
            failure_diagnostics=failure_diagnostics,
        )

    from aps_cp_sat.cold_rolling_pipeline import ColdRollingPipeline
    from aps_cp_sat.domain.models import ColdRollingRequest

    planner_cfg = _to_planner_config(config)
    result = ColdRollingPipeline().run(
        ColdRollingRequest(
            orders_path=Path(orders_path),
            steel_info_path=Path(steel_info_path),
            output_path=Path(output_path),
            config=planner_cfg,
        )
    )
    return result.schedule_df, result.rounds_df
