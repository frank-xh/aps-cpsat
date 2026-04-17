from __future__ import annotations

"""
Export layer contract.

This module only renders already-decoded schedules into files. It must not
change plan semantics, add bridge rows, or repair infeasibility.
"""

from pathlib import Path
from time import perf_counter
from typing import Dict, Iterable, Tuple

import pandas as pd
from openpyxl.utils import get_column_letter

from aps_cp_sat.config.rule_config import RuleConfig
from aps_cp_sat.rules import RULE_REGISTRY, RuleKey


EXPORT_RULE_KEYS = [
    RuleKey.LINE_COMPATIBILITY,
    RuleKey.TEMPERATURE_OVERLAP,
    RuleKey.WIDTH_TRANSITION,
    RuleKey.THICKNESS_TRANSITION,
    RuleKey.CROSS_GROUP_BRIDGE,
    RuleKey.CAMPAIGN_TON_MAX,
    RuleKey.CAMPAIGN_TON_MIN,
    RuleKey.REVERSE_WIDTH_COUNT,
    RuleKey.REVERSE_WIDTH_TOTAL,
    RuleKey.VIRTUAL_USAGE,
    RuleKey.VIRTUAL_RATIO,
]


_LINE_ZH = {"big_roll": "大辊线", "small_roll": "小辊线"}
_EDGE_TYPE_ZH = {"DIRECT_EDGE": "直接相邻", "REAL_BRIDGE_EDGE": "实物桥接", "VIRTUAL_BRIDGE_EDGE": "虚拟桥接"}
_CANDIDATE_STATUS_ZH = {
    "ASSIGNED_CANDIDATE": "候选已分配",
    "DROPPED_CANDIDATE": "候选已剔除",
    "UNROUTABLE_SLOT_MEMBER": "不可路由槽成员",
}
_SUMMARY_MODE_ZH = {"OFFICIAL": "正式口径", "CANDIDATE_ANALYSIS": "候选口径"}
_RESULT_USAGE_ZH = {"OFFICIAL": "正式", "PARTIAL_OFFICIAL": "部分正式", "ANALYSIS_ONLY": "仅分析", "NOT_EXPORTED": "未导出"}
_ACCEPTANCE_ZH = {
    "OFFICIAL_FULL_SCHEDULE": "正式全排结果",
    "PARTIAL_SCHEDULE_WITH_DROPS": "部分可接受(含剔除)",
    "BEST_SEARCH_CANDIDATE_ANALYSIS": "最佳搜索候选(仅分析)",
}
_FAILURE_MODE_ZH = {
    "FAILED_ROUTING_SEARCH": "路由搜索失败",
    "FAILED_STRONG_INFEASIBILITY_SIGNAL": "强不可行信号",
    "FAILED_TIME_BUDGET": "时间预算超限",
    "FAILED_NO_CANDIDATE": "无候选",
    "FAILED_IMPLEMENTATION_ERROR": "实现错误",
}
_VIOLATION_SRC_ZH = {
    "slot_router_diagnostics": "槽位路由诊断",
    "candidate_allocation_only": "候选分配推导",
    "candidate_validate": "候选校验",
    "UNAVAILABLE_FOR_CANDIDATE": "候选不可得",
}
_CONFIDENCE_ZH = {"high": "高", "medium": "中", "low": "低"}


def _zh_enum(value: object, mapping: Dict[str, str]) -> object:
    if value is None or value is pd.NA:
        return value
    s = str(value)
    return mapping.get(s, value)


def _zh_line(value: object) -> object:
    return _zh_enum(value, _LINE_ZH)


def _kv_to_zh(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if list(out.columns) == ["metric", "value"]:
        out = out.rename(columns={"metric": "指标", "value": "值"})
    elif list(out.columns) == ["指标", "值"]:
        return out
    return out


def _campaign_ton_penalty(df: pd.DataFrame, rule: RuleConfig) -> Tuple[int, float]:
    if df.empty or "is_virtual" not in df.columns:
        return 0, 0.0
    real = df[~df["is_virtual"]].copy()
    if real.empty:
        return 0, 0.0
    tons_by = real.groupby(["line", "campaign_id"])["tons"].sum().reset_index(name="campaign_tons")
    low = tons_by[tons_by["campaign_tons"] < float(rule.campaign_ton_min)].copy()
    if low.empty:
        return 0, 0.0
    low["gap_tons"] = float(rule.campaign_ton_min) - low["campaign_tons"]
    return int(len(low)), float(low["gap_tons"].sum())


def _autosize_excel(writer: pd.ExcelWriter) -> None:
    for ws in writer.book.worksheets:
        for idx, col in enumerate(
            ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
            start=1,
        ):
            max_len = 0
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(idx)].width = min(60, max(10, max_len + 2))


def _flatten_dict(data: Dict[str, object], prefix: str = "") -> Iterable[tuple[str, object]]:
    for key, value in data.items():
        full_key = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            yield from _flatten_dict(value, full_key)
        else:
            yield full_key, value


def _num_int(value: object) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        return int(value)
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0


def _num_float(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return float(int(value))
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _build_diagnostics_report(diagnostics: Dict[str, object] | None) -> pd.DataFrame:
    if not isinstance(diagnostics, dict) or not diagnostics:
        return pd.DataFrame(columns=["section", "metric", "value"])
    rows = []
    for section, value in diagnostics.items():
        if isinstance(value, dict):
            for item, item_value in _flatten_dict(value):
                rows.append((str(section), str(item), item_value))
        else:
            rows.append(("general", str(section), value))
    return pd.DataFrame(rows, columns=["section", "metric", "value"])


def _candidate_line_membership_matches(row: pd.Series, line: str) -> bool:
    if str(row.get("line", "")) == str(line):
        return True
    candidate_lines = [v.strip() for v in str(row.get("candidate_lines", "") or "").split(",") if v.strip()]
    return str(line) in candidate_lines


def _build_candidate_line_summary(candidate_schedule_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    candidate_df = candidate_schedule_df.copy() if isinstance(candidate_schedule_df, pd.DataFrame) else pd.DataFrame()
    for line in ["big_roll", "small_roll"]:
        assigned = candidate_df[
            (candidate_df.get("line", pd.Series(dtype=str)) == line)
            & (~candidate_df.get("drop_flag", pd.Series(dtype=bool)).fillna(False))
        ].copy() if not candidate_df.empty else pd.DataFrame()
        dropped = candidate_df[
            candidate_df.apply(lambda r: _candidate_line_membership_matches(r, line), axis=1)
            & candidate_df.get("drop_flag", pd.Series(dtype=bool)).fillna(False)
        ].copy() if not candidate_df.empty else pd.DataFrame()
        slot_sizes = assigned.groupby("slot_no").size() if ("slot_no" in assigned.columns and not assigned.empty) else pd.Series(dtype=int)
        edge_series = assigned.get("selected_edge_type", pd.Series(dtype=str)).astype(str) if not assigned.empty else pd.Series(dtype=str)
        rows.append(
            {
                "line": line,
                "candidate_assigned_orders": int(len(assigned)),
                "candidate_assigned_tons": round(float(assigned["tons"].sum()) if "tons" in assigned.columns else 0.0, 1),
                "candidate_slot_count": int(assigned["slot_no"].nunique()) if "slot_no" in assigned.columns and not assigned.empty else 0,
                "candidate_avg_slot_order_count": round(float(slot_sizes.mean()) if not slot_sizes.empty else 0.0, 2),
                "candidate_max_slot_order_count": int(slot_sizes.max()) if not slot_sizes.empty else 0,
                "candidate_unroutable_slot_count": int(assigned.loc[assigned.get("slot_unroutable_flag", pd.Series(dtype=bool)).fillna(False), "slot_no"].nunique()) if "slot_no" in assigned.columns and not assigned.empty else 0,
                "candidate_dropped_order_count": int(dropped["order_id"].nunique()) if not dropped.empty and "order_id" in dropped.columns else int(len(dropped)),
                "candidate_direct_edge_count": int((edge_series == "DIRECT_EDGE").sum()) if not edge_series.empty else 0,
                "candidate_real_bridge_edge_count": int((edge_series == "REAL_BRIDGE_EDGE").sum()) if not edge_series.empty else 0,
                "candidate_virtual_bridge_edge_count": int((edge_series == "VIRTUAL_BRIDGE_EDGE").sum()) if not edge_series.empty else 0,
            }
        )
    return pd.DataFrame(rows)


def _extract_raw_slots(
    failure_diagnostics: Dict[str, object] | None,
    engine_meta: Dict[str, object] | None,
) -> list[dict]:
    for source in (failure_diagnostics, engine_meta):
        if not isinstance(source, dict):
            continue
        raw_slots = source.get("slot_route_details")
        if isinstance(raw_slots, dict):
            raw_slots = list(raw_slots.values())
        if isinstance(raw_slots, list) and raw_slots:
            return [dict(item) for item in raw_slots if isinstance(item, dict)]
    return []


def _build_candidate_schedule_sheet(candidate_schedule_df: pd.DataFrame, line: str) -> pd.DataFrame:
    candidate_df = candidate_schedule_df.copy() if isinstance(candidate_schedule_df, pd.DataFrame) else pd.DataFrame()
    if candidate_df.empty:
        return pd.DataFrame(
            columns=[
                "order_id",
                "slot_no",
                "candidate_position",
                "candidate_slot_member_index",
                "tons",
                "width",
                "thickness",
                "steel_group",
                "line_capability",
                "candidate_status",
                "slot_unroutable_flag",
                "slot_route_risk_score",
                "selected_edge_type",
                "analysis_only",
                "official_usable",
                "result_usage",
            ]
        )
    out = candidate_df[
        (candidate_df.get("line", pd.Series(dtype=str)) == line)
        & (~candidate_df.get("drop_flag", pd.Series(dtype=bool)).fillna(False))
    ].copy()
    if out.empty:
        return pd.DataFrame(
            columns=[
                "order_id",
                "slot_no",
                "candidate_position",
                "candidate_slot_member_index",
                "tons",
                "width",
                "thickness",
                "steel_group",
                "line_capability",
                "candidate_status",
                "slot_unroutable_flag",
                "slot_route_risk_score",
                "selected_edge_type",
                "analysis_only",
                "official_usable",
                "result_usage",
            ]
        )
    if "candidate_position" not in out.columns:
        out["candidate_position"] = 0
    if "candidate_slot_member_index" not in out.columns:
        out["candidate_slot_member_index"] = out.get("candidate_position", 0)
    out["analysis_only"] = True
    out["official_usable"] = False
    out["result_usage"] = "ANALYSIS_ONLY"
    out = out.sort_values(
        ["slot_no", "candidate_position", "candidate_slot_member_index", "order_id"],
        kind="mergesort",
    ).reset_index(drop=True)
    return out[
        [
            "order_id",
            "slot_no",
            "candidate_position",
            "candidate_slot_member_index",
            "tons",
            "width",
            "thickness",
            "steel_group",
            "line_capability",
            "candidate_status",
            "slot_unroutable_flag",
            "slot_route_risk_score",
            "selected_edge_type",
            "analysis_only",
            "official_usable",
            "result_usage",
        ]
    ]


def _build_slot_frames(raw_slots: list[dict], candidate_schedule_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    slot_rows: list[dict] = []
    unroutable_rows: list[dict] = []
    if raw_slots:
        for item in raw_slots:
            min_width = item.get("min_width", 0) or 0
            max_width = item.get("max_width", 0) or 0
            min_thickness = item.get("min_thickness", 0) or 0
            max_thickness = item.get("max_thickness", 0) or 0
            width_span = item.get("width_span", (float(max_width) - float(min_width)) if max_width or min_width else 0)
            thickness_span = item.get("thickness_span", (float(max_thickness) - float(min_thickness)) if max_thickness or min_thickness else 0)
            row = {
                "line": item.get("line", ""),
                "slot_no": item.get("slot_no", 0),
                "slot_order_count": item.get("order_count", 0),
                "slot_tons": item.get("slot_tons", 0),
                "order_count_over_cap": item.get("order_count_over_cap", 0),
                "slot_route_risk_score": item.get("slot_route_risk_score", 0),
                "pair_gap_proxy": item.get("pair_gap_proxy", 0),
                "span_risk": item.get("span_risk", 0),
                "degree_risk": item.get("degree_risk", 0),
                "isolated_order_penalty": item.get("isolated_order_penalty", 0),
                "dominant_unroutable_reason": item.get("dominant_unroutable_reason", ""),
                "template_coverage_ratio": item.get("template_coverage_ratio", 0),
                "missing_template_edge_count": item.get("missing_template_edge_count", item.get("missing_pair_count", 0)),
                "zero_in_orders": item.get("zero_in_orders", 0),
                "zero_out_orders": item.get("zero_out_orders", 0),
                "width_span": width_span,
                "thickness_span": thickness_span,
                "steel_group_count": item.get("steel_group_count", 0),
                "top_isolated_orders": item.get("top_isolated_orders", []),
                "status": item.get("status", ""),
            }
            slot_rows.append(row)
            if str(item.get("status", "")) == "UNROUTABLE_SLOT":
                unroutable_rows.append(
                    {
                        "line": row["line"],
                        "slot_no": row["slot_no"],
                        "order_count": row["slot_order_count"],
                        "slot_route_risk_score": row["slot_route_risk_score"],
                        "template_coverage_ratio": row["template_coverage_ratio"],
                        "missing_template_edge_count": row["missing_template_edge_count"],
                        "zero_in_orders": row["zero_in_orders"],
                        "zero_out_orders": row["zero_out_orders"],
                        "width_span": row["width_span"],
                        "thickness_span": row["thickness_span"],
                        "steel_group_count": row["steel_group_count"],
                        "dominant_unroutable_reason": row["dominant_unroutable_reason"],
                        "top_isolated_orders": row["top_isolated_orders"],
                    }
                )
    elif isinstance(candidate_schedule_df, pd.DataFrame) and not candidate_schedule_df.empty:
        assigned = candidate_schedule_df[
            (~candidate_schedule_df.get("drop_flag", pd.Series(dtype=bool)).fillna(False))
            & candidate_schedule_df.get("line", pd.Series(dtype=str)).astype(str).ne("")
        ].copy()
        if not assigned.empty:
            for (line, slot_no), grp in assigned.groupby(["line", "slot_no"], dropna=False):
                slot_unroutable = bool(grp.get("slot_unroutable_flag", pd.Series(dtype=bool)).fillna(False).any())
                row = {
                    "line": line,
                    "slot_no": slot_no,
                    "slot_order_count": int(len(grp)),
                    "slot_tons": round(float(grp["tons"].sum()) if "tons" in grp.columns else 0.0, 1),
                    "order_count_over_cap": 0,
                    "slot_route_risk_score": int(grp.get("slot_route_risk_score", pd.Series(dtype=float)).fillna(0).max()) if "slot_route_risk_score" in grp.columns else 0,
                    "pair_gap_proxy": pd.NA,
                    "span_risk": pd.NA,
                    "degree_risk": pd.NA,
                    "isolated_order_penalty": pd.NA,
                    "dominant_unroutable_reason": "UNROUTABLE_SLOT_MEMBER" if slot_unroutable else "",
                    "template_coverage_ratio": pd.NA,
                    "missing_template_edge_count": pd.NA,
                    "zero_in_orders": pd.NA,
                    "zero_out_orders": pd.NA,
                    "width_span": round(float(grp["width"].max() - grp["width"].min()), 1) if "width" in grp.columns else pd.NA,
                    "thickness_span": round(float(grp["thickness"].max() - grp["thickness"].min()), 3) if "thickness" in grp.columns else pd.NA,
                    "steel_group_count": int(grp["steel_group"].astype(str).nunique()) if "steel_group" in grp.columns else pd.NA,
                    "top_isolated_orders": [],
                    "status": "UNROUTABLE_SLOT" if slot_unroutable else "ASSIGNED_CANDIDATE",
                }
                slot_rows.append(row)
                if slot_unroutable:
                    unroutable_rows.append(
                        {
                            "line": row["line"],
                            "slot_no": row["slot_no"],
                            "order_count": row["slot_order_count"],
                            "slot_route_risk_score": row["slot_route_risk_score"],
                            "template_coverage_ratio": row["template_coverage_ratio"],
                            "missing_template_edge_count": row["missing_template_edge_count"],
                            "zero_in_orders": row["zero_in_orders"],
                            "zero_out_orders": row["zero_out_orders"],
                            "width_span": row["width_span"],
                            "thickness_span": row["thickness_span"],
                            "steel_group_count": row["steel_group_count"],
                            "dominant_unroutable_reason": row["dominant_unroutable_reason"],
                            "top_isolated_orders": row["top_isolated_orders"],
                        }
                    )
    slot_summary_df = pd.DataFrame(
        slot_rows,
        columns=[
            "line",
            "slot_no",
            "slot_order_count",
            "slot_tons",
            "order_count_over_cap",
            "slot_route_risk_score",
            "pair_gap_proxy",
            "span_risk",
            "degree_risk",
            "isolated_order_penalty",
            "dominant_unroutable_reason",
        ],
    )
    unroutable_slots_df = pd.DataFrame(
        unroutable_rows,
        columns=[
            "line",
            "slot_no",
            "order_count",
            "template_coverage_ratio",
            "missing_template_edge_count",
            "zero_in_orders",
            "zero_out_orders",
            "width_span",
            "thickness_span",
            "steel_group_count",
            "top_isolated_orders",
            "slot_route_risk_score",
            "dominant_unroutable_reason",
        ],
    )
    return slot_summary_df, unroutable_slots_df


def _build_candidate_violation_summary(
    candidate_schedule_df: pd.DataFrame,
    raw_slots: list[dict],
    validation_summary: dict,
    best_candidate_routing_feasible: bool,
) -> pd.DataFrame:
    candidate_df = candidate_schedule_df.copy() if isinstance(candidate_schedule_df, pd.DataFrame) else pd.DataFrame()
    if raw_slots:
        candidate_unroutable_slot_count = int(len([r for r in raw_slots if str(r.get("status", "")) == "UNROUTABLE_SLOT"]))
        candidate_bad_slot_count = candidate_unroutable_slot_count
        candidate_zero_in = int(sum(int(r.get("zero_in_orders", 0) or 0) for r in raw_slots))
        candidate_zero_out = int(sum(int(r.get("zero_out_orders", 0) or 0) for r in raw_slots))
        candidate_period_reverse_viol = int(sum(int(r.get("period_reverse_count_violation_count", 0) or 0) for r in raw_slots))
        slot_source = "slot_router_diagnostics"
        slot_conf = "high"
    elif not candidate_df.empty:
        assigned = candidate_df[
            (~candidate_df.get("drop_flag", pd.Series(dtype=bool)).fillna(False))
            & candidate_df.get("line", pd.Series(dtype=str)).astype(str).ne("")
        ].copy()
        if not assigned.empty and "slot_no" in assigned.columns:
            slot_unroutable = assigned.get("slot_unroutable_flag", pd.Series(dtype=bool)).fillna(False)
            candidate_unroutable_slot_count = int(assigned.loc[slot_unroutable, "slot_no"].nunique())
            candidate_bad_slot_count = candidate_unroutable_slot_count
        else:
            candidate_unroutable_slot_count = 0
            candidate_bad_slot_count = 0
        # Without per-slot diagnostics we cannot confidently infer zero-in/out counts.
        candidate_zero_in = pd.NA
        candidate_zero_out = pd.NA
        candidate_period_reverse_viol = pd.NA
        slot_source = "candidate_allocation_only"
        slot_conf = "medium"
    else:
        candidate_unroutable_slot_count = 0
        candidate_bad_slot_count = 0
        candidate_zero_in = pd.NA
        candidate_zero_out = pd.NA
        candidate_period_reverse_viol = pd.NA
        slot_source = "UNAVAILABLE_FOR_CANDIDATE"
        slot_conf = "low"
    unavailable = pd.NA
    candidate_direct_reverse = validation_summary.get("direct_reverse_step_violation_count", unavailable) if best_candidate_routing_feasible else unavailable
    candidate_virtual_attach_reverse = validation_summary.get("virtual_attach_reverse_violation_count", unavailable) if best_candidate_routing_feasible else unavailable
    candidate_bridge_count_violation = validation_summary.get("bridge_count_violation_count", unavailable) if best_candidate_routing_feasible else unavailable
    candidate_invalid_virtual_spec = validation_summary.get("invalid_virtual_spec_count", unavailable) if best_candidate_routing_feasible else unavailable
    rows = [
        ("candidate_unroutable_slot_count", _num_int(validation_summary.get("unroutable_slot_count", 0)), candidate_unroutable_slot_count, slot_source, slot_conf),
        ("candidate_bad_slot_count", 0, candidate_bad_slot_count, slot_source, slot_conf),
        ("candidate_zero_in_order_count", 0, candidate_zero_in, slot_source if raw_slots else "UNAVAILABLE_FOR_CANDIDATE", "high" if raw_slots else "low"),
        ("candidate_zero_out_order_count", 0, candidate_zero_out, slot_source if raw_slots else "UNAVAILABLE_FOR_CANDIDATE", "high" if raw_slots else "low"),
        ("candidate_direct_reverse_step_violation_count", _num_int(validation_summary.get("direct_reverse_step_violation_count", 0)), candidate_direct_reverse, "candidate_validate" if best_candidate_routing_feasible else "UNAVAILABLE_FOR_CANDIDATE", "high" if best_candidate_routing_feasible else "low"),
        ("candidate_virtual_attach_reverse_violation_count", _num_int(validation_summary.get("virtual_attach_reverse_violation_count", 0)), candidate_virtual_attach_reverse, "candidate_validate" if best_candidate_routing_feasible else "UNAVAILABLE_FOR_CANDIDATE", "high" if best_candidate_routing_feasible else "low"),
        ("candidate_period_reverse_count_violation_count", _num_int(validation_summary.get("period_reverse_count_violation_count", 0)), candidate_period_reverse_viol, slot_source, slot_conf),
        ("candidate_bridge_count_violation_count", _num_int(validation_summary.get("bridge_count_violation_count", 0)), candidate_bridge_count_violation, "candidate_validate" if best_candidate_routing_feasible else "UNAVAILABLE_FOR_CANDIDATE", "high" if best_candidate_routing_feasible else "low"),
        ("candidate_invalid_virtual_spec_count", _num_int(validation_summary.get("invalid_virtual_spec_count", 0)), candidate_invalid_virtual_spec, "candidate_validate" if best_candidate_routing_feasible else "UNAVAILABLE_FOR_CANDIDATE", "high" if best_candidate_routing_feasible else "low"),
    ]
    return pd.DataFrame(rows, columns=["metric", "official_value", "candidate_value", "violation_count_source", "violation_count_confidence"])


def export_schedule_results(
    final_df: pd.DataFrame,
    rounds_df: pd.DataFrame,
    dropped_df: pd.DataFrame,
    output_path: str,
    input_order_count: int,
    rule: RuleConfig,
    t_start: float | None = None,
    engine_used: str = "unknown",
    fallback_used: bool = False,
    fallback_type: str = "",
    fallback_reason: str = "",
    equivalence_summary: Dict[str, object] | None = None,
    failure_diagnostics: Dict[str, object] | None = None,
    engine_meta: Dict[str, object] | None = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    final_df = final_df.copy()
    rounds_df = rounds_df.copy()
    dropped_df = dropped_df.copy()
    final_df["engine_used"] = str(engine_used)
    rounds_df["engine_used"] = str(engine_used)

    low_cnt, low_gap = _campaign_ton_penalty(final_df, rule)
    diagnostics_report = _build_diagnostics_report(failure_diagnostics)
    fallback_diag = failure_diagnostics.get("fallback", {}) if isinstance(failure_diagnostics, dict) else {}
    em = engine_meta if isinstance(engine_meta, dict) else {}
    main_path = str(fallback_diag.get("main_path", engine_used))
    used_local_routing = bool(fallback_diag.get("used_local_routing", False))
    local_routing_role = str(fallback_diag.get("local_routing_role", "not_used"))
    bridge_modeling = str(fallback_diag.get("bridge_modeling", "template_based"))
    joint_estimates = failure_diagnostics.get("joint_estimates", {}) if isinstance(failure_diagnostics, dict) else {}
    routing_feasible = bool(fallback_diag.get("routing_feasible", False))
    template_pair_ok = bool(fallback_diag.get("template_pair_ok", False))
    adjacency_rule_ok = bool(fallback_diag.get("adjacency_rule_ok", False))
    bridge_expand_ok = bool(fallback_diag.get("bridge_expand_ok", False))
    unroutable_slot_count = int(fallback_diag.get("unroutable_slot_count", 0))
    strict_template_edges_enabled = bool(fallback_diag.get("strict_template_edges_enabled", True))
    result_acceptance_status = str(fallback_diag.get("result_acceptance_status", "UNKNOWN"))
    failure_mode = str(fallback_diag.get("failure_mode", ""))
    evidence_level = str(fallback_diag.get("evidence_level", "OK"))
    top_infeasibility_signals = fallback_diag.get("top_infeasibility_signals", [])
    template_graph_health = str(fallback_diag.get("template_graph_health", "UNKNOWN"))
    precheck_autorelax_applied = bool(fallback_diag.get("precheck_autorelax_applied", False))
    solve_attempt_count = int(fallback_diag.get("solve_attempt_count", 0))
    fallback_attempt_count = int(fallback_diag.get("fallback_attempt_count", 0))
    early_stop_reason = str(fallback_diag.get("early_stop_reason", ""))
    export_failed_result_for_debug = bool(fallback_diag.get("export_failed_result_for_debug", False))
    export_analysis_on_failure = bool(fallback_diag.get("export_analysis_on_failure", False))
    export_best_candidate_analysis = bool(fallback_diag.get("export_best_candidate_analysis", False))
    final_export_performed = bool(fallback_diag.get("final_export_performed", True))
    official_exported = bool(fallback_diag.get("official_exported", routing_feasible))
    analysis_exported = bool(fallback_diag.get("analysis_exported", False))
    result_usage = str(fallback_diag.get("result_usage", "OFFICIAL" if routing_feasible else "UNKNOWN"))
    best_candidate_available = bool(fallback_diag.get("best_candidate_available", False))
    best_candidate_type = str(fallback_diag.get("best_candidate_type", ""))
    best_candidate_objective = float(fallback_diag.get("best_candidate_objective", 0.0) or 0.0)
    best_candidate_search_status = str(fallback_diag.get("best_candidate_search_status", ""))
    best_candidate_routing_feasible = bool(fallback_diag.get("best_candidate_routing_feasible", False))
    best_candidate_unroutable_slot_count = int(fallback_diag.get("best_candidate_unroutable_slot_count", 0) or 0)
    export_consistency_ok = bool(fallback_diag.get("export_consistency_ok", True))
    # Keep console / engine_meta / export consistent: engine_meta is treated as the source of truth when present.
    if isinstance(em, dict) and em:
        result_acceptance_status = str(em.get("result_acceptance_status", result_acceptance_status))
        failure_mode = str(em.get("failure_mode", failure_mode))
        result_usage = str(em.get("result_usage", result_usage))
        best_candidate_available = bool(em.get("best_candidate_available", best_candidate_available))
        best_candidate_type = str(em.get("best_candidate_type", best_candidate_type))
        best_candidate_search_status = str(em.get("best_candidate_search_status", best_candidate_search_status))
        export_consistency_ok = bool(
            export_consistency_ok
            and (str(fallback_diag.get("result_acceptance_status", "")) in {"", str(result_acceptance_status)})
            and (str(fallback_diag.get("failure_mode", "")) in {"", str(failure_mode)})
            and (str(fallback_diag.get("result_usage", "")) in {"", str(result_usage)})
        )
    assignment_pressure_mode = str(fallback_diag.get("assignment_pressure_mode", "normal"))
    template_build_seconds = float(fallback_diag.get("template_build_seconds", 0.0))
    joint_master_seconds = float(fallback_diag.get("joint_master_seconds", 0.0))
    local_router_seconds = float(fallback_diag.get("local_router_seconds", 0.0))
    fallback_total_seconds = float(fallback_diag.get("fallback_total_seconds", 0.0))
    total_run_seconds = float(fallback_diag.get("total_run_seconds", 0.0))
    candidate_schedule_df = em.get("candidate_schedule_df") if isinstance(em.get("candidate_schedule_df"), pd.DataFrame) else pd.DataFrame()
    candidate_big_roll_df = em.get("candidate_big_roll_df") if isinstance(em.get("candidate_big_roll_df"), pd.DataFrame) else pd.DataFrame()
    candidate_small_roll_df = em.get("candidate_small_roll_df") if isinstance(em.get("candidate_small_roll_df"), pd.DataFrame) else pd.DataFrame()
    candidate_line_summary_df = _build_candidate_line_summary(candidate_schedule_df)
    raw_slots = _extract_raw_slots(failure_diagnostics if isinstance(failure_diagnostics, dict) else None, em)
    candidate_backfill_mode = (
        final_df.empty
        and (
            str(result_acceptance_status) == "BEST_SEARCH_CANDIDATE_ANALYSIS"
            or (bool(best_candidate_available) and not bool(official_exported))
        )
        and isinstance(candidate_schedule_df, pd.DataFrame)
        and not candidate_schedule_df.empty
    )
    if final_df.empty and raw_slots:
        computed_unroutable = int(len([r for r in raw_slots if str(r.get("status", "")) == "UNROUTABLE_SLOT"]))
        if computed_unroutable > 0 and unroutable_slot_count <= 0:
            unroutable_slot_count = computed_unroutable

    summary_df = pd.DataFrame(
        [
            ("结果引擎路径", str(engine_used)),
            ("主路径", main_path),
            ("是否触发 fallback", "是" if bool(fallback_used) else "否"),
            ("fallback 类型", str(fallback_type)),
            ("fallback 原因", str(fallback_reason)),
            ("是否使用局部路由", "是" if used_local_routing else "否"),
            ("局部路由角色", local_routing_role),
            ("bridge modeling", bridge_modeling),
            ("routing_feasible", routing_feasible),
            ("template_pair_ok", template_pair_ok),
            ("adjacency_rule_ok", adjacency_rule_ok),
            ("bridge_expand_ok", bridge_expand_ok),
            ("unroutable_slot_count", unroutable_slot_count),
            ("strict_template_edges_enabled", strict_template_edges_enabled),
            ("result_acceptance_status", result_acceptance_status),
            ("failure_mode", failure_mode),
            ("evidence_level", evidence_level),
            ("export_failed_result_for_debug", export_failed_result_for_debug),
            ("export_analysis_on_failure", export_analysis_on_failure),
            ("export_best_candidate_analysis", export_best_candidate_analysis),
            ("final_export_performed", final_export_performed),
            ("official_exported", official_exported),
            ("analysis_exported", analysis_exported),
            ("result_usage", result_usage),
            ("best_candidate_available", best_candidate_available),
            ("best_candidate_type", best_candidate_type),
            ("best_candidate_objective", best_candidate_objective),
            ("best_candidate_search_status", best_candidate_search_status),
            ("best_candidate_routing_feasible", best_candidate_routing_feasible),
            ("best_candidate_unroutable_slot_count", best_candidate_unroutable_slot_count),
            ("template_graph_health", template_graph_health),
            ("precheck_autorelax_applied", precheck_autorelax_applied),
            ("solve_attempt_count", solve_attempt_count),
            ("fallback_attempt_count", fallback_attempt_count),
            ("early_stop_reason", early_stop_reason),
            ("template_build_seconds", template_build_seconds),
            ("joint_master_seconds", joint_master_seconds),
            ("local_router_seconds", local_router_seconds),
            ("fallback_total_seconds", fallback_total_seconds),
            ("total_run_seconds", total_run_seconds),
            ("吨位上限语义", "HARD"),
            ("吨位下限语义", "STRONG_SOFT"),
            ("虚拟占比语义", "STRONG_SOFT"),
            ("逆宽次数/总量语义", "HARD"),
            ("真实订单数", int((~final_df["is_virtual"]).sum()) if "is_virtual" in final_df.columns else int(len(final_df))),
            (f"低吨位轧期数(<{int(rule.campaign_ton_min)})", int(low_cnt)),
            ("低吨位缺口吨位", round(float(low_gap), 1)),
            ("剔除订单数(唯一)", int(dropped_df["order_id"].nunique())) if "order_id" in dropped_df.columns else ("剔除订单数(唯一)", int(len(dropped_df))),
        ],
        columns=["指标", "值"],
    )
    if isinstance(equivalence_summary, dict) and equivalence_summary:
        for k, v in equivalence_summary.items():
            summary_df.loc[len(summary_df)] = (f"等价性.{k}", v)
    if isinstance(top_infeasibility_signals, list):
        for idx, item in enumerate(top_infeasibility_signals[:5], start=1):
            summary_df.loc[len(summary_df)] = (f"不可行信号.top{idx}", item)
    if isinstance(joint_estimates, dict):
        for k, v in joint_estimates.items():
            summary_df.loc[len(summary_df)] = (f"主模型估计.{k}", v)

    # -------------------------------------------------------------------------
    # Constructive LNS path: add solver-specific diagnostics to summary
    # -------------------------------------------------------------------------
    if str(engine_used) == "constructive_lns" or str(main_path) == "constructive_lns":
        lns_engine_meta = em.get("lns_engine_meta", {}) or {}
        lns_diag = em.get("lns_diagnostics", {}) or {}
        cut_diag = lns_diag.get("campaign_cut_diags", {}) if isinstance(lns_diag.get("campaign_cut_diags", {}), dict) else {}
        rounds_summary = lns_diag.get("rounds_summary", {}) or {}

        def _summary_cut_or_em(key: str, default=0):
            return cut_diag.get(key, em.get(key, default))

        # Route C: Read bridge expansion mode for summary
        bridge_expansion_mode = str(em.get("bridge_expansion_mode", lns_engine_meta.get("bridge_expansion_mode", "disabled")))
        virtual_bridge_enabled = bool(em.get("virtual_bridge_edge_enabled_in_constructive", lns_engine_meta.get("virtual_bridge_edge_enabled_in_constructive", False)))
        real_bridge_enabled = bool(em.get("real_bridge_edge_enabled_in_constructive", lns_engine_meta.get("real_bridge_edge_enabled_in_constructive", False)))
        constructive_edge_policy = str(em.get("constructive_edge_policy", lns_engine_meta.get("constructive_edge_policy", "direct_only")))
        bridge_edge_leak_detected = bool(em.get("bridge_edge_leak_detected", False))

        # Build route label string
        if constructive_edge_policy == "direct_only":
            route_label = "路线C(direct_only,禁用所有桥接边)"
        elif constructive_edge_policy == "direct_plus_real_bridge":
            route_label = "路线B(禁用虚拟桥接边,允许实物桥接)"
        elif bridge_expansion_mode == "disabled":
            route_label = "路线B(bridge_expansion=disabled)"
        else:
            route_label = "路线A(支持桥接展开)"

        lns_rows = [
            ("LNS路径.初始合法链数", int(lns_diag.get("constructive_build_diags", {}).get("total_chains", 0) if isinstance(lns_diag.get("constructive_build_diags"), dict) else 0)),
            ("LNS路径.初始切段数", int(lns_diag.get("initial_planned_count", 0))),
            ("LNS路径.最终切段数", int(lns_diag.get("final_planned_count", 0))),
            ("LNS路径.ALNS接受轮数", int(rounds_summary.get("accepted_count", 0))),
            ("LNS路径.ALNS总轮数", int(lns_engine_meta.get("n_rounds_ran", 0))),
            ("LNS路径.ALNS计划轮数", int(lns_engine_meta.get("n_rounds_planned", 0))),
            ("LNS路径.求解状态", str(em.get("lns_status", "UNKNOWN"))),
            ("LNS路径.初始剔除数", int(lns_diag.get("initial_dropped_count", 0))),
            ("LNS路径.最终剔除数", int(lns_diag.get("final_dropped_count", 0))),
            ("LNS路径.剔除变化量", int(lns_diag.get("final_dropped_count", 0) - lns_diag.get("initial_dropped_count", 0))),
            ("LNS路径.订单改善量", int(lns_diag.get("improvement_delta_orders", 0))),
            ("LNS路径.总求解秒数", round(float(lns_engine_meta.get("total_seconds", 0.0)), 2)),
            ("LNS路径.无全局联合模型", bool(lns_engine_meta.get("no_global_joint_model", True))),
            ("LNS路径.无槽位分桶", bool(lns_engine_meta.get("no_slot_bucket", True))),
            ("LNS路径.无非法惩罚", bool(lns_engine_meta.get("no_illegal_penalty", True))),
            # Route C: Edge policy fields
            ("LNS路径.桥接展开模式", bridge_expansion_mode),
            ("LNS路径.边策略", constructive_edge_policy),
            ("LNS路径.虚拟桥接边启用", "是" if virtual_bridge_enabled else "否"),
            ("LNS路径.实物桥接边启用", "是" if real_bridge_enabled else "否"),
            ("LNS路径.求解路线", route_label),
            ("LNS路径.桥接边泄漏", "泄漏!" if bridge_edge_leak_detected else "无泄漏"),
            ("Underfilled重构.尝试", int(_summary_cut_or_em("underfilled_reconstruction_attempts", 0) or 0)),
            ("Underfilled重构.成功", int(_summary_cut_or_em("underfilled_reconstruction_success", 0) or 0)),
            ("Underfilled重构.valid增量", int(_summary_cut_or_em("underfilled_reconstruction_valid_delta", 0) or 0)),
            ("Underfilled重构.underfilled下降", int(_summary_cut_or_em("underfilled_reconstruction_underfilled_delta", 0) or 0)),
            ("Underfilled重构.耗时秒", round(float(_summary_cut_or_em("underfilled_reconstruction_seconds", 0.0) or 0.0), 3)),
            ("Repair实物桥.尝试", int(_summary_cut_or_em("repair_only_real_bridge_attempts", 0) or 0)),
            ("Repair实物桥.成功", int(_summary_cut_or_em("repair_only_real_bridge_success", 0) or 0)),
            ("Repair实物桥.候选总数", int(_summary_cut_or_em("repair_only_real_bridge_candidates_total", 0) or 0)),
            ("Repair实物桥.候选保留", int(_summary_cut_or_em("repair_only_real_bridge_candidates_kept", 0) or 0)),
            ("Repair实物桥.raw行数", int(_summary_cut_or_em("repair_bridge_raw_rows_total", 0) or 0)),
            ("Repair实物桥.matched行数", int(_summary_cut_or_em("repair_bridge_matched_rows_total", 0) or 0)),
            ("Repair实物桥.band启用", "是" if bool(_summary_cut_or_em("repair_bridge_boundary_band_enabled", True)) else "否"),
            ("Repair实物桥.band测试pair数", int(_summary_cut_or_em("repair_bridge_band_pairs_tested", 0) or 0)),
            ("Repair实物桥.band命中数", int(_summary_cut_or_em("repair_bridge_band_hits", 0) or 0)),
            ("Repair实物桥.单点命中数", int(_summary_cut_or_em("repair_bridge_single_point_hits", 0) or 0)),
            ("Repair实物桥.band-only命中数", int(_summary_cut_or_em("repair_bridge_band_only_hits", 0) or 0)),
            ("Repair实物桥.band最优距离", int(_summary_cut_or_em("repair_bridge_band_best_distance", -1) or -1)),
            ("Repair实物桥.端点微调启用", "是" if bool(_summary_cut_or_em("repair_bridge_endpoint_adjustment_enabled", True)) else "否"),
            ("Repair实物桥.微调方案数", int(_summary_cut_or_em("repair_bridge_adjustments_generated", 0) or 0)),
            ("Repair实物桥.微调测试pair数", int(_summary_cut_or_em("repair_bridge_adjustment_pairs_tested", 0) or 0)),
            ("Repair实物桥.微调命中数", int(_summary_cut_or_em("repair_bridge_adjustment_hits", 0) or 0)),
            ("Repair实物桥.仅微调命中数", int(_summary_cut_or_em("repair_bridge_adjustment_only_hits", 0) or 0)),
            ("Repair实物桥.最优微调代价", int(_summary_cut_or_em("repair_bridge_best_adjustment_cost", -1) or -1)),
            ("Repair实物桥.candidate命中", int(_summary_cut_or_em("repair_bridge_candidates_matched", 0) or 0)),
            ("Repair实物桥.candidate-pair非法拒绝", int(_summary_cut_or_em("repair_bridge_candidates_rejected_pair_invalid", 0) or 0)),
            ("Repair实物桥.candidate-吨位拒绝", int(_summary_cut_or_em("repair_bridge_candidates_rejected_ton_invalid", 0) or 0)),
            ("Repair实物桥.candidate-score拒绝", int(_summary_cut_or_em("repair_bridge_candidates_rejected_score_worse", 0) or 0)),
            ("Repair实物桥.candidate接受", int(_summary_cut_or_em("repair_bridge_candidates_accepted", 0) or 0)),
            ("Repair实物桥.精确非法pair数", int(_summary_cut_or_em("repair_bridge_exact_invalid_pair_count", 0) or 0)),
            ("Repair实物桥.frontier错位数", int(_summary_cut_or_em("repair_bridge_frontier_mismatch_count", 0) or 0)),
            ("Repair实物桥.宽度非法pair数", int(_summary_cut_or_em("repair_bridge_pair_invalid_width", 0) or 0)),
            ("Repair实物桥.厚度非法pair数", int(_summary_cut_or_em("repair_bridge_pair_invalid_thickness", 0) or 0)),
            ("Repair实物桥.温度非法pair数", int(_summary_cut_or_em("repair_bridge_pair_invalid_temp", 0) or 0)),
            ("Repair实物桥.钢种组非法pair数", int(_summary_cut_or_em("repair_bridge_pair_invalid_group", 0) or 0)),
            ("Repair实物桥.未知非法pair数", int(_summary_cut_or_em("repair_bridge_pair_invalid_unknown", 0) or 0)),
            ("Repair实物桥.吨位救援尝试", int(_summary_cut_or_em("repair_bridge_ton_rescue_attempts", 0) or 0)),
            ("Repair实物桥.吨位救援成功", int(_summary_cut_or_em("repair_bridge_ton_rescue_success", 0) or 0)),
            ("Repair实物桥.吨位救援窗口数", int(_summary_cut_or_em("repair_bridge_ton_rescue_windows_tested", 0) or 0)),
            ("Repair实物桥.吨位救援valid增量", int(_summary_cut_or_em("repair_bridge_ton_rescue_valid_delta", 0) or 0)),
            ("Repair实物桥.吨位救援underfilled增量", int(_summary_cut_or_em("repair_bridge_ton_rescue_underfilled_delta", 0) or 0)),
            ("Repair实物桥.吨位救援订单增量", int(_summary_cut_or_em("repair_bridge_ton_rescue_scheduled_orders_delta", 0) or 0)),
            ("Repair实物桥.当前块吨位不足", int(_summary_cut_or_em("repair_bridge_filtered_ton_below_min_current_block", 0) or 0)),
            ("Repair实物桥.扩邻后仍吨位不足", int(_summary_cut_or_em("repair_bridge_filtered_ton_below_min_even_after_neighbor_expansion", 0) or 0)),
            ("Repair实物桥.扩邻后无法切段", int(_summary_cut_or_em("repair_bridge_filtered_ton_split_not_found", 0) or 0)),
            ("Repair实物桥.吨位救援无收益", int(_summary_cut_or_em("repair_bridge_filtered_ton_rescue_no_gain", 0) or 0)),
            ("Repair实物桥.key不匹配次数", int(_summary_cut_or_em("repair_bridge_endpoint_key_mismatch_count", 0) or 0)),
            ("Repair实物桥字段不匹配次数", int(_summary_cut_or_em("repair_bridge_field_name_mismatch_count", 0) or 0)),
            ("Repair实物桥pack不一致次数", int(_summary_cut_or_em("repair_bridge_inconsistency_count", 0) or 0)),
            ("Repair实物桥.耗时秒", round(float(_summary_cut_or_em("repair_only_real_bridge_seconds", 0.0) or 0.0), 3)),
        ]
        for row in lns_rows:
            summary_df.loc[len(summary_df)] = row

    runtime_rows = [
        # ---- Run path fingerprints ----
        ("运行路径.pipeline", str(em.get("run_path_fingerprint_pipeline", ""))),
        ("运行路径.constructive_builder", str(em.get("run_path_fingerprint_constructive_builder", ""))),
        ("运行路径.campaign_cutter", str(em.get("run_path_fingerprint_campaign_cutter", ""))),
        ("运行路径.constructive_lns_master", str(em.get("run_path_fingerprint_constructive_lns_master", ""))),
        ("结果引擎路径", str(engine_used)),
        ("主路径", main_path),
        ("是否触发 fallback", "是" if bool(fallback_used) else "否"),
        ("fallback 类型", str(fallback_type)),
        ("fallback 原因", str(fallback_reason)),
        ("是否使用局部路由", "是" if used_local_routing else "否"),
        ("局部路由角色", local_routing_role),
        ("bridge modeling", bridge_modeling),
        ("routing_feasible", routing_feasible),
        ("template_pair_ok", template_pair_ok),
        ("adjacency_rule_ok", adjacency_rule_ok),
        ("bridge_expand_ok", bridge_expand_ok),
        ("unroutable_slot_count", unroutable_slot_count),
        ("strict_template_edges_enabled", strict_template_edges_enabled),
        ("result_acceptance_status", result_acceptance_status),
        ("failure_mode", failure_mode),
        ("evidence_level", evidence_level),
        ("export_failed_result_for_debug", export_failed_result_for_debug),
        ("export_analysis_on_failure", export_analysis_on_failure),
        ("export_best_candidate_analysis", export_best_candidate_analysis),
        ("final_export_performed", final_export_performed),
        ("official_exported", official_exported),
        ("analysis_exported", analysis_exported),
        ("result_usage", result_usage),
        ("best_candidate_available", best_candidate_available),
        ("best_candidate_type", best_candidate_type),
        ("best_candidate_objective", best_candidate_objective),
        ("best_candidate_search_status", best_candidate_search_status),
        ("best_candidate_routing_feasible", best_candidate_routing_feasible),
        ("best_candidate_unroutable_slot_count", best_candidate_unroutable_slot_count),
        ("export_consistency_ok", export_consistency_ok),
            ("template_build_seconds", template_build_seconds),
            ("joint_master_seconds", joint_master_seconds),
            ("local_router_seconds", local_router_seconds),
            ("fallback_total_seconds", fallback_total_seconds),
            ("total_run_seconds", total_run_seconds),
            ("输入订单数", int(input_order_count)),
        ("输出真实订单数", int((~final_df["is_virtual"]).sum()) if "is_virtual" in final_df.columns else int(len(final_df))),
        ("输出虚拟板坯数", int(final_df["is_virtual"].sum()) if "is_virtual" in final_df.columns else 0),
        ("低吨位轧期数", int(low_cnt)),
        ("低吨位缺口吨位", round(float(low_gap), 1)),
        (f"剔除订单数(唯一订单)", int(dropped_df["order_id"].nunique())) if "order_id" in dropped_df.columns else ("剔除订单数(唯一订单)", int(len(dropped_df))),
    ]
    # -------------------------------------------------------------------------
    # Constructive LNS path: add to runtime rows
    # -------------------------------------------------------------------------
    # Route C: Read bridge expansion mode for runtime summary
    bridge_expansion_mode = str(em.get("bridge_expansion_mode", em.get("lns_engine_meta", {}).get("bridge_expansion_mode", "disabled")))
    virtual_bridge_enabled = bool(em.get("virtual_bridge_edge_enabled_in_constructive", em.get("lns_engine_meta", {}).get("virtual_bridge_edge_enabled_in_constructive", False)))
    real_bridge_enabled = bool(em.get("real_bridge_edge_enabled_in_constructive", em.get("lns_engine_meta", {}).get("real_bridge_edge_enabled_in_constructive", False)))
    constructive_edge_policy = str(em.get("constructive_edge_policy", em.get("lns_engine_meta", {}).get("constructive_edge_policy", "direct_only")))
    bridge_edge_leak_detected = bool(em.get("bridge_edge_leak_detected", False))

    # Build route label
    if constructive_edge_policy == "direct_only":
        route_label = "路线C(direct_only,禁用所有桥接边)"
    elif constructive_edge_policy == "direct_plus_real_bridge":
        route_label = "路线B(禁用虚拟桥接边,允许实物桥接)"
    elif bridge_expansion_mode == "disabled":
        route_label = "路线B(bridge_expansion=disabled)"
    else:
        route_label = "路线A(支持桥接展开)"

    if str(engine_used) == "constructive_lns" or str(main_path) == "constructive_lns":
        lns_engine_meta = em.get("lns_engine_meta", {}) or {}
        lns_diag = em.get("lns_diagnostics", {}) or {}
        cut_diag = lns_diag.get("campaign_cut_diags", {}) if isinstance(lns_diag.get("campaign_cut_diags", {}), dict) else {}
        rounds_summary = lns_diag.get("rounds_summary", {}) or {}

        def _cut_or_em(key: str, default=0):
            return cut_diag.get(key, em.get(key, default))

        runtime_rows.extend([
            ("LNS.ALNS接受轮数", int(rounds_summary.get("accepted_count", 0))),
            ("LNS.ALNS总轮数", int(lns_engine_meta.get("n_rounds_ran", 0))),
            ("LNS.求解状态", str(em.get("lns_status", "UNKNOWN"))),
            ("LNS.初始合法链数", int(lns_diag.get("constructive_build_diags", {}).get("total_chains", 0) if isinstance(lns_diag.get("constructive_build_diags"), dict) else 0)),
            ("LNS.初始切段数", int(lns_diag.get("initial_planned_count", 0))),
            ("LNS.最终切段数", int(lns_diag.get("final_planned_count", 0))),
            ("LNS.剔除变化量", int(lns_diag.get("final_dropped_count", 0) - lns_diag.get("initial_dropped_count", 0))),
            ("LNS.总求解秒数", round(float(lns_engine_meta.get("total_seconds", 0.0)), 2)),
            # Route C: Edge policy fields for runtime summary
            ("LNS.桥接展开模式", bridge_expansion_mode),
            ("LNS.边策略", constructive_edge_policy),
            ("LNS.虚拟桥接边启用", "是" if virtual_bridge_enabled else "否"),
            ("LNS.实物桥接边启用", "是" if real_bridge_enabled else "否"),
            ("LNS.求解路线", route_label),
            ("LNS.桥接边泄漏", "泄漏!" if bridge_edge_leak_detected else "无泄漏"),
            # ---- Decode order integrity gate (fail-fast from constructive_lns decode path) ----
            ("LNS.decode_order_integrity_ok", bool(em.get("decode_order_integrity_ok", True))),
            ("LNS.decode_order_mismatch_count", int(em.get("decode_order_mismatch_count", 0))),
            ("LNS.decode_demoted_order_count", int(em.get("decode_demoted_order_count", 0))),
            ("LNS.decode_order_mismatch_campaigns", str(em.get("decode_order_mismatch_campaigns", []))),
            # ---- Tail repair attempt/success counts ----
            ("Tail.RECUT尝试", int(lns_diag.get("tail_repair_recut_attempts", lns_diag.get("campaign_cut_diags", {}).get("tail_rebalance_summary", {}).get("tail_repair_recut_attempts", 0) if isinstance(lns_diag.get("campaign_cut_diags"), dict) else 0))),
            ("Tail.RECUT成功", int(lns_diag.get("tail_repair_recut_success", lns_diag.get("campaign_cut_diags", {}).get("tail_rebalance_summary", {}).get("tail_repair_recut_success", 0) if isinstance(lns_diag.get("campaign_cut_diags"), dict) else 0))),
            ("Tail.SHIFT尝试", int(lns_diag.get("tail_repair_shift_attempts", lns_diag.get("campaign_cut_diags", {}).get("tail_rebalance_summary", {}).get("tail_repair_shift_attempts", 0) if isinstance(lns_diag.get("campaign_cut_diags"), dict) else 0))),
            ("Tail.SHIFT成功", int(lns_diag.get("tail_repair_shift_success", lns_diag.get("campaign_cut_diags", {}).get("tail_rebalance_summary", {}).get("tail_repair_shift_success", 0) if isinstance(lns_diag.get("campaign_cut_diags"), dict) else 0))),
            ("Tail.FILL尝试", int(lns_diag.get("tail_repair_fill_attempts", lns_diag.get("campaign_cut_diags", {}).get("tail_rebalance_summary", {}).get("tail_repair_fill_attempts", 0) if isinstance(lns_diag.get("campaign_cut_diags"), dict) else 0))),
            ("Tail.FILL成功", int(lns_diag.get("tail_repair_fill_success", lns_diag.get("campaign_cut_diags", {}).get("tail_rebalance_summary", {}).get("tail_repair_fill_success", 0) if isinstance(lns_diag.get("campaign_cut_diags"), dict) else 0))),
            ("Tail.MERGE尝试", int(lns_diag.get("tail_repair_merge_attempts", lns_diag.get("campaign_cut_diags", {}).get("tail_rebalance_summary", {}).get("tail_repair_merge_attempts", 0) if isinstance(lns_diag.get("campaign_cut_diags"), dict) else 0))),
            ("Tail.MERGE成功", int(lns_diag.get("tail_repair_merge_success", lns_diag.get("campaign_cut_diags", {}).get("tail_rebalance_summary", {}).get("tail_repair_merge_success", 0) if isinstance(lns_diag.get("campaign_cut_diags"), dict) else 0))),
            ("Underfilled重构.尝试", int(cut_diag.get("underfilled_reconstruction_attempts", 0))),
            ("Underfilled重构.成功", int(cut_diag.get("underfilled_reconstruction_success", 0))),
            ("Underfilled重构.测试block数", int(cut_diag.get("underfilled_reconstruction_blocks_tested", 0))),
            ("Underfilled重构.跳过block数", int(_cut_or_em("underfilled_reconstruction_blocks_skipped", 0) or 0)),
            ("Underfilled重构.valid_before", int(_cut_or_em("underfilled_reconstruction_valid_before", 0) or 0)),
            ("Underfilled重构.valid_after", int(_cut_or_em("underfilled_reconstruction_valid_after", 0) or 0)),
            ("Underfilled重构.underfilled_before", int(_cut_or_em("underfilled_reconstruction_underfilled_before", 0) or 0)),
            ("Underfilled重构.underfilled_after", int(_cut_or_em("underfilled_reconstruction_underfilled_after", 0) or 0)),
            ("Underfilled重构.救回segment数", int(cut_diag.get("underfilled_reconstruction_segments_salvaged", 0))),
            ("Underfilled重构.救回订单数", int(cut_diag.get("underfilled_reconstruction_orders_salvaged", 0))),
            ("Underfilled重构.valid增量", int(cut_diag.get("underfilled_reconstruction_valid_delta", 0))),
            ("Underfilled重构.underfilled下降", int(cut_diag.get("underfilled_reconstruction_underfilled_delta", 0))),
            ("Underfilled重构.未进入原因", str(_cut_or_em("underfilled_reconstruction_not_entered_reason", ""))),
            ("Underfilled重构.耗时秒", round(float(_cut_or_em("underfilled_reconstruction_seconds", 0.0) or 0.0), 3)),
            ("Repair实物桥.尝试", int(cut_diag.get("repair_only_real_bridge_attempts", 0))),
            ("Repair实物桥.成功", int(cut_diag.get("repair_only_real_bridge_success", 0))),
            ("Repair实物桥.候选总数", int(_cut_or_em("repair_only_real_bridge_candidates_total", 0) or 0)),
            ("Repair实物桥.候选保留", int(_cut_or_em("repair_only_real_bridge_candidates_kept", 0) or 0)),
            ("Repair实物桥.pack类型", str(_cut_or_em("repair_bridge_pack_type", ""))),
            ("Repair实物桥.pack含real行", "是" if bool(_cut_or_em("repair_bridge_pack_has_real_rows", False)) else "否"),
            ("Repair实物桥.pack real总行数", int(_cut_or_em("repair_bridge_pack_real_rows_total", 0) or 0)),
            ("Repair实物桥.pack virtual总行数", int(_cut_or_em("repair_bridge_pack_virtual_rows_total", 0) or 0)),
            ("Repair实物桥.raw行数", int(_cut_or_em("repair_bridge_raw_rows_total", 0) or 0)),
            ("Repair实物桥.matched行数", int(_cut_or_em("repair_bridge_matched_rows_total", 0) or 0)),
            ("Repair实物桥.kept行数", int(_cut_or_em("repair_bridge_kept_rows_total", 0) or 0)),
            ("Repair实物桥.band启用", "是" if bool(_cut_or_em("repair_bridge_boundary_band_enabled", True)) else "否"),
            ("Repair实物桥.band测试pair数", int(_cut_or_em("repair_bridge_band_pairs_tested", 0) or 0)),
            ("Repair实物桥.band命中数", int(_cut_or_em("repair_bridge_band_hits", 0) or 0)),
            ("Repair实物桥.单点命中数", int(_cut_or_em("repair_bridge_single_point_hits", 0) or 0)),
            ("Repair实物桥.band-only命中数", int(_cut_or_em("repair_bridge_band_only_hits", 0) or 0)),
            ("Repair实物桥.band最优距离", int(_cut_or_em("repair_bridge_band_best_distance", -1) or -1)),
            ("Repair实物桥.端点微调启用", "是" if bool(_cut_or_em("repair_bridge_endpoint_adjustment_enabled", True)) else "否"),
            ("Repair实物桥.微调方案数", int(_cut_or_em("repair_bridge_adjustments_generated", 0) or 0)),
            ("Repair实物桥.微调测试pair数", int(_cut_or_em("repair_bridge_adjustment_pairs_tested", 0) or 0)),
            ("Repair实物桥.微调命中数", int(_cut_or_em("repair_bridge_adjustment_hits", 0) or 0)),
            ("Repair实物桥.仅微调命中数", int(_cut_or_em("repair_bridge_adjustment_only_hits", 0) or 0)),
            ("Repair实物桥.最优微调代价", int(_cut_or_em("repair_bridge_best_adjustment_cost", -1) or -1)),
            ("Repair实物桥.candidate命中", int(_cut_or_em("repair_bridge_candidates_matched", 0) or 0)),
            ("Repair实物桥.candidate-pair非法拒绝", int(_cut_or_em("repair_bridge_candidates_rejected_pair_invalid", 0) or 0)),
            ("Repair实物桥.candidate-吨位拒绝", int(_cut_or_em("repair_bridge_candidates_rejected_ton_invalid", 0) or 0)),
            ("Repair实物桥.candidate-score拒绝", int(_cut_or_em("repair_bridge_candidates_rejected_score_worse", 0) or 0)),
            ("Repair实物桥.candidate接受", int(_cut_or_em("repair_bridge_candidates_accepted", 0) or 0)),
            ("Repair实物桥.精确非法pair数", int(_cut_or_em("repair_bridge_exact_invalid_pair_count", 0) or 0)),
            ("Repair实物桥.frontier错位数", int(_cut_or_em("repair_bridge_frontier_mismatch_count", 0) or 0)),
            ("Repair实物桥.宽度非法pair数", int(_cut_or_em("repair_bridge_pair_invalid_width", 0) or 0)),
            ("Repair实物桥.厚度非法pair数", int(_cut_or_em("repair_bridge_pair_invalid_thickness", 0) or 0)),
            ("Repair实物桥.温度非法pair数", int(_cut_or_em("repair_bridge_pair_invalid_temp", 0) or 0)),
            ("Repair实物桥.钢种组非法pair数", int(_cut_or_em("repair_bridge_pair_invalid_group", 0) or 0)),
            ("Repair实物桥.未知非法pair数", int(_cut_or_em("repair_bridge_pair_invalid_unknown", 0) or 0)),
            ("Repair实物桥.吨位救援尝试", int(_cut_or_em("repair_bridge_ton_rescue_attempts", 0) or 0)),
            ("Repair实物桥.吨位救援成功", int(_cut_or_em("repair_bridge_ton_rescue_success", 0) or 0)),
            ("Repair实物桥.吨位救援窗口数", int(_cut_or_em("repair_bridge_ton_rescue_windows_tested", 0) or 0)),
            ("Repair实物桥.吨位救援valid增量", int(_cut_or_em("repair_bridge_ton_rescue_valid_delta", 0) or 0)),
            ("Repair实物桥.吨位救援underfilled增量", int(_cut_or_em("repair_bridge_ton_rescue_underfilled_delta", 0) or 0)),
            ("Repair实物桥.吨位救援订单增量", int(_cut_or_em("repair_bridge_ton_rescue_scheduled_orders_delta", 0) or 0)),
            ("Repair实物桥.当前块吨位不足", int(_cut_or_em("repair_bridge_filtered_ton_below_min_current_block", 0) or 0)),
            ("Repair实物桥.扩邻后仍吨位不足", int(_cut_or_em("repair_bridge_filtered_ton_below_min_even_after_neighbor_expansion", 0) or 0)),
            ("Repair实物桥.扩邻后无法切段", int(_cut_or_em("repair_bridge_filtered_ton_split_not_found", 0) or 0)),
            ("Repair实物桥.吨位救援无收益", int(_cut_or_em("repair_bridge_filtered_ton_rescue_no_gain", 0) or 0)),
            ("Repair实物桥.endpoint key不匹配", int(_cut_or_em("repair_bridge_endpoint_key_mismatch_count", 0) or 0)),
            ("Repair实物桥.字段名不匹配", int(_cut_or_em("repair_bridge_field_name_mismatch_count", 0) or 0)),
            ("Repair实物桥.pack不一致", int(_cut_or_em("repair_bridge_inconsistency_count", 0) or 0)),
            ("Repair实物桥.过滤-direct已可行", int(_cut_or_em("repair_only_real_bridge_filtered_direct_feasible", 0) or 0)),
            ("Repair实物桥.过滤-pair非法", int(_cut_or_em("repair_only_real_bridge_filtered_pair_invalid", 0) or 0)),
            ("Repair实物桥.过滤-吨位非法", int(_cut_or_em("repair_only_real_bridge_filtered_ton_invalid", 0) or 0)),
            ("Repair实物桥.过滤-score较差", int(_cut_or_em("repair_only_real_bridge_filtered_score_worse", 0) or 0)),
            ("Repair实物桥.过滤-桥数量超限", int(_cut_or_em("repair_only_real_bridge_filtered_bridge_limit_exceeded", 0) or 0)),
            ("Repair实物桥.过滤-多重集非法", int(_cut_or_em("repair_only_real_bridge_filtered_multiplicity_invalid", 0) or 0)),
            ("Repair实物桥.过滤-非实物桥", int(_cut_or_em("repair_only_real_bridge_filtered_bridge_path_not_real", 0) or 0)),
            ("Repair实物桥.过滤-桥路径缺失", int(_cut_or_em("repair_only_real_bridge_filtered_bridge_path_missing", 0) or 0)),
            ("Repair实物桥.过滤-block订单不一致", int(_cut_or_em("repair_only_real_bridge_filtered_block_order_mismatch", 0) or 0)),
            ("Repair实物桥.过滤-line不匹配", int(_cut_or_em("repair_only_real_bridge_filtered_line_mismatch", 0) or 0)),
            ("Repair实物桥.过滤-block成员不匹配", int(_cut_or_em("repair_only_real_bridge_filtered_block_membership_mismatch", 0) or 0)),
            ("Repair实物桥.过滤-path载荷为空", int(_cut_or_em("repair_only_real_bridge_filtered_bridge_path_payload_empty", 0) or 0)),
            ("Repair实物桥.使用segment数", int(cut_diag.get("repair_only_real_bridge_used_segments", 0))),
            ("Repair实物桥.救回订单数", int(cut_diag.get("repair_only_real_bridge_used_orders", 0))),
            ("Repair实物桥.未进入原因", str(_cut_or_em("repair_only_real_bridge_not_entered_reason", ""))),
            ("Repair实物桥.耗时秒", round(float(_cut_or_em("repair_only_real_bridge_seconds", 0.0) or 0.0), 3)),
            # ---- Final segment salvage diagnostics ----
            ("Salvage.尝试次数", int(lns_diag.get("final_segment_salvage_attempts", 0))),
            ("Salvage.成功次数", int(lns_diag.get("final_segment_salvage_success_count", 0))),
            ("Salvage.保留clean片数", int(lns_diag.get("final_segment_salvaged_piece_count", 0))),
            ("Salvage.降级fragments数", int(lns_diag.get("final_segment_demoted_fragment_count", 0))),
            ("Salvage.整段丢弃次数", int(lns_diag.get("final_segment_full_drop_count", 0))),
        ])

    if isinstance(failure_diagnostics, dict):
        runtime_rows.extend((f"诊断.{k}", v) for k, v in _flatten_dict(failure_diagnostics))
    runtime_report = pd.DataFrame(runtime_rows, columns=["指标", "值"])

    equivalence_report = pd.DataFrame(columns=["检查项", "值"])
    if isinstance(equivalence_summary, dict) and equivalence_summary:
        equivalence_report = pd.DataFrame([(str(k), v) for k, v in equivalence_summary.items()], columns=["检查项", "值"])

    # ---- 统一初始化 campaign_slot 一致性诊断变量（提前初始化，确保 diagnostics_snapshot / engine_meta 注入可用） ----
    campaign_slot_consistency_ok = True
    campaign_slot_fixed_count = 0
    campaign_slot_fix_warnings: list[str] = []

    diagnostics_snapshot = pd.DataFrame(
        [
            ("main_path", main_path),
            ("engine_used", str(engine_used)),
            ("fallback_used", bool(fallback_used)),
            ("fallback_type", str(fallback_type)),
            ("fallback_reason", str(fallback_reason)),
            ("assignment_pressure_mode", assignment_pressure_mode),
            ("used_local_routing", bool(used_local_routing)),
            ("local_routing_role", str(local_routing_role)),
            ("bridge_modeling", str(bridge_modeling)),
            ("result_acceptance_status", result_acceptance_status),
            ("failure_mode", failure_mode),
            ("evidence_level", evidence_level),
            ("export_failed_result_for_debug", export_failed_result_for_debug),
            ("export_analysis_on_failure", export_analysis_on_failure),
            ("export_best_candidate_analysis", export_best_candidate_analysis),
            ("final_export_performed", final_export_performed),
            ("official_exported", official_exported),
            ("analysis_exported", analysis_exported),
            ("result_usage", result_usage),
            ("best_candidate_available", best_candidate_available),
            ("best_candidate_type", best_candidate_type),
            ("best_candidate_objective", best_candidate_objective),
            ("best_candidate_search_status", best_candidate_search_status),
            ("best_candidate_routing_feasible", best_candidate_routing_feasible),
            ("best_candidate_unroutable_slot_count", best_candidate_unroutable_slot_count),
            ("export_consistency_ok", export_consistency_ok),
            ("template_graph_health", template_graph_health),
            ("precheck_autorelax_applied", precheck_autorelax_applied),
            ("solve_attempt_count", solve_attempt_count),
            ("fallback_attempt_count", fallback_attempt_count),
            ("early_stop_reason", early_stop_reason),
            ("template_build_seconds", template_build_seconds),
            ("joint_master_seconds", joint_master_seconds),
            ("local_router_seconds", local_router_seconds),
            ("fallback_total_seconds", fallback_total_seconds),
            ("total_run_seconds", total_run_seconds),
            # ---- Decode order integrity gate (fail-fast from constructive_lns decode path) ----
            ("decode_order_integrity_ok", bool(em.get("decode_order_integrity_ok", True))),
            ("decode_order_mismatch_count", int(em.get("decode_order_mismatch_count", 0))),
            ("decode_demoted_order_count", int(em.get("decode_demoted_order_count", 0))),
            ("decode_order_mismatch_campaigns", str(em.get("decode_order_mismatch_campaigns", []))),
            # ---- campaign_slot_no 一致性诊断 ----
            ("campaign_slot_no_consistency_ok", campaign_slot_consistency_ok),
            ("campaign_slot_no_groups_fixed_count", campaign_slot_fixed_count),
            # ---- Run path fingerprints ----
            ("run_path_fingerprint_pipeline", str(em.get("run_path_fingerprint_pipeline", ""))),
            ("run_path_fingerprint_constructive_builder", str(em.get("run_path_fingerprint_constructive_builder", ""))),
            ("run_path_fingerprint_campaign_cutter", str(em.get("run_path_fingerprint_campaign_cutter", ""))),
            ("run_path_fingerprint_constructive_lns_master", str(em.get("run_path_fingerprint_constructive_lns_master", ""))),
            # ---- Tail repair diagnostics ----
            ("tail_repair_recut_attempts", int(em.get("tail_repair_recut_attempts", 0))),
            ("tail_repair_recut_success", int(em.get("tail_repair_recut_success", 0))),
            ("tail_repair_shift_attempts", int(em.get("tail_repair_shift_attempts", 0))),
            ("tail_repair_shift_success", int(em.get("tail_repair_shift_success", 0))),
            ("tail_repair_fill_attempts", int(em.get("tail_repair_fill_attempts", 0))),
            ("tail_repair_fill_success", int(em.get("tail_repair_fill_success", 0))),
            ("tail_repair_merge_attempts", int(em.get("tail_repair_merge_attempts", 0))),
            ("tail_repair_merge_success", int(em.get("tail_repair_merge_success", 0))),
            # ---- Final segment salvage diagnostics ----
            ("final_segment_salvage_attempts", int(em.get("final_segment_salvage_attempts", 0))),
            ("final_segment_salvage_success_count", int(em.get("final_segment_salvage_success_count", 0))),
            ("final_segment_salvaged_piece_count", int(em.get("final_segment_salvaged_piece_count", 0))),
            ("final_segment_demoted_fragment_count", int(em.get("final_segment_demoted_fragment_count", 0))),
            ("final_segment_full_drop_count", int(em.get("final_segment_full_drop_count", 0))),
        ],
        columns=["key", "value"],
    )

    # ---- 扩展 engine_meta 注入 campaign_slot 诊断 ----
    if isinstance(em, dict):
        em["campaign_slot_no_consistency_ok"] = campaign_slot_consistency_ok
        em["campaign_slot_no_groups_fixed_count"] = campaign_slot_fixed_count
        if campaign_slot_fix_warnings:
            em["campaign_slot_fix_warnings"] = campaign_slot_fix_warnings[:10]  # 最多保留10条

    precheck_summary_df = pd.DataFrame(
        [
            ("total_orders", int(input_order_count)),
            ("total_tons", round(float(final_df["tons"].sum() + dropped_df["tons"].sum()) if "tons" in final_df.columns and "tons" in dropped_df.columns else 0.0, 1)),
            ("line_capability_histogram", str((dropped_df.get("line_capability", pd.Series(dtype=str)).value_counts().to_dict() if "line_capability" in dropped_df.columns else {}))),
            ("suspicious_data_count", int(fallback_diag.get("suspicious_data_count", 0) if isinstance(fallback_diag, dict) else 0)),
        ],
        columns=["metric", "value"],
    )

    feasibility_evidence_df = pd.DataFrame(columns=["metric", "value"])
    if isinstance(failure_diagnostics, dict) and isinstance(failure_diagnostics.get("feasibility_evidence"), dict):
        fe = failure_diagnostics["feasibility_evidence"]
        rows = [("evidence_level", failure_diagnostics.get("fallback", {}).get("evidence_level", "OK"))]
        for key, value in _flatten_dict(fe):
            rows.append((str(key), value))
        feasibility_evidence_df = pd.DataFrame(rows, columns=["metric", "value"])

    template_summary_df = pd.DataFrame(columns=["line", "candidate_pairs", "templates_built", "template_coverage_ratio", "avg_virtual_count", "max_virtual_count", "rejected_by_chain_limit", "direct_edge_count", "real_bridge_edge_count", "virtual_bridge_edge_count"])
    if isinstance(failure_diagnostics, dict):
        template_section = failure_diagnostics.get("template", {})
        rows = []
        for line in ["big_roll", "small_roll"]:
            rows.append(
                {
                    "line": line,
                    "candidate_pairs": template_section.get(f"build_debug.{line}.candidate_pairs", 0),
                    "templates_built": template_section.get(f"{line}.kept_templates", template_section.get(f"build_debug.{line}.kept_templates", 0)),
                    "template_coverage_ratio": template_section.get(f"{line}.coverage_ratio", 0.0),
                    "avg_virtual_count": template_section.get(f"build_debug.{line}.avg_virtual_count", 0.0),
                    "max_virtual_count": template_section.get(f"build_debug.{line}.max_virtual_count", 0),
                    "rejected_by_chain_limit": template_section.get(f"build_debug.{line}.rejected_by_chain_limit", 0),
                    "direct_edge_count": template_section.get(f"build_debug.{line}.direct_edge_count", 0),
                    "real_bridge_edge_count": template_section.get(f"build_debug.{line}.real_bridge_edge_count", 0),
                    "virtual_bridge_edge_count": template_section.get(f"build_debug.{line}.virtual_bridge_edge_count", 0),
                }
            )
        template_summary_df = pd.DataFrame(rows)

    master_allocation_summary_df = pd.DataFrame(columns=["line", "slot_no", "slot_order_count", "slot_tons", "order_count_over_cap", "slot_route_risk_score", "pair_gap_proxy", "span_risk", "degree_risk", "isolated_order_penalty", "selected_bridge_mix"])
    if raw_slots:
        master_allocation_summary_df = pd.DataFrame(
            [
                {
                    "line": r.get("line", ""),
                    "slot_no": r.get("slot_no", 0),
                    "slot_order_count": r.get("order_count", 0),
                    "slot_tons": r.get("slot_tons", 0),
                    "order_count_over_cap": r.get("order_count_over_cap", 0),
                    "slot_route_risk_score": r.get("slot_route_risk_score", 0),
                    "pair_gap_proxy": r.get("pair_gap_proxy", 0),
                    "span_risk": r.get("span_risk", 0),
                    "degree_risk": r.get("degree_risk", 0),
                    "isolated_order_penalty": r.get("isolated_order_penalty", 0),
                    "selected_bridge_mix": r.get("selected_bridge_mix", ""),
                }
                for r in raw_slots
            ]
        )

    unroutable_slots_df = pd.DataFrame(
        columns=[
            "line", "slot_no", "order_count", "template_coverage_ratio", "missing_pair_count",
            "line_slot_order_cap", "order_count_over_cap",
            "zero_in_orders", "zero_out_orders", "min_width", "max_width",
            "min_thickness", "max_thickness", "steel_group_count",
            "estimated_reverse_count", "estimated_virtual_blocks",
            "isolated_order_penalty", "period_reverse_count",
            "period_reverse_count_violation_count", "reverse_count_definition",
            "top_isolated_orders", "slot_route_risk_score", "status",
        ]
    )
    if raw_slots:
        unroutable_slots_df = pd.DataFrame([r for r in raw_slots if str(r.get("status", "")) == "UNROUTABLE_SLOT"])

    rule_snapshot = pd.DataFrame(
        [
            (spec.key.value, spec.zh_name, spec.en_name, spec.solver_active, spec.validation_active, spec.export_visible)
            for spec in RULE_REGISTRY.export_visible_specs()
        ],
        columns=["rule_key", "zh_name", "en_name", "solver_active", "validation_active", "export_visible"],
    )

    flow_df = pd.DataFrame(
        [
            ("总订单数(输入-规划粒度)", int(input_order_count)),
            ("真实订单数(输出)", int((~final_df["is_virtual"]).sum()) if "is_virtual" in final_df.columns else int(len(final_df))),
            ("虚拟板坯数(输出)", int(final_df["is_virtual"].sum()) if "is_virtual" in final_df.columns else 0),
            ("剔除订单数", int(len(dropped_df))),
        ],
        columns=["指标", "值"],
    )

    campaign_tons_df = pd.DataFrame(columns=["产线", "产线代码", "轧期序号", "轧期槽位号", "实物吨位", "虚拟吨位", "总吨位", "实物卷数", "虚拟卷数"])
    if not final_df.empty and {"line", "tons", "is_virtual"}.issubset(set(final_df.columns)):
        c = final_df.copy()
        c["实物吨位"] = c.apply(lambda r: float(r["tons"]) if not bool(r["is_virtual"]) else 0.0, axis=1)
        c["虚拟吨位"] = c.apply(lambda r: float(r["tons"]) if bool(r["is_virtual"]) else 0.0, axis=1)
        c["实物卷数"] = c["is_virtual"].map(lambda v: 0 if bool(v) else 1)
        c["虚拟卷数"] = c["is_virtual"].map(lambda v: 1 if bool(v) else 0)

        # ---- 关键修复：轧期槽位号必须来自 master_slot，不允许用 cumcount 或 fallback ----
        # 1. 优先使用 master_slot（构造性LNS路径的正确语义）
        if "master_slot" in c.columns:
            c["轧期槽位号_raw"] = pd.to_numeric(c["master_slot"], errors="coerce").fillna(0).astype(int)
        else:
            c["轧期槽位号_raw"] = 0

        # 2. 一致性自检：每个 (line, campaign_id) 分组内的 master_slot 必须唯一
        if {"line", "campaign_id", "轧期槽位号_raw"}.issubset(set(c.columns)) and not c.empty:
            grp_check = c.groupby(["line", "campaign_id"], dropna=False)["轧期槽位号_raw"].nunique()
            inconsistent = grp_check[grp_check > 1]
            if not inconsistent.empty:
                campaign_slot_consistency_ok = False
                for (line_val, camp_val), n_unique in inconsistent.items():
                    campaign_slot_fix_warnings.append(
                        f"INCONSISTENT_master_slot: line={line_val}, campaign={camp_val}, "
                        f"unique_slot_values={n_unique}"
                    )
                    campaign_slot_fixed_count += 1

        # 3. 对每个 (line, campaign_id) 分组，取众数的 master_slot 作为该 campaign 的槽位号
        if {"line", "campaign_id", "轧期槽位号_raw"}.issubset(set(c.columns)) and not c.empty:
            def _mode_or_first(series: pd.Series) -> int:
                """取众数，若无众数则取第一个值（两者都为空时返回0）"""
                s = series.dropna()
                if s.empty:
                    return 0
                vc = s.value_counts()
                if len(vc) > 0:
                    return int(vc.index[0])  # 众数（最常见的槽位号）
                return int(s.iloc[0])

            mode_slot = c.groupby(["line", "campaign_id"], dropna=False)["轧期槽位号_raw"].transform(_mode_or_first)
            c["轧期槽位号"] = mode_slot.astype(int)
        else:
            c["轧期槽位号"] = c["轧期槽位号_raw"]

        # 4. 移除临时列
        c = c.drop(columns=["轧期槽位号_raw"], errors="ignore")

        # 5. 聚合时：吨位/卷数用 sum，槽位号取众数（同一 campaign 内应该相同）
        # 注意：槽位号取 first（与 master_slot 模式一致），吨位/卷数取 sum
        campaign_tons_df = (
            c.groupby(["line", "campaign_id"], as_index=False, dropna=False).agg(
                实物吨位=("实物吨位", "sum"),
                虚拟吨位=("虚拟吨位", "sum"),
                实物卷数=("实物卷数", "sum"),
                虚拟卷数=("虚拟卷数", "sum"),
                轧期槽位号=("轧期槽位号", "first"),  # 槽位号取第一个（因为已经是众数）
            )
            .rename(columns={"line": "产线代码", "campaign_id": "轧期序号"})
        )
        campaign_tons_df["总吨位"] = campaign_tons_df["实物吨位"] + campaign_tons_df["虚拟吨位"]
        campaign_tons_df["产线"] = campaign_tons_df["产线代码"].map({"big_roll": "大辊线", "small_roll": "小辊线"}).fillna(campaign_tons_df["产线代码"])
        campaign_tons_df = campaign_tons_df[["产线", "产线代码", "轧期序号", "轧期槽位号", "实物吨位", "虚拟吨位", "总吨位", "实物卷数", "虚拟卷数"]]
        for col in ["实物吨位", "虚拟吨位", "总吨位"]:
            campaign_tons_df[col] = campaign_tons_df[col].round(1)

    show_cols = [
        "global_seq", "line_name", "line", "line_seq", "master_slot", "campaign_id", "campaign_id_hint", "campaign_seq", "campaign_real_seq",
        "order_id", "source_order_id", "parent_order_id", "lot_id", "is_virtual", "grade", "steel_group",
        "width", "thickness", "temp_min", "temp_max", "temp_mean", "tons", "backlog", "due_date", "due_bucket",
    ]
    schedule = final_df[[c for c in show_cols if c in final_df.columns]].copy()

    # 检测是否为 constructive_lns 路径
    is_constructive_lns = str(engine_used) == "constructive_lns" or str(main_path) == "constructive_lns"

    # ---- 关键修复：campaign_slot_no 必须严格使用 master_slot，不允许 fallback 到 cumcount ----
    # 修复点：对每个 (line, campaign_id) 分组，取众数的 master_slot 作为该 campaign 的槽位号
    if "master_slot" in schedule.columns:
        # 对每个 (line, campaign_id) 分组，取众数的 master_slot 作为该 campaign 的槽位号
        def _mode_or_first(series: pd.Series) -> int:
            """取众数，若无众数则取第一个值（两者都为空时返回0）"""
            s = series.dropna()
            if s.empty:
                return 0
            vc = s.value_counts()
            if len(vc) > 0:
                return int(vc.index[0])  # 众数（最常见的槽位号）
            return int(s.iloc[0])

        if {"line", "campaign_id", "master_slot"}.issubset(set(schedule.columns)) and not schedule.empty:
            master_slot_numeric = pd.to_numeric(schedule["master_slot"], errors="coerce").fillna(0).astype(int)
            mode_slot = schedule.groupby(["line", "campaign_id"], dropna=False)["master_slot"].transform(_mode_or_first)
            schedule["campaign_slot_no"] = pd.to_numeric(mode_slot, errors="coerce").fillna(0).astype(int)
        else:
            schedule["campaign_slot_no"] = pd.to_numeric(schedule["master_slot"], errors="coerce").fillna(0).astype(int)
    elif "campaign_no" in schedule.columns:
        # 旧路径：fallback 到 campaign_no
        schedule["campaign_slot_no"] = pd.to_numeric(schedule["campaign_no"], errors="coerce").fillna(0).astype(int)
    else:
        schedule["campaign_slot_no"] = 0

    # A. constructive_lns 路径：保留字符串型 campaign_id
    if is_constructive_lns:
        if "campaign_id" in schedule.columns:
            # 优先取 campaign_id，若为空则回退到 campaign_id_hint
            schedule["campaign_id"] = schedule["campaign_id"].fillna(
                schedule["campaign_id_hint"] if "campaign_id_hint" in schedule.columns else pd.NA
            ).fillna("")
        # 补充 diagnostics 标识
        if isinstance(em, dict):
            em["campaign_id_string_preserved"] = True
    # B. 旧路径：保持数值化兼容
    else:
        if "campaign_id" in schedule.columns:
            schedule["campaign_id"] = pd.to_numeric(schedule["campaign_id"], errors="coerce").fillna(0).astype(int)

    # 删除临时的 campaign_id_hint 列（不再需要导出）
    if "campaign_id_hint" in schedule.columns:
        schedule = schedule.drop(columns=["campaign_id_hint"])
    if "is_virtual" in schedule.columns:
        schedule["是否虚拟"] = schedule["is_virtual"].map(lambda x: "是" if bool(x) else "否")
        schedule = schedule.drop(columns=["is_virtual"])
    schedule = schedule.rename(
        columns={
            "global_seq": "全局序号",
            "line_name": "产线",
            "line": "产线代码",
            "line_seq": "产线内序号",
            "master_slot": "master_slot",  # 保留原始槽位号（用于调试）
            "campaign_id": "轧期编号",
            "campaign_slot_no": "轧期槽位号",
            "campaign_seq": "轧期内序号",
            "campaign_real_seq": "轧期内真实序号",
            "order_id": "物料号",
            "source_order_id": "来源订单号",
            "parent_order_id": "父订单号",
            "lot_id": "规划批号",
            "grade": "牌号",
            "steel_group": "钢种组",
            "width": "宽度",
            "thickness": "厚度",
            "temp_min": "温度下限",
            "temp_max": "温度上限",
            "temp_mean": "均热温度",
            "tons": "吨位",
            "backlog": "欠交量",
            "due_date": "交期",
            "due_bucket": "交期桶",
        }
    )
    # 排序和轧期内序号重算：按产线代码、轧期编号（或槽位号）、产线内序号
    if {"产线代码", "轧期编号", "产线内序号"}.issubset(set(schedule.columns)):
        # constructive_lns 路径下用字符串 campaign_id 排序，旧路径用数值排序
        schedule = schedule.sort_values(
            ["产线代码", "轧期编号", "产线内序号"],
            kind="mergesort",
            na_position="first",
        ).reset_index(drop=True)
        # 重新计算轧期内序号（同一 campaign 内的顺序）
        schedule["轧期内序号"] = schedule.groupby(["产线代码", "轧期编号"], dropna=False).cumcount() + 1
        schedule["轧期内真实序号"] = schedule["轧期内序号"]

    big_sheet = schedule[schedule["产线代码"] == "big_roll"].drop(columns=["产线代码"], errors="ignore").reset_index(drop=True) if "产线代码" in schedule.columns else schedule.copy()
    small_sheet = schedule[schedule["产线代码"] == "small_roll"].drop(columns=["产线代码"], errors="ignore").reset_index(drop=True) if "产线代码" in schedule.columns else schedule.iloc[0:0].copy()
    if candidate_backfill_mode:
        big_sheet = _build_candidate_schedule_sheet(candidate_schedule_df, "big_roll")
        small_sheet = _build_candidate_schedule_sheet(candidate_schedule_df, "small_roll")

    # ---- 轧期槽位号一致性自检（写大辊/小辊排程之前） ----
    # 对每个 (line, campaign_id) 分组，检查 campaign_slot_no 是否只有一个唯一值
    sheet_consistency_warnings: list[str] = []
    sheet_slot_fixed_count = 0
    sheet_slot_consistency_ok = True

    for sheet_name, sheet_df in [("大辊线", big_sheet), ("小辊线", small_sheet)]:
        if sheet_df.empty:
            continue
        if "轧期编号" not in sheet_df.columns or "轧期槽位号" not in sheet_df.columns:
            continue
        # 对每条产线的 sheet 检查槽位号一致性
        if "line" in schedule.columns:
            line_col = schedule["line"].iloc[0] if not schedule.empty else ""
        else:
            line_col = sheet_name
        for camp_id, grp in sheet_df.groupby("轧期编号", dropna=False):
            unique_slots = grp["轧期槽位号"].dropna().unique()
            if len(unique_slots) > 1:
                sheet_slot_consistency_ok = False
                sheet_slot_fixed_count += 1
                mode_slot = grp["轧期槽位号"].mode()
                fixed_slot = int(mode_slot.iloc[0]) if len(mode_slot) > 0 else int(grp["轧期槽位号"].iloc[0])
                sheet_consistency_warnings.append(
                    f"INCONSISTENT_CAMPAIGN_SLOT: sheet={sheet_name}, campaign={camp_id}, "
                    f"unique_slots={list(unique_slots)}, fixed_to={fixed_slot}"
                )
                # 修正：将该 campaign 内的所有行改为众数槽位号
                if sheet_name == "大辊线":
                    big_sheet.loc[big_sheet["轧期编号"] == camp_id, "轧期槽位号"] = fixed_slot
                elif sheet_name == "小辊线":
                    small_sheet.loc[small_sheet["轧期编号"] == camp_id, "轧期槽位号"] = fixed_slot

    # 合并两次一致性检查结果
    campaign_slot_consistency_ok = campaign_slot_consistency_ok and sheet_slot_consistency_ok
    campaign_slot_fixed_count += sheet_slot_fixed_count
    campaign_slot_fix_warnings.extend(sheet_consistency_warnings)

    if not campaign_slot_consistency_ok:
        print(
            f"[APS][WARNING] campaign_slot_no 一致性检查失败: "
            f"共 {campaign_slot_fixed_count} 个 campaign 的槽位号不一致，已修正。\n"
            + "\n".join(campaign_slot_fix_warnings[:5])  # 只打印前5条
        )
    elif campaign_slot_fixed_count == 0 and is_constructive_lns:
        # 仅在 constructive_lns 路径且无修正时提示确认
        print(
            f"[APS][campaign_slot确认] constructive_lns 路径: "
            f"campaign_slot_no 一致性检查通过, 所有 campaign 的槽位号均来自 master_slot"
        )

    rounds_out = rounds_df.copy()
    if "line" in rounds_out.columns:
        rounds_out["产线"] = rounds_out["line"].map({"big_roll": "大辊线", "small_roll": "小辊线"}).fillna(rounds_out["line"])

    dropped_reason_stats = pd.DataFrame(columns=["剔除原因", "数量", "占比"])
    dropped_out = dropped_df.copy()
    if not dropped_out.empty and "dominant_drop_reason" in dropped_out.columns:
        stats = dropped_out["dominant_drop_reason"].fillna("UNKNOWN").value_counts().reset_index()
        stats.columns = ["剔除原因", "数量"]
        stats["占比"] = (stats["数量"] / max(1, int(stats["数量"].sum()))).round(4)
        dropped_reason_stats = stats

    scheduled_orders = _num_int((~final_df["is_virtual"]).sum()) if "is_virtual" in final_df.columns else _num_int(len(final_df))
    unscheduled_orders = _num_int(dropped_out["order_id"].nunique()) if not dropped_out.empty and "order_id" in dropped_out.columns else _num_int(len(dropped_out))
    scheduled_tons = round(_num_float(final_df["tons"].sum()) if "tons" in final_df.columns else 0.0, 1)
    unscheduled_tons = round(_num_float(dropped_out["tons"].sum()) if "tons" in dropped_out.columns else 0.0, 1)
    selected_direct_edge_count = _num_int(fallback_diag.get("bridge_mix_summary", {}).get("selected_direct_edge_count", 0) if isinstance(fallback_diag, dict) else 0)
    selected_real_bridge_edge_count = _num_int(fallback_diag.get("bridge_mix_summary", {}).get("selected_real_bridge_edge_count", 0) if isinstance(fallback_diag, dict) else 0)
    selected_virtual_bridge_edge_count = _num_int(fallback_diag.get("bridge_mix_summary", {}).get("selected_virtual_bridge_edge_count", 0) if isinstance(fallback_diag, dict) else 0)
    if not final_df.empty and "line" in final_df.columns:
        real_mask = (~final_df["is_virtual"]) if "is_virtual" in final_df.columns else pd.Series([True] * len(final_df), index=final_df.index)
        big_roll_scheduled_orders = _num_int(((final_df["line"] == "big_roll") & real_mask).sum())
        small_roll_scheduled_orders = _num_int(((final_df["line"] == "small_roll") & real_mask).sum())
    else:
        big_roll_scheduled_orders = 0
        small_roll_scheduled_orders = 0
    if candidate_backfill_mode:
        candidate_assigned = candidate_schedule_df[
            (~candidate_schedule_df.get("drop_flag", pd.Series(dtype=bool)).fillna(False))
            & candidate_schedule_df.get("line", pd.Series(dtype=str)).astype(str).ne("")
        ].copy()
        scheduled_orders = int(len(candidate_assigned))
        scheduled_tons = round(float(candidate_assigned["tons"].sum()) if "tons" in candidate_assigned.columns else 0.0, 1)
        if not candidate_assigned.empty:
            big_roll_scheduled_orders = int((candidate_assigned.get("line", pd.Series(dtype=str)) == "big_roll").sum())
            small_roll_scheduled_orders = int((candidate_assigned.get("line", pd.Series(dtype=str)) == "small_roll").sum())
            edge_series = candidate_assigned.get("selected_edge_type", pd.Series(dtype=str)).astype(str)
            selected_direct_edge_count = int((edge_series == "DIRECT_EDGE").sum())
            selected_real_bridge_edge_count = int((edge_series == "REAL_BRIDGE_EDGE").sum())
            selected_virtual_bridge_edge_count = int((edge_series == "VIRTUAL_BRIDGE_EDGE").sum())
    max_big_roll_slot_order_count = _num_int(fallback_diag.get("big_roll_max_slot_order_count", joint_estimates.get("big_roll_max_slot_order_count", 0) if isinstance(joint_estimates, dict) else 0))
    max_small_roll_slot_order_count = _num_int(fallback_diag.get("small_roll_max_slot_order_count", joint_estimates.get("small_roll_max_slot_order_count", 0) if isinstance(joint_estimates, dict) else 0))

    run_summary_df = pd.DataFrame(
        [
            ("profile", str(fallback_diag.get("profile_name", "default"))),
            ("acceptance", result_acceptance_status),
            ("failure_mode", failure_mode),
            ("best_candidate_available", int(best_candidate_available)),
            ("best_candidate_type", best_candidate_type),
            ("best_candidate_objective", best_candidate_objective),
            ("best_candidate_search_status", best_candidate_search_status),
            ("best_candidate_routing_feasible", int(best_candidate_routing_feasible)),
            ("best_candidate_unroutable_slot_count", best_candidate_unroutable_slot_count),
            ("export_consistency_ok", int(export_consistency_ok)),
            ("routing_feasible", int(routing_feasible)),
            ("evidence_level", evidence_level),
            ("total_orders", int(input_order_count)),
            ("scheduled_orders", scheduled_orders),
            ("unscheduled_orders", unscheduled_orders),
            ("scheduled_tons", scheduled_tons),
            ("unscheduled_tons", unscheduled_tons),
            ("total_run_seconds", total_run_seconds),
            ("template_build_seconds", template_build_seconds),
            ("joint_master_seconds", joint_master_seconds),
            ("fallback_total_seconds", fallback_total_seconds),
        ],
        columns=["metric", "value"],
    )

    line_summary_rows = []
    slot_detail_rows = []
    # raw_slots already extracted from failure_diagnostics/engine_meta to support candidate backfill
    for line in ["big_roll", "small_roll"]:
        line_sched = final_df[final_df["line"] == line].copy() if (not final_df.empty and "line" in final_df.columns) else final_df.iloc[0:0].copy()
        real_orders = line_sched[(~line_sched["is_virtual"]) if "is_virtual" in line_sched.columns else pd.Series([True] * len(line_sched), index=line_sched.index)]
        slot_count = int(line_sched["campaign_id"].nunique()) if ("campaign_id" in line_sched.columns and not line_sched.empty) else 0
        avg_slot_order_count = float(real_orders.groupby("campaign_id").size().mean()) if ("campaign_id" in real_orders.columns and not real_orders.empty) else 0.0
        max_slot_order_count = int(real_orders.groupby("campaign_id").size().max()) if ("campaign_id" in real_orders.columns and not real_orders.empty) else 0
        unr_line = [r for r in raw_slots if str(r.get("line", "")) == line and str(r.get("status", "")) == "UNROUTABLE_SLOT"]
        edge_series = line_sched["selected_edge_type"].astype(str) if "selected_edge_type" in line_sched.columns else pd.Series(dtype=str)
        line_summary_rows.append(
            {
                "line": line,
                "scheduled_orders": int(len(real_orders)),
                "scheduled_tons": round(float(line_sched["tons"].sum()) if "tons" in line_sched.columns else 0.0, 1),
                "slot_count": slot_count,
                "avg_slot_order_count": round(avg_slot_order_count, 2),
                "max_slot_order_count": max_slot_order_count,
                "unroutable_slot_count": int(len(unr_line)),
                "direct_edge_count": int((edge_series == "DIRECT_EDGE").sum()) if not edge_series.empty else 0,
                "real_bridge_edge_count": int((edge_series == "REAL_BRIDGE_EDGE").sum()) if not edge_series.empty else 0,
                "virtual_bridge_edge_count": int((edge_series == "VIRTUAL_BRIDGE_EDGE").sum()) if not edge_series.empty else 0,
            }
        )
    line_summary_df = pd.DataFrame(
        line_summary_rows,
        columns=[
            "line", "scheduled_orders", "scheduled_tons", "slot_count", "avg_slot_order_count", "max_slot_order_count",
            "unroutable_slot_count", "direct_edge_count", "real_bridge_edge_count", "virtual_bridge_edge_count",
        ],
    )
    if final_df.empty and not candidate_line_summary_df.empty:
        line_summary_df = candidate_line_summary_df.rename(
            columns={
                "candidate_assigned_orders": "scheduled_orders",
                "candidate_assigned_tons": "scheduled_tons",
                "candidate_slot_count": "slot_count",
                "candidate_avg_slot_order_count": "avg_slot_order_count",
                "candidate_max_slot_order_count": "max_slot_order_count",
                "candidate_unroutable_slot_count": "unroutable_slot_count",
                "candidate_direct_edge_count": "direct_edge_count",
                "candidate_real_bridge_edge_count": "real_bridge_edge_count",
                "candidate_virtual_bridge_edge_count": "virtual_bridge_edge_count",
            }
        )[
            [
                "line", "scheduled_orders", "scheduled_tons", "slot_count", "avg_slot_order_count", "max_slot_order_count",
                "unroutable_slot_count", "direct_edge_count", "real_bridge_edge_count", "virtual_bridge_edge_count",
            ]
        ]
        line_summary_df["summary_mode"] = "CANDIDATE_ANALYSIS"
    else:
        line_summary_df["summary_mode"] = "OFFICIAL"

    slot_summary_df, unroutable_slots_df = _build_slot_frames(raw_slots, candidate_schedule_df)

    validation_summary = failure_diagnostics.get("validation_summary", {}) if isinstance(failure_diagnostics, dict) else {}
    hard_violation_total_official = 0
    for k in [
        "width_jump_violation_cnt",
        "thickness_violation_cnt",
        "temp_conflict_cnt",
        "non_pc_direct_switch_cnt",
        "direct_reverse_step_violation_count",
        "virtual_attach_reverse_violation_count",
        "period_reverse_count_violation_count",
        "bridge_count_violation_count",
        "invalid_virtual_spec_count",
    ]:
        try:
            hard_violation_total_official += int(validation_summary.get(k, 0) or 0)
        except Exception:
            continue
    candidate_violation_summary_df = _build_candidate_violation_summary(
        candidate_schedule_df,
        raw_slots,
        validation_summary if isinstance(validation_summary, dict) else {},
        bool(best_candidate_routing_feasible),
    )
    violation_summary_df = pd.DataFrame(
        [
            ("hard_constraint_violation_count", int(hard_violation_total_official)),
            ("direct_reverse_step_violation_count", int(validation_summary.get("direct_reverse_step_violation_count", 0))),
            ("virtual_attach_reverse_violation_count", int(validation_summary.get("virtual_attach_reverse_violation_count", 0))),
            ("period_reverse_count_violation_count", int(validation_summary.get("period_reverse_count_violation_count", 0))),
            ("bridge_count_violation_count", int(validation_summary.get("bridge_count_violation_count", 0))),
            ("invalid_virtual_spec_count", int(validation_summary.get("invalid_virtual_spec_count", 0))),
            ("unroutable_slot_count", unroutable_slot_count),
            ("hard_cap_not_enforced", bool(fallback_diag.get("hard_cap_not_enforced", False) or joint_estimates.get("hard_cap_not_enforced", False))),
            ("isolation_risk_not_effective", any(bool(r.get("isolation_risk_not_effective", False)) for r in raw_slots)),
        ],
        columns=["metric", "value"],
    )
    if final_df.empty and not candidate_violation_summary_df.empty:
        violation_summary_df = candidate_violation_summary_df.copy()

    reason_hist = dropped_reason_stats.set_index("剔除原因")["数量"].to_dict() if not dropped_reason_stats.empty else {}
    _dropped_unique_count = int(dropped_out["order_id"].nunique()) if not dropped_out.empty and "order_id" in dropped_out.columns else int(len(dropped_out))
    _dropped_row_count = int(len(dropped_out))
    _duplicated_dropped_rows = max(0, _dropped_row_count - _dropped_unique_count)
    unscheduled_summary_df = pd.DataFrame(
        [
            ("unscheduled_order_count", int(dropped_out["order_id"].nunique())) if not dropped_out.empty and "order_id" in dropped_out.columns else ("unscheduled_order_count", int(len(dropped_out))),
            ("unscheduled_tons", round(float(dropped_out["tons"].sum()) if "tons" in dropped_out.columns else 0.0, 1)),
            ("dropped_row_count", _dropped_row_count),
            ("duplicated_dropped_rows_count", _duplicated_dropped_rows),
            ("drop_ratio_unique", round(float(_dropped_unique_count / max(1, input_order_count)), 4)),
            ("dropped_reason_histogram", str(reason_hist)),
            ("globally_isolated_order_count", int(dropped_out["globally_isolated"].fillna(False).sum()) if "globally_isolated" in dropped_out.columns else 0),
            ("low_priority_drop_count", int((dropped_out.get("dominant_drop_reason", pd.Series(dtype=str)) == "LOW_PRIORITY_DROP").sum())),
            ("no_feasible_line_count", int((dropped_out.get("dominant_drop_reason", pd.Series(dtype=str)) == "NO_FEASIBLE_LINE").sum())),
            ("slot_routing_risk_too_high_count", int((dropped_out.get("dominant_drop_reason", pd.Series(dtype=str)) == "SLOT_ROUTING_RISK_TOO_HIGH").sum())),
        ],
        columns=["metric", "value"],
    )

    bridge_summary_df = pd.DataFrame(
        [
            ("selected_direct_edge_count", selected_direct_edge_count),
            ("selected_real_bridge_edge_count", selected_real_bridge_edge_count),
            ("selected_virtual_bridge_edge_count", selected_virtual_bridge_edge_count),
            ("direct_edge_ratio", float(fallback_diag.get("bridge_mix_summary", {}).get("direct_edge_ratio", 0.0) if isinstance(fallback_diag, dict) else 0.0)),
            ("real_bridge_ratio", float(fallback_diag.get("bridge_mix_summary", {}).get("real_bridge_ratio", 0.0) if isinstance(fallback_diag, dict) else 0.0)),
            ("virtual_bridge_ratio", float(fallback_diag.get("bridge_mix_summary", {}).get("virtual_bridge_ratio", 0.0) if isinstance(fallback_diag, dict) else 0.0)),
            ("max_bridge_count_used", int(validation_summary.get("max_bridge_count_used", 0))),
            ("avg_virtual_count", float(validation_summary.get("averageTemplateVirtualCount", joint_estimates.get("averageTemplateVirtualCount", 0.0) if isinstance(joint_estimates, dict) else 0.0) or 0.0)),
            # Route C: Bridge edge leak detection for constructive_lns path
            ("bridge_edge_leak_detected", bridge_edge_leak_detected),
            ("constructive_edge_policy", constructive_edge_policy if is_constructive_lns else ""),
        ],
        columns=["metric", "value"],
    )

    progress_metrics_df = pd.DataFrame(
        [[
            str(fallback_diag.get("profile_name", "default")),
            result_acceptance_status,
            failure_mode,
            int(best_candidate_available),
            best_candidate_type,
            int(export_consistency_ok),
            int(routing_feasible),
            scheduled_orders,
            unscheduled_orders,
            _num_int(dropped_out["order_id"].nunique()) if not dropped_out.empty and "order_id" in dropped_out.columns else _num_int(len(dropped_out)),
            unscheduled_tons,
            big_roll_scheduled_orders,
            small_roll_scheduled_orders,
            unroutable_slot_count,
            max_big_roll_slot_order_count,
            max_small_roll_slot_order_count,
            selected_direct_edge_count,
            selected_real_bridge_edge_count,
            selected_virtual_bridge_edge_count,
            int(validation_summary.get("direct_reverse_step_violation_count", 0)),
            int(validation_summary.get("virtual_attach_reverse_violation_count", 0)),
            int(validation_summary.get("period_reverse_count_violation_count", 0)),
            int(validation_summary.get("invalid_virtual_spec_count", 0)),
            (pd.NA if candidate_backfill_mode else int(hard_violation_total_official)),
            template_build_seconds,
            joint_master_seconds,
            fallback_total_seconds,
            total_run_seconds,
        ]],
        columns=[
            "profile", "acceptance", "failure_mode", "best_candidate_available", "best_candidate_type",
            "export_consistency_ok",
            "routing_feasible", "scheduled_orders", "unscheduled_orders", "dropped_order_count", "dropped_tons",
            "big_roll_scheduled_orders", "small_roll_scheduled_orders", "unroutable_slot_count",
            "max_big_roll_slot_order_count", "max_small_roll_slot_order_count",
            "selected_direct_edge_count", "selected_real_bridge_edge_count", "selected_virtual_bridge_edge_count",
            "direct_reverse_step_violation_count", "virtual_attach_reverse_violation_count",
            "period_reverse_count_violation_count", "invalid_virtual_spec_count",
            "hard_constraint_violation_count",
            "template_build_seconds", "joint_master_seconds", "fallback_total_seconds", "total_run_seconds",
        ],
    )

    drop_and_bridge_details_df = pd.DataFrame(columns=[
        "order_id", "dropped", "dominant_drop_reason", "secondary_reasons", "globally_isolated",
        "candidate_lines", "risk_summary", "selected_edge_type", "bridge_count", "bridge_type",
        "virtual_widths_used", "virtual_thicknesses_used",
    ])
    detail_rows = []
    if not final_df.empty:
        for _, row in final_df.iterrows():
            is_virtual = bool(row.get("is_virtual", False))
            detail_rows.append(
                {
                    "order_id": row.get("order_id", ""),
                    "dropped": False,
                    "dominant_drop_reason": "",
                    "secondary_reasons": "",
                    "globally_isolated": False,
                    "candidate_lines": "",
                    "risk_summary": "",
                    "selected_edge_type": row.get("selected_edge_type", ""),
                    "bridge_count": int(row.get("bridge_count", 0) or 0),
                    "bridge_type": row.get("selected_edge_type", ""),
                    "virtual_widths_used": row.get("width", "") if is_virtual else "",
                    "virtual_thicknesses_used": row.get("thickness", "") if is_virtual else "",
                }
            )
    if not dropped_out.empty:
        for _, row in dropped_out.iterrows():
            detail_rows.append(
                {
                    "order_id": row.get("order_id", ""),
                    "dropped": True,
                    "dominant_drop_reason": row.get("dominant_drop_reason", row.get("drop_reason", "")),
                    "secondary_reasons": row.get("secondary_reasons", ""),
                    "globally_isolated": bool(row.get("globally_isolated", False)),
                    "candidate_lines": row.get("candidate_lines", ""),
                    "risk_summary": row.get("risk_summary", ""),
                    "selected_edge_type": "",
                    "bridge_count": 0,
                    "bridge_type": "",
                    "virtual_widths_used": "",
                    "virtual_thicknesses_used": "",
                }
            )
    if detail_rows:
        drop_and_bridge_details_df = pd.DataFrame(detail_rows)

    drop_and_bridge_summary_df = pd.DataFrame(
        [
            ("dropped_order_count", int(len(dropped_out))),
            ("selected_direct_edge_count", int(fallback_diag.get("bridge_mix_summary", {}).get("selected_direct_edge_count", 0) if isinstance(fallback_diag, dict) else 0)),
            ("selected_real_bridge_edge_count", int(fallback_diag.get("bridge_mix_summary", {}).get("selected_real_bridge_edge_count", 0) if isinstance(fallback_diag, dict) else 0)),
            ("selected_virtual_bridge_edge_count", int(fallback_diag.get("bridge_mix_summary", {}).get("selected_virtual_bridge_edge_count", 0) if isinstance(fallback_diag, dict) else 0)),
            ("max_bridge_count_used", int((failure_diagnostics.get("validation_summary", {}) if isinstance(failure_diagnostics, dict) else {}).get("max_bridge_count_used", 0))),
            ("bridge_count_violation_count", int((failure_diagnostics.get("validation_summary", {}) if isinstance(failure_diagnostics, dict) else {}).get("bridge_count_violation_count", 0))),
            ("invalid_virtual_spec_count", int((failure_diagnostics.get("validation_summary", {}) if isinstance(failure_diagnostics, dict) else {}).get("invalid_virtual_spec_count", 0))),
            ("dropped_reason_histogram", str((dropped_reason_stats.set_index("剔除原因")["数量"].to_dict() if not dropped_reason_stats.empty else {}))),
        ],
        columns=["metric", "value"],
    )
    candidate_big_roll_out = candidate_big_roll_df.copy() if isinstance(candidate_big_roll_df, pd.DataFrame) else pd.DataFrame()
    candidate_small_roll_out = candidate_small_roll_df.copy() if isinstance(candidate_small_roll_df, pd.DataFrame) else pd.DataFrame()
    if candidate_big_roll_out.empty:
        candidate_big_roll_out = candidate_schedule_df[candidate_schedule_df.get("line", pd.Series(dtype=str)) == "big_roll"].copy() if not candidate_schedule_df.empty else pd.DataFrame()
    if candidate_small_roll_out.empty:
        candidate_small_roll_out = candidate_schedule_df[candidate_schedule_df.get("line", pd.Series(dtype=str)) == "small_roll"].copy() if not candidate_schedule_df.empty else pd.DataFrame()

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Main schedule sheets: keep official naming, but translate candidate drafts when backfilled.
        if candidate_backfill_mode:
            big_sheet_out = big_sheet.rename(
                columns={
                    "order_id": "物料号",
                    "slot_no": "槽位号",
                    "candidate_position": "槽位内序号",
                    "candidate_slot_member_index": "槽位内成员序号",
                    "tons": "吨位",
                    "width": "宽度",
                    "thickness": "厚度",
                    "steel_group": "钢种组",
                    "line_capability": "产线能力",
                    "candidate_status": "候选状态",
                    "slot_unroutable_flag": "槽不可路由标记",
                    "slot_route_risk_score": "槽路由风险",
                    "selected_edge_type": "选中边类型",
                    "analysis_only": "仅分析",
                    "official_usable": "可下发",
                    "result_usage": "结果用途",
                }
            )
            if "候选状态" in big_sheet_out.columns:
                big_sheet_out["候选状态"] = big_sheet_out["候选状态"].map(lambda v: _zh_enum(v, _CANDIDATE_STATUS_ZH))
            if "选中边类型" in big_sheet_out.columns:
                big_sheet_out["选中边类型"] = big_sheet_out["选中边类型"].map(lambda v: _zh_enum(v, _EDGE_TYPE_ZH))
            small_sheet_out = small_sheet.rename(
                columns={
                    "order_id": "物料号",
                    "slot_no": "槽位号",
                    "candidate_position": "槽位内序号",
                    "candidate_slot_member_index": "槽位内成员序号",
                    "tons": "吨位",
                    "width": "宽度",
                    "thickness": "厚度",
                    "steel_group": "钢种组",
                    "line_capability": "产线能力",
                    "candidate_status": "候选状态",
                    "slot_unroutable_flag": "槽不可路由标记",
                    "slot_route_risk_score": "槽路由风险",
                    "selected_edge_type": "选中边类型",
                    "analysis_only": "仅分析",
                    "official_usable": "可下发",
                    "result_usage": "结果用途",
                }
            )
            if "候选状态" in small_sheet_out.columns:
                small_sheet_out["候选状态"] = small_sheet_out["候选状态"].map(lambda v: _zh_enum(v, _CANDIDATE_STATUS_ZH))
            if "选中边类型" in small_sheet_out.columns:
                small_sheet_out["选中边类型"] = small_sheet_out["选中边类型"].map(lambda v: _zh_enum(v, _EDGE_TYPE_ZH))
        else:
            big_sheet_out = big_sheet
            small_sheet_out = small_sheet
        big_sheet_out.to_excel(writer, sheet_name="大辊线排程", index=False)
        small_sheet_out.to_excel(writer, sheet_name="小辊线排程", index=False)
        campaign_tons_df.to_excel(writer, sheet_name="轧期吨位统计", index=False)
        rounds_out.rename(columns={"round": "轮次", "line": "产线代码"}).to_excel(writer, sheet_name="轮次统计", index=False)
        summary_df.to_excel(writer, sheet_name="总览指标", index=False)
        flow_df.to_excel(writer, sheet_name="订单去向", index=False)
        dropped_out.to_excel(writer, sheet_name="剔除订单", index=False)
        dropped_reason_stats.to_excel(writer, sheet_name="剔除原因统计", index=False)
        runtime_report.to_excel(writer, sheet_name="标准化运行报表", index=False)
        diagnostics_report.rename(columns={"section": "分区", "metric": "指标", "value": "值"}).to_excel(writer, sheet_name="诊断摘要", index=False)
        equivalence_report.to_excel(writer, sheet_name="等价性检查", index=False)
        diagnostics_snapshot.rename(columns={"key": "字段", "value": "值"}).to_excel(writer, sheet_name="求解路径快照", index=False)
        rule_snapshot.rename(
            columns={
                "rule_key": "规则键",
                "zh_name": "中文名",
                "en_name": "英文名",
                "solver_active": "求解启用",
                "validation_active": "校验启用",
                "export_visible": "导出可见",
            }
        ).to_excel(writer, sheet_name="规则快照", index=False)
        _kv_to_zh(precheck_summary_df).to_excel(writer, sheet_name="数据预检汇总", index=False)
        _kv_to_zh(feasibility_evidence_df).to_excel(writer, sheet_name="不可行证据", index=False)
        template_summary_df.assign(产线=template_summary_df.get("line", "").map(_zh_line)).rename(
            columns={
                "line": "产线代码",
                "candidate_pairs": "候选pair数",
                "templates_built": "模板数",
                "template_coverage_ratio": "模板覆盖率",
                "avg_virtual_count": "平均虚拟段数",
                "max_virtual_count": "最大虚拟段数",
                "rejected_by_chain_limit": "链长拒绝数",
                "direct_edge_count": "直接边数",
                "real_bridge_edge_count": "实物桥接边数",
                "virtual_bridge_edge_count": "虚拟桥接边数",
            }
        )[[
            "产线", "产线代码", "候选pair数", "模板数", "模板覆盖率", "平均虚拟段数", "最大虚拟段数",
            "链长拒绝数", "直接边数", "实物桥接边数", "虚拟桥接边数",
        ]].to_excel(writer, sheet_name="模板汇总", index=False)
        master_allocation_summary_df.assign(产线=master_allocation_summary_df.get("line", "").map(_zh_line)).rename(
            columns={
                "line": "产线代码",
                "slot_no": "槽位号",
                "slot_order_count": "槽内订单数",
                "slot_tons": "槽内吨位",
                "order_count_over_cap": "超cap订单数",
                "slot_route_risk_score": "槽路由风险",
                "pair_gap_proxy": "pair缺口风险",
                "span_risk": "跨度风险",
                "degree_risk": "低度数风险",
                "isolated_order_penalty": "孤点惩罚",
                "selected_bridge_mix": "桥接mix",
            }
        )[[
            "产线", "产线代码", "槽位号", "槽内订单数", "槽内吨位", "超cap订单数",
            "槽路由风险", "pair缺口风险", "跨度风险", "低度数风险", "孤点惩罚", "桥接mix",
        ]].to_excel(writer, sheet_name="主模型分配汇总", index=False)
        unroutable_slots_df.assign(产线=unroutable_slots_df.get("line", "").map(_zh_line)).rename(
            columns={
                "line": "产线代码",
                "slot_no": "槽位号",
                "order_count": "槽内订单数",
                "template_coverage_ratio": "模板覆盖率",
                "missing_template_edge_count": "缺失模板边数",
                "zero_in_orders": "零入度订单数",
                "zero_out_orders": "零出度订单数",
                "width_span": "宽度跨度",
                "thickness_span": "厚度跨度",
                "steel_group_count": "钢种组数量",
                "slot_route_risk_score": "槽路由风险",
                "dominant_unroutable_reason": "主不可路由原因",
                "top_isolated_orders": "孤点订单Top",
            }
        )[[
            "产线", "产线代码", "槽位号", "槽内订单数", "模板覆盖率", "缺失模板边数",
            "零入度订单数", "零出度订单数", "宽度跨度", "厚度跨度", "钢种组数量",
            "槽路由风险", "主不可路由原因", "孤点订单Top",
        ]].to_excel(writer, sheet_name="不可路由槽位", index=False)
        _kv_to_zh(drop_and_bridge_summary_df).to_excel(writer, sheet_name="剔除与桥接汇总", index=False)
        run_summary_zh = _kv_to_zh(run_summary_df)
        if "指标" in run_summary_zh.columns and "值" in run_summary_zh.columns:
            run_summary_zh["指标"] = run_summary_zh["指标"].map(
                {
                    "profile": "Profile",
                    "acceptance": "结果类型",
                    "failure_mode": "失败模式",
                    "best_candidate_available": "是否有最佳候选",
                    "best_candidate_type": "最佳候选类型",
                    "best_candidate_objective": "最佳候选目标值",
                    "best_candidate_search_status": "最佳候选状态",
                    "best_candidate_routing_feasible": "最佳候选路由可行",
                    "best_candidate_unroutable_slot_count": "最佳候选不可路由槽位数",
                    "export_consistency_ok": "导出口径一致",
                    "routing_feasible": "路由可行",
                    "evidence_level": "证据等级",
                    "total_orders": "总订单数",
                    "scheduled_orders": "已排订单数",
                    "unscheduled_orders": "未排订单数",
                    "scheduled_tons": "已排吨位",
                    "unscheduled_tons": "未排吨位",
                    "total_run_seconds": "总耗时(s)",
                    "template_build_seconds": "模板耗时(s)",
                    "joint_master_seconds": "主模型耗时(s)",
                    "fallback_total_seconds": "fallback耗时(s)",
                }
            ).fillna(run_summary_zh["指标"])
            def _zh_run_value(k: str, v: object) -> object:
                if k in {"结果类型", "acceptance"}:
                    return _zh_enum(v, _ACCEPTANCE_ZH)
                if k in {"失败模式", "failure_mode"}:
                    return _zh_enum(v, _FAILURE_MODE_ZH)
                return v
            run_summary_zh["值"] = [
                _zh_run_value(str(k), v) for k, v in zip(run_summary_zh["指标"].tolist(), run_summary_zh["值"].tolist())
            ]
        run_summary_zh.to_excel(writer, sheet_name="运行汇总", index=False)
        line_summary_df.assign(产线=line_summary_df.get("line", "").map(_zh_line)).rename(
            columns={
                "line": "产线代码",
                "scheduled_orders": "已排订单数",
                "scheduled_tons": "已排吨位",
                "slot_count": "槽位数",
                "avg_slot_order_count": "平均槽位订单数",
                "max_slot_order_count": "最大槽位订单数",
                "unroutable_slot_count": "不可路由槽位数",
                "direct_edge_count": "直接边数",
                "real_bridge_edge_count": "实物桥接边数",
                "virtual_bridge_edge_count": "虚拟桥接边数",
                "summary_mode": "统计口径",
            }
        ).assign(**{"统计口径": line_summary_df.get("summary_mode", "").map(lambda v: _zh_enum(v, _SUMMARY_MODE_ZH))})[[
            "产线", "产线代码", "已排订单数", "已排吨位", "槽位数", "平均槽位订单数", "最大槽位订单数",
            "不可路由槽位数", "直接边数", "实物桥接边数", "虚拟桥接边数", "统计口径",
        ]].to_excel(writer, sheet_name="产线汇总", index=False)
        violation_out = _kv_to_zh(violation_summary_df) if list(violation_summary_df.columns) == ["metric", "value"] else violation_summary_df.copy()
        if list(violation_out.columns) == ["metric", "official_value", "candidate_value", "violation_count_source", "violation_count_confidence"]:
            violation_out = violation_out.rename(
                columns={
                    "metric": "指标",
                    "official_value": "正式值",
                    "candidate_value": "候选值",
                    "violation_count_source": "统计来源",
                    "violation_count_confidence": "置信度",
                }
            )
            violation_out["统计来源"] = violation_out["统计来源"].map(lambda v: _zh_enum(v, _VIOLATION_SRC_ZH))
            violation_out["置信度"] = violation_out["置信度"].map(lambda v: _zh_enum(v, _CONFIDENCE_ZH))
        else:
            violation_out = _kv_to_zh(violation_out).rename(columns={"指标": "指标", "值": "值"})
            if "指标" in violation_out.columns:
                violation_out["指标"] = violation_out["指标"].map(
                    {
                        "hard_constraint_violation_count": "硬约束违例总数",
                        "direct_reverse_step_violation_count": "直接逆宽违例数(>20)",
                        "virtual_attach_reverse_violation_count": "虚拟接入逆宽违例数(>250)",
                        "period_reverse_count_violation_count": "辊期逆宽次数违例数(>5)",
                        "bridge_count_violation_count": "桥链长度违例数(>5)",
                        "invalid_virtual_spec_count": "虚拟规格违例数",
                        "unroutable_slot_count": "不可路由槽位数",
                        "hard_cap_not_enforced": "hard cap 未生效",
                        "isolation_risk_not_effective": "孤点风险未生效",
                    }
                ).fillna(violation_out["指标"])
        violation_out.to_excel(writer, sheet_name="违规汇总", index=False)
        _kv_to_zh(unscheduled_summary_df).to_excel(writer, sheet_name="未排汇总", index=False)
        _kv_to_zh(bridge_summary_df).to_excel(writer, sheet_name="桥接汇总", index=False)
        slot_summary_df.assign(产线=slot_summary_df.get("line", "").map(_zh_line)).rename(
            columns={
                "line": "产线代码",
                "slot_no": "槽位号",
                "slot_order_count": "槽内订单数",
                "slot_tons": "槽内吨位",
                "order_count_over_cap": "超cap订单数",
                "slot_route_risk_score": "槽路由风险",
                "pair_gap_proxy": "pair缺口风险",
                "span_risk": "跨度风险",
                "degree_risk": "低度数风险",
                "isolated_order_penalty": "孤点惩罚",
                "dominant_unroutable_reason": "主不可路由原因",
            }
        )[[
            "产线", "产线代码", "槽位号", "槽内订单数", "槽内吨位", "超cap订单数",
            "槽路由风险", "pair缺口风险", "跨度风险", "低度数风险", "孤点惩罚", "主不可路由原因",
        ]].to_excel(writer, sheet_name="槽位汇总", index=False)
        drop_and_bridge_details_df.rename(
            columns={
                "order_id": "物料号",
                "dropped": "是否剔除",
                "dominant_drop_reason": "主剔除原因",
                "secondary_reasons": "次要原因",
                "globally_isolated": "全局孤点",
                "candidate_lines": "候选产线",
                "risk_summary": "风险摘要",
                "selected_edge_type": "选中边类型",
                "bridge_count": "桥段数",
                "bridge_type": "桥接类型",
                "virtual_widths_used": "虚拟宽度使用",
                "virtual_thicknesses_used": "虚拟厚度使用",
            }
        ).to_excel(writer, sheet_name="剔除与桥接明细", index=False)
        progress_out = progress_metrics_df.copy()
        progress_out = progress_out.rename(
            columns={
                "profile": "Profile",
                "acceptance": "结果类型",
                "failure_mode": "失败模式",
                "best_candidate_available": "是否有最佳候选",
                "best_candidate_type": "最佳候选类型",
                "export_consistency_ok": "导出口径一致",
                "routing_feasible": "路由可行",
                "scheduled_orders": "已排订单数",
                "unscheduled_orders": "未排订单数",
                "dropped_order_count": "剔除订单数",
                "dropped_tons": "剔除吨位",
                "big_roll_scheduled_orders": "大辊已排订单数",
                "small_roll_scheduled_orders": "小辊已排订单数",
                "unroutable_slot_count": "不可路由槽位数",
                "max_big_roll_slot_order_count": "大辊最大槽订单数",
                "max_small_roll_slot_order_count": "小辊最大槽订单数",
                "selected_direct_edge_count": "直接边数(选中)",
                "selected_real_bridge_edge_count": "实物桥接边数(选中)",
                "selected_virtual_bridge_edge_count": "虚拟桥接边数(选中)",
                "direct_reverse_step_violation_count": "直接逆宽违例数(>20)",
                "virtual_attach_reverse_violation_count": "虚拟接入逆宽违例数(>250)",
                "period_reverse_count_violation_count": "辊期逆宽次数违例数(>5)",
                "invalid_virtual_spec_count": "虚拟规格违例数",
                "hard_constraint_violation_count": "硬约束违例总数",
                "template_build_seconds": "模板耗时(s)",
                "joint_master_seconds": "主模型耗时(s)",
                "fallback_total_seconds": "fallback耗时(s)",
                "total_run_seconds": "总耗时(s)",
            }
        )
        if "结果类型" in progress_out.columns:
            progress_out["结果类型"] = progress_out["结果类型"].map(lambda v: _zh_enum(v, _ACCEPTANCE_ZH))
        if "失败模式" in progress_out.columns:
            progress_out["失败模式"] = progress_out["失败模式"].map(lambda v: _zh_enum(v, _FAILURE_MODE_ZH))
        progress_out.to_excel(writer, sheet_name="进展指标", index=False)
        if best_candidate_available:
            candidate_big_roll_out.to_excel(writer, sheet_name="大辊线候选排程", index=False)
            candidate_small_roll_out.to_excel(writer, sheet_name="小辊线候选排程", index=False)
            candidate_line_summary_df.to_excel(writer, sheet_name="产线汇总_候选", index=False)
            candidate_violation_summary_df.to_excel(writer, sheet_name="违规汇总_候选", index=False)
        _autosize_excel(writer)

    final_df.to_csv(out_path.with_suffix(".csv"), index=False, encoding="utf-8-sig")
    elapsed = (perf_counter() - t_start) if t_start is not None else None
    if elapsed is None:
        print(f"[APS] 排程完成, 引擎={engine_used}, 输出={out_path}, result_usage={result_usage}, official_exported={official_exported}, analysis_exported={analysis_exported}")
    else:
        print(f"[APS] 排程完成, 引擎={engine_used}, 总耗时={elapsed:.1f}s, 输出={out_path}, result_usage={result_usage}, official_exported={official_exported}, analysis_exported={analysis_exported}")
    if best_candidate_available:
        print(
            f"[APS][导出一致性] acceptance={result_acceptance_status}, failure_mode={failure_mode}, "
            f"result_usage={result_usage}, best_candidate_available={int(best_candidate_available)}, "
            f"best_candidate_type={best_candidate_type}, best_candidate_search_status={best_candidate_search_status}, "
            f"export_consistency_ok={export_consistency_ok}"
        )
    if str(result_usage) == "ANALYSIS_ONLY" and best_candidate_available:
        print(
            f"[APS][候选导出] candidate_schedule_exported={not candidate_schedule_df.empty}, "
            f"candidate_line_summary_exported={not candidate_line_summary_df.empty}, "
            f"candidate_violation_summary_exported={not candidate_violation_summary_df.empty}, "
            "ANALYSIS_ONLY, NOT_ROUTING_FEASIBLE, 不可下发"
        )

    # Route B: Detect and warn about virtual bridge edge leakage in disabled mode
    if bridge_expansion_mode == "disabled" and virtual_bridge_enabled is False:
        # Route C (direct_only): Check if any bridge edge appears in the result (leak detection)
        virtual_edge_leak_count = selected_virtual_bridge_edge_count
        real_edge_leak_count = selected_real_bridge_edge_count
        is_direct_only = not virtual_bridge_enabled and not real_bridge_enabled

        if is_direct_only:
            # Route C: both virtual and real bridge edges must be 0
            if real_edge_leak_count > 0 or virtual_edge_leak_count > 0:
                print(
                    f"[APS][WARNING] 路线C(direct_only)桥接泄漏检测: "
                    f"allow_virtual={virtual_bridge_enabled}, allow_real={real_bridge_enabled}, "
                    f"但结果中包含 {real_edge_leak_count} 条 REAL_BRIDGE_EDGE, "
                    f"{virtual_edge_leak_count} 条 VIRTUAL_BRIDGE_EDGE！"
                    f"这是不期望的行为，结果违反了 direct_only 语义。"
                    f"请检查 constructive_sequence_builder / local_inserter_cp_sat 中的 allow_* 设置。"
                )
            else:
                print(
                    f"[APS][路线C确认] constructive_edge_policy=direct_only, "
                    f"结果中无桥接边泄漏, direct_only 语义正确"
                )
        elif virtual_edge_leak_count > 0:
            print(
                f"[APS][WARNING] 路线B桥接泄漏检测: bridge_expansion_mode=disabled, "
                f"但结果中包含 {virtual_edge_leak_count} 条 VIRTUAL_BRIDGE_EDGE 边！"
                f"这是不期望的行为，结果可能不符合禁用展开的语义。"
                f"请检查 constructive_lns_master 中的 allow_virtual_bridge_edge_in_constructive 设置。"
            )
        else:
            print(
                f"[APS][路线B确认] bridge_expansion_mode=disabled, "
                f"虚拟桥接边启用=false, 结果中无 VIRTUAL_BRIDGE_EDGE 泄漏, 路线B语义正确"
            )

    return final_df, rounds_df
